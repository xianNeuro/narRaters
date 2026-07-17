#!/usr/bin/env python3
"""
Web-based interface to visually view subjects' raw recall texts, corrected texts,
parsed texts, and recall ratings.

Run with: python web-interface.py
Then open browser to: http://127.0.0.1:5000/pipeline-config
"""

# Defer pandas import (saves ~15s startup). Module __getattr__ does not run for `pd` inside
# this file's functions (LOAD_GLOBAL only sees the module dict), so we use a tiny proxy that
# replaces itself with the real pandas module on first attribute access.
import os
import sys
import re
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_file, redirect, session, g
import json
import shutil
import hashlib
import secrets
import threading
from functools import wraps


class _LazyPandas:
    def __getattr__(self, name):
        import pandas as _pd
        globals()['pd'] = _pd
        return getattr(_pd, name)


pd = _LazyPandas()

# Package root: bundled templates, scripts, static (pip wheel or source checkout).
SCRIPT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = SCRIPT_DIR.parent
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

# The ``narraters`` package (cli, runtime_install, paths) lives under ``src/`` in a
# source checkout. Put that on sys.path so ``import narraters.*`` resolves to the
# bundled source even when the installed/editable metadata is stale or points at a
# moved/missing directory (otherwise step execution fails with
# "ModuleNotFoundError: No module named 'narraters.runtime_install'").
_SRC_DIR = PACKAGE_ROOT / "src"
if (_SRC_DIR / "narraters").is_dir() and str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

from helpers.software_paths import resolve_runtime_project_root

# Workspace root: where data/ and output/ live (may differ after ``pip install``).
WORKSPACE_ROOT = resolve_runtime_project_root(script_dir=SCRIPT_DIR)
PROJECT_ROOT = WORKSPACE_ROOT
SCRIPTS_DIR = PACKAGE_ROOT / "scripts"

from helpers.anthropic_ids import DEFAULT_ANTHROPIC_RECALL_MATCH_MODEL, provider_for_model
from helpers.feedback_links import (
    BUG_ISSUE_URL,
    DISCUSSIONS_URL,
    FEEDBACK_ISSUE_URL,
    ISSUES_URL,
    REPO_URL,
)
from helpers.step_types import (
    AUDIO_TRANSCRIBE,
    CAUSAL_RATING,
    EVENT_SEGMENT,
    SENTENCE_CORRECT,
    TEXT_MATCHING,
    TEXT_PARSING,
    audio_scope_for_step,
    normalize_pipeline_config,
    normalize_pipeline_step,
    normalize_step_type,
    step_runtime_key,
)


# Run pipeline subprocesses from the data workspace; serve UI assets from PACKAGE_ROOT.
os.chdir(str(WORKSPACE_ROOT))

# Initialize Flask app with template + static folders pinned to the package root.
# Both must be explicit: the module is loaded under the synthetic name
# ``narraters._legacy_server``, so Flask cannot derive its root path and would
# otherwise fall back to the current working directory (which we just chdir'd to
# WORKSPACE_ROOT). That made Flask's built-in ``static`` endpoint serve from
# ``WORKSPACE_ROOT/static`` (the user's data folder, no static/ there) and shadow
# the routes below — so every asset 404'd whenever the project root differed from
# the package (the documented ``pip install`` / NARRATERS_PROJECT_ROOT flows).
app = Flask(
    __name__,
    template_folder=str(PACKAGE_ROOT / 'templates'),
    static_folder=str(PACKAGE_ROOT / 'static'),
)


def _load_or_create_session_secret() -> bytes:
    """Persist the Flask session secret so users aren't logged out on every restart.

    Stored as a 0600 file under ~/.narraters/. If the directory isn't writable
    (rare — readonly $HOME), fall back to a process-lifetime random key.
    """
    secret_dir = Path.home() / ".narraters"
    secret_path = secret_dir / "session_secret"
    try:
        if secret_path.exists():
            data = secret_path.read_bytes().strip()
            if len(data) >= 32:
                return data
        secret_dir.mkdir(mode=0o700, exist_ok=True)
        token = secrets.token_hex(32).encode("ascii")
        secret_path.write_bytes(token)
        try:
            os.chmod(secret_path, 0o600)
        except OSError:
            pass
        return token
    except OSError:
        return secrets.token_hex(32).encode("ascii")


app.secret_key = _load_or_create_session_secret()

# Tighten session cookie defaults. SECURE is set below only when auth is enforced
# (production runs behind a TLS proxy); local http-only dev keeps it off so the
# cookie still works. SameSite=Lax + HttpOnly are cheap wins in every mode.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)


# --- App-level authentication (Flask-Login) ---------------------------------
#
# Password login gated by REQUIRE_AUTH, which `narraters serve --production`
# turns on (and which can be forced on anywhere via NARRATERS_REQUIRE_AUTH=1).
# When off (local dev), the before_request gate is a no-op and the UI is
# login-free. Accounts live in users.json (see narraters.accounts); manage them
# with `narraters users add/passwd/remove/list`.
from flask_login import (  # noqa: E402
    LoginManager,
    UserMixin,
    current_user,
    login_user,
    logout_user,
)

REQUIRE_AUTH = os.environ.get("NARRATERS_REQUIRE_AUTH") == "1"

if REQUIRE_AUTH:
    # Production sits behind a TLS-terminating reverse proxy (see HOSTING.md),
    # so the session cookie can be marked Secure (sent over HTTPS only).
    app.config["SESSION_COOKIE_SECURE"] = True

login_manager = LoginManager(app)
login_manager.login_view = "login_page"


class User(UserMixin):
    """A logged-in account. ``id`` is the username (the key in users.json)."""

    def __init__(self, username):
        self.id = username


@login_manager.user_loader
def _load_user(user_id):
    # Restore the session only if the account still exists (handles deletions).
    if user_id and user_id in load_users():
        return User(user_id)
    return None


def _safe_next(target):
    """Return ``target`` only if it is a safe local path (blocks open redirects)."""
    if target and target.startswith("/") and not target.startswith("//"):
        return target
    return None


# Endpoints reachable without logging in (the login page and static assets).
_AUTH_EXEMPT_ENDPOINTS = {"login_page", "static"}


@app.before_request
def _require_login():
    """Blanket auth gate: when REQUIRE_AUTH is on, every request must be logged
    in except the exempt endpoints. Covers routes that lack an explicit
    decorator (most /api/* endpoints), so nothing is accidentally left open."""
    if not REQUIRE_AUTH or current_user.is_authenticated:
        return None
    if request.endpoint in _AUTH_EXEMPT_ENDPOINTS or request.path.startswith("/static/") or request.path == "/favicon.ico":
        return None
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": "Authentication required"}), 401
    from urllib.parse import quote
    return redirect("/login?next=" + quote(request.full_path))


# --- Admin account -----------------------------------------------------------
#
# The account named "admin" gets the user-management panel (/admin) instead of
# the rating overview. In production the identity comes from the authenticated
# login; in local no-auth mode the free-text rater name suffices (a local
# operator already owns users.json, so this grants nothing extra).
ADMIN_USERNAME = 'admin'


def _is_admin():
    """True when the current request comes from the admin account."""
    if REQUIRE_AUTH:
        return current_user.is_authenticated and current_user.get_id() == ADMIN_USERNAME
    return session.get('username') == ADMIN_USERNAME


def admin_required(f):
    """Server-side admin gate for /api/admin/* (UI hiding alone is not enough)."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not _is_admin():
            return jsonify({'success': False, 'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return wrapper


def _apply_recall_rmatch_api_key(env, step_options, api_key):
    """Deprecated: rMatch recall path removed. Kept as no-op for older callers."""
    return


# Story segmentation presets that call Ollama (must match keys in scripts/2_story-event-segment.py SUPPORTED_MODELS).
_EVENT_SEGMENT_OLLAMA_MODEL_KEYS = frozenset(
    {"gemma4-e4b-ollama", "llama5.3-ollama", "llama3.3-ollama"}
)
_EVENT_SEGMENT_OLLAMA_DEFAULT_TAGS = {
    "gemma4-e4b-ollama": "gemma4:e4b",
    "llama5.3-ollama": "llama5.3",
    "llama3.3-ollama": "llama3.3",
}


def _gemma4_preflight_report(step_type: str, method: str | None, step_options: dict | None) -> dict | None:
    """
    When a step uses local Gemma-4 via Ollama E4B, run a lightweight check.
    Returns the report dict, or None if this step does not use that backend.
    """
    opts = step_options if isinstance(step_options, dict) else {}
    m = (method or "").strip().lower()
    try:
        from helpers.gemma_environment import check_ollama_gemma_e4b_environment
    except Exception as e:
        return {
            "ok": False,
            "errors": [f"Could not import Gemma check helper: {e}"],
            "warnings": [],
            "details": {},
        }

    if step_type == "textMatching" and m == "gemma-ollama":
        tag = (
            str(
                opts.get("recall_rating_ollama_model")
                or opts.get("recall_ollama_model")
                or opts.get("ollama_model")
                or ""
            ).strip()
            or None
        ) or os.environ.get("RECALL_RATING_OLLAMA_MODEL") or "gemma4:e4b"
        return check_ollama_gemma_e4b_environment(model_tag=tag)
    if step_type == "textParsing" and m == "gemma-ollama":
        tag = (
            str(opts.get("recall_parse_ollama_model") or opts.get("ollama_model") or "").strip()
            or None
        ) or os.environ.get("RECALL_PARSE_OLLAMA_MODEL") or "gemma4:e4b"
        return check_ollama_gemma_e4b_environment(model_tag=tag)
    if step_type == "sentenceCorrect" and m == "gemma-ollama":
        tag = (
            str(opts.get("ollama_model") or opts.get("spell_gram_ollama_model") or "").strip()
            or None
        ) or os.environ.get("SPELL_GRAM_OLLAMA_MODEL") or "gemma4:e4b"
        return check_ollama_gemma_e4b_environment(model_tag=tag)
    if step_type == "eventSegment" and m == "api":
        api_model = str(opts.get("model") or opts.get("event_segment_model") or "").strip()
        if api_model in _EVENT_SEGMENT_OLLAMA_MODEL_KEYS:
            default_tag = _EVENT_SEGMENT_OLLAMA_DEFAULT_TAGS.get(api_model, "gemma4:e4b")
            tag = (
                str(opts.get("ollama_model") or opts.get("event_segment_ollama_model") or "").strip()
                or None
            ) or os.environ.get("EVENT_SEGMENT_OLLAMA_MODEL") or default_tag
            return check_ollama_gemma_e4b_environment(model_tag=tag)
    return None


def _format_subprocess_step_error(result, prefix='Processing failed'):
    """Build a readable error string from subprocess stdout+stderr (tracebacks often span both)."""
    stderr = getattr(result, 'stderr', None) or ''
    stdout = getattr(result, 'stdout', None) or ''
    combined = (stderr + '\n' + stdout).strip()
    lines = [ln for ln in combined.split('\n') if ln.strip()]
    meaningful = None
    for line in reversed(lines):
        ls = line.strip()
        if ls.startswith('File ') or ls.startswith('Traceback'):
            continue
        low = ls.lower()
        if any(
            x in low
            for x in ('error', 'exception', 'failed', 'runtimeerror', 'importerror', 'valueerror')
        ):
            meaningful = ls
            break
    if meaningful is None:
        meaningful = '\n'.join(lines[-15:]) if lines else (stderr or stdout or 'Unknown error')
    if len(meaningful) > 4500:
        meaningful = meaningful[:4500] + '\n...(truncated)'
    return f'{prefix}: {meaningful}'


# --- User account storage location ---
#
# Pip-installed copies of the app share a single ``site-packages`` directory across users
# and may be read-only, so accounts must NOT live inside the package itself. Default to
# ``~/.narraters/`` (overridable via ``NARRATERS_DATA_DIR``), and migrate the legacy
# in-package files on first run so existing source-checkout users keep their account.
def _resolve_account_data_dir():
    override = (os.environ.get("NARRATERS_DATA_DIR") or "").strip()
    if override:
        return Path(override).expanduser().resolve()
    return Path.home() / ".narraters"


ACCOUNT_DATA_DIR = _resolve_account_data_dir()
USERS_FILE = ACCOUNT_DATA_DIR / "users.json"

# Legacy locations from earlier in-package storage (migrated on first run).
_LEGACY_USERS_FILE = PACKAGE_ROOT / "server" / "users.json"


def _migrate_legacy_user_file_once():
    try:
        if _LEGACY_USERS_FILE.exists() and not USERS_FILE.exists():
            ACCOUNT_DATA_DIR.mkdir(parents=True, exist_ok=True)
            try:
                USERS_FILE.write_bytes(_LEGACY_USERS_FILE.read_bytes())
                # Restrict to owner-only so other users on the same machine cannot read it.
                try:
                    os.chmod(USERS_FILE, 0o600)
                except OSError:
                    pass
                print(f"Migrated legacy users.json -> {USERS_FILE}")
            except Exception as e:
                print(f"Warning: could not migrate legacy users.json: {e}")
    except Exception:
        pass


_migrate_legacy_user_file_once()


def _pipeline_config_path():
    """Writable pipeline config location (pip installs use read-only site-packages)."""
    return ACCOUNT_DATA_DIR / "pipeline_config.json"


_LEGACY_PIPELINE_FILE = PACKAGE_ROOT / "pipeline_config.json"


def _migrate_legacy_pipeline_config_once():
    try:
        target = _pipeline_config_path()
        if _LEGACY_PIPELINE_FILE.exists() and not target.exists():
            ACCOUNT_DATA_DIR.mkdir(parents=True, exist_ok=True)
            try:
                target.write_bytes(_LEGACY_PIPELINE_FILE.read_bytes())
                try:
                    os.chmod(target, 0o600)
                except OSError:
                    pass
                print(f"Migrated legacy pipeline_config.json -> {target}")
            except Exception as e:
                print(f"Warning: could not migrate legacy pipeline_config.json: {e}")
    except Exception:
        pass


_migrate_legacy_pipeline_config_once()


# User accounts, password hashing, and verification now live in
# ``narraters.accounts`` so the CLI (``narraters users ...``) can reuse them
# without importing this Flask server module. Imported under the historical
# names so the rest of this file is unchanged.
from narraters.accounts import (  # noqa: E402
    add_user,
    get_batch_visibility,
    get_benchmark_pass,
    is_safe_username,
    load_users,
    remove_user,
    save_users,
    set_batch_visible,
    set_benchmark_pass,
    set_password,
    hash_password,
    verify_password,
    verify_user,
)
from narraters import __version__  # noqa: E402


def get_users():
    """Get current users dictionary."""
    return load_users()


# --- Fast directory listing cache -------------------------------------------
#
# The dashboard formerly called Path.glob() many times per (item × step) — each
# call re-scans the directory on disk. With dozens of subjects and steps that is
# thousands of redundant filesystem scans and is the main cause of slow loads.
#
# Instead we scan each directory once and answer every glob / exists / edit-file
# query from an in-memory listing. The cache is request-scoped (Flask ``g``) so
# it is always fresh on the next request; outside a request context (helper
# scripts, tests) it falls back to a short TTL so a single batch run stays
# correct.
import fnmatch as _fnmatch
import time as _time

_DIR_CACHE_TTL = 3.0  # seconds, only used outside a Flask request context
_dir_cache_fallback: dict = {}


def _dir_cache_store():
    """Return the dict used to cache directory listings for this scope."""
    try:
        if g:  # truthy only inside an active request/app context
            cache = getattr(g, '_dir_listing_cache', None)
            if cache is None:
                cache = {}
                g._dir_listing_cache = cache
            return cache, False
    except RuntimeError:
        pass
    return _dir_cache_fallback, True


def _list_dir_names(directory):
    """Cached list of entry names in ``directory`` (one os.scandir per dir).

    Returns [] for a missing/!dir path (that fact is cached too).
    """
    if directory is None:
        return []
    key = str(directory)
    cache, use_ttl = _dir_cache_store()
    now = _time.monotonic()
    hit = cache.get(key)
    if hit is not None:
        names, ts = hit
        if not use_ttl or (now - ts) < _DIR_CACHE_TTL:
            return names
    try:
        with os.scandir(directory) as it:
            names = [e.name for e in it]
    except PermissionError:
        # Do not cache permission failures — user may grant FDA and retry without restart.
        return []
    except (FileNotFoundError, NotADirectoryError, OSError):
        names = []
    cache[key] = (names, now)
    return names


def _dir_exists_cached(directory):
    """True if ``directory`` exists and is non-empty-listable (cached scandir)."""
    if directory is None:
        return False
    # A real, empty directory still "exists"; distinguish via a cheap probe only
    # when the cached listing is empty.
    names = _list_dir_names(directory)
    if names:
        return True
    try:
        return Path(directory).is_dir()
    except OSError:
        return False


def cached_glob(directory, pattern):
    """Drop-in for ``Path(directory).glob(pattern)`` backed by the dir cache.

    Only supports flat (non-recursive) patterns, which is all the dashboard
    uses. Returns a list of ``Path`` objects.
    """
    if directory is None:
        return []
    base = Path(directory)
    return [base / n for n in _list_dir_names(directory) if _fnmatch.fnmatch(n, pattern)]


def glob_files_in_dir(directory, pattern):
    """Glob flat files in ``directory``, with a direct scandir fallback when the dir cache is empty."""
    hits = cached_glob(directory, pattern)
    if hits:
        return hits
    try:
        base = Path(directory)
        if not base.is_dir():
            return []
        with os.scandir(base) as it:
            return [
                base / entry.name
                for entry in it
                if entry.is_file(follow_symlinks=False) and _fnmatch.fnmatch(entry.name, pattern)
            ]
    except PermissionError:
        print(f"Warning: permission denied listing {directory}")
        return []
    except OSError as e:
        print(f"Warning: could not list {directory}: {e}")
        return []


def cached_path_exists(path):
    """Cached ``Path.exists()`` for a single file, via its parent dir listing."""
    if path is None:
        return False
    p = Path(path)
    return p.name in _list_dir_names(p.parent)


def is_user_edit_file(filename):
    """Check if a filename is a user-edit file (matches _{username}-edit pattern before extension)."""
    return bool(re.search(r'_\w+-edit\.', str(filename)))


# Markers that only appear in downstream recall/causal pipeline outputs, never in a
# genuine story-events file. Recall/causal exports embed the "{story}_events-..."
# prefix in their default filenames (e.g.
# ``the_siren_events_the_siren_sub-01_rate-recall_manual_Rater-edit.xlsx``), which
# would otherwise be mistaken for the story-events file and shadow it.
_NON_STORY_EVENTS_MARKERS = (
    'rate-recall', 'recall-version', 'recall-match',
    'causal-rating', 'causal-linguistic', 'textparsing',
)


def is_story_events_filename(filename):
    """True unless the name carries a downstream-step marker (recall/causal output)."""
    low = str(filename).lower()
    return not any(marker in low for marker in _NON_STORY_EVENTS_MARKERS)


def get_edit_suffix():
    """Get the edit suffix for the current logged-in user (e.g., '_username-edit')."""
    username = session.get('username', 'human')
    return f"_{username}-edit"


def extract_edit_username(filename):
    """Extract the username from a user-edit filename. Returns None if not an edit file."""
    match = re.search(r'_(\w+)-edit\.', str(filename))
    return match.group(1) if match else None


def _list_edit_files(directory, base_prefix, middle='', ext='.txt'):
    """All user-edit files for ``{base_prefix}{middle}``, newest first.

    Accepts both the canonical name ``{base_prefix}{middle}_{user}-edit{ext}`` and
    longer names that carry method/model tags inherited from earlier pipeline
    steps, e.g. ``{base_prefix}_spell-X{middle}-Y_{user}-edit{ext}``. Word-boundary
    guarded so ``sub-01`` does not pick up ``sub-012``.
    """
    if not directory or not _dir_exists_cached(directory):
        return []
    pattern = f"{base_prefix}*{middle}*_*-edit{ext}" if middle else f"{base_prefix}*_*-edit{ext}"
    out = []
    for f in sorted(cached_glob(directory, pattern), key=lambda p: p.stat().st_mtime, reverse=True):
        if not is_user_edit_file(f.name):
            continue
        stem = f.stem
        if middle:
            if not (stem.startswith(f"{base_prefix}_") or stem.startswith(f"{base_prefix}{middle}")):
                continue
        else:
            if not (stem == base_prefix or stem.startswith(f"{base_prefix}_")):
                continue
        out.append(f)
    return out


def find_best_edit_file(directory, base_prefix, middle='', ext='.txt'):
    """Find the best user-edit file, preferring current user's edit.
    base_prefix: e.g., 'subj001'
    middle: e.g., '_parsed' or '_rate-recall' or '_events'
    ext: e.g., '.txt' or '.xlsx'
    Returns Path or None.
    """
    edits = _list_edit_files(directory, base_prefix, middle, ext)
    if not edits:
        return None
    try:
        username = session.get('username', 'human')
    except RuntimeError:
        username = 'human'
    for f in edits:
        if extract_edit_username(f.name) == username:
            return f
    if username != 'human':
        for f in edits:
            if extract_edit_username(f.name) == 'human':
                return f
    return edits[0]


def has_any_edit_file(directory, base_prefix, middle='', ext='.txt'):
    """Check if any user-edit file exists for the given pattern."""
    return bool(_list_edit_files(directory, base_prefix, middle, ext))


def get_all_edit_versions(directory, base_prefix, middle='', ext='.txt'):
    """Get list of all available user-edit version names for a file pattern.
    Returns list of strings like ['editor1-edit', 'editor2-edit'].
    """
    versions = []
    seen = set()
    for f in _list_edit_files(directory, base_prefix, middle, ext):
        username = extract_edit_username(f.name)
        if username:
            version_name = f"{username}-edit"
            if version_name not in seen:
                seen.add(version_name)
                versions.append(version_name)
    return versions


# --- Manage folder for user activity records ---
#
# Same rationale as USERS_FILE: keep user-specific account/activity data in the per-user
# data dir (``~/.narraters/manage/`` by default) rather than inside the installed package.
MANAGE_DIR = ACCOUNT_DATA_DIR / "manage"
_LEGACY_MANAGE_DIR = PACKAGE_ROOT / "manage"


def _migrate_legacy_manage_dir_once():
    legacy_records = _LEGACY_MANAGE_DIR / "user_records.json"
    new_records = MANAGE_DIR / "user_records.json"
    try:
        if legacy_records.exists() and not new_records.exists():
            MANAGE_DIR.mkdir(parents=True, exist_ok=True)
            new_records.write_bytes(legacy_records.read_bytes())
            try:
                os.chmod(new_records, 0o600)
            except OSError:
                pass
            print(f"Migrated legacy user_records.json -> {new_records}")
    except Exception as e:
        print(f"Warning: could not migrate legacy user_records.json: {e}")


_migrate_legacy_manage_dir_once()


def get_user_records():
    """Load user records from the per-user manage folder."""
    records_file = MANAGE_DIR / "user_records.json"
    if records_file.exists():
        try:
            with open(records_file, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_user_records(records):
    """Save user records to the per-user manage folder (owner-only permissions)."""
    try:
        MANAGE_DIR.mkdir(parents=True, exist_ok=True)
        records_file = MANAGE_DIR / "user_records.json"
        with open(records_file, "w") as f:
            json.dump(records, f, indent=2)
        try:
            os.chmod(records_file, 0o600)
        except OSError:
            pass
    except Exception as e:
        print(f"Error saving user records: {e}")


def log_user_login(username):
    """Log a user login event to manage folder."""
    records = get_user_records()
    if username not in records:
        records[username] = {
            'username': username,
            'password_hash': '',
            'created_at': datetime.now().isoformat(),
            'logins': [],
            'edits': []
        }
    records[username]['logins'].append({
        'timestamp': datetime.now().isoformat()
    })
    save_user_records(records)


def log_user_edit(username, edit_type, subject_id, filename):
    """Log a user edit event to manage folder."""
    records = get_user_records()
    if username not in records:
        records[username] = {
            'username': username,
            'password_hash': '',
            'created_at': datetime.now().isoformat(),
            'logins': [],
            'edits': []
        }
    records[username]['edits'].append({
        'timestamp': datetime.now().isoformat(),
        'type': edit_type,
        'subject': subject_id,
        'file': filename
    })
    save_user_records(records)


# Supported audio/video file extensions for transcription and playback
SUPPORTED_AUDIO_EXTENSIONS = ('.wav', '.mp3', '.mp4', '.m4a', '.flac', '.ogg', '.webm', '.aac')

# Directory paths (relative to project root)
DATA_DIR = WORKSPACE_ROOT / 'data'
STORY_EVENTS_DIR = WORKSPACE_ROOT / 'data' / '3_story_events'
STORY_TRANSCRIPT_DIR = WORKSPACE_ROOT / 'data' / '2_story_transcript'
STORY_AUDIO_DIR = WORKSPACE_ROOT / 'data' / '1_story_audio'
RECALL_AUDIO_DIR = WORKSPACE_ROOT / 'data' / '4_recall_audio'
OUTPUT_DIR = WORKSPACE_ROOT / 'output'
RECALL_CORRECTED_DIR = OUTPUT_DIR / 'recall_corrected'
RECALL_PARSED_DIR = OUTPUT_DIR / 'recall_parsed'
RECALL_RATED_DIR = OUTPUT_DIR / 'recall_rated'
RECALL_AUDIO_TRANSCRIBED_DIR = OUTPUT_DIR / 'recall_audio-transcribed'
STORY_AUDIO_TRANSCRIBED_DIR = OUTPUT_DIR / 'story_audio-transcribed'
CAUSAL_RATED_DIR = OUTPUT_DIR / 'causal_rated'

# --- Benchmark mode -------------------------------------------------------
# `narraters serve --benchmark` sets NARRATERS_BENCHMARK=1. In this mode the
# landing page is the text-matching benchmark overview (see /benchmark): a list
# of recall files under benchmark/unrated/ to rate, with results written under
# benchmark/rated/<username>/.
#
# NARRATERS_BENCHMARK_DIR overrides where that benchmark/ tree lives (default:
# WORKSPACE_ROOT/benchmark). Hosting setups point it at a dedicated, group-shared
# directory the admin can rsync into/out of as a different user than the service
# user (see HOSTING.md "Sync benchmark data"). It is admin-set in systemd, never
# request-controlled.
BENCHMARK_MODE = os.environ.get('NARRATERS_BENCHMARK') == '1'
BENCHMARK_DIR = Path(
    os.environ.get('NARRATERS_BENCHMARK_DIR') or (WORKSPACE_ROOT / 'benchmark')
).expanduser().resolve()
BENCHMARK_UNRATED_DIR = BENCHMARK_DIR / 'unrated'
BENCHMARK_RATED_DIR = BENCHMARK_DIR / 'rated'

# Excel files to check for raw recall data
EXCEL_FILES = [
    f for f in (WORKSPACE_ROOT / 'data').glob('summary_*.xlsx')
] if (WORKSPACE_ROOT / 'data').exists() else []

# Step type to default output directory (used when pipeline config has no outputPath)
STEP_TYPE_DEFAULT_OUTPUT = {
    'sentenceCorrect': RECALL_CORRECTED_DIR,
    'textParsing': RECALL_PARSED_DIR,
    'textMatching': RECALL_RATED_DIR,
    'causalRating': CAUSAL_RATED_DIR,
    'eventSegment': STORY_EVENTS_DIR,
    'audioTranscribe:story': STORY_AUDIO_TRANSCRIBED_DIR,
    'audioTranscribe:recall': RECALL_AUDIO_TRANSCRIBED_DIR,
}


def _resolve_path_from_config(path_str):
    """Convert path string from pipeline config to absolute Path.

    Supports POSIX absolute paths (``/Users/...``), Windows absolute paths
    (``C:\\Users\\...``), tilde expansion (``~/Desktop/...``), and project-
    relative paths (resolved against ``WORKSPACE_ROOT``).
    """
    if not path_str or not path_str.strip():
        return None
    p = path_str.strip().rstrip('/').rstrip('\\')
    if not p:
        return None
    p_obj = Path(p).expanduser()
    if p_obj.is_absolute():
        return p_obj
    # Paths saved without a leading ``/`` or ``~`` (e.g. ``Dropbox/proj/...``).
    if p.startswith(('Dropbox/', 'Users/', 'Library/')):
        try:
            return Path.home() / p
        except RuntimeError:
            pass
    return WORKSPACE_ROOT / p


def _path_for_client(p):
    """Render a file/directory path as a string for the web client.

    Returns a project-relative POSIX path when ``p`` lives under
    ``WORKSPACE_ROOT``;
    otherwise returns the absolute path so externally-located input/output
    folders (e.g. ones the user drag-dropped from Finder/Explorer) survive
    round-tripping through the UI.
    """
    try:
        return str(Path(p).relative_to(WORKSPACE_ROOT))
    except (ValueError, TypeError):
        return str(p)


def get_output_dir_for_step_type(step_type):
    """Get output directory for a step type from pipeline config, or default path.
    Returns Path or None.
    """
    config = get_pipeline_config()
    if config and config.get('steps'):
        for step in config['steps']:
            if step_runtime_key(step) == step_type:
                out = step.get('outputPath', '')
                if out:
                    return _resolve_path_from_config(out)
    # Fallback to default
    return STEP_TYPE_DEFAULT_OUTPUT.get(step_type)


def _iter_pipeline_config_io_dirs():
    """Yield resolved input/output directories from the active pipeline config."""
    config = get_pipeline_config()
    if not config or not config.get('steps'):
        return
    for step in config['steps']:
        for key in ('inputPath', 'outputPath'):
            d = _resolve_path_from_config(step.get(key, ''))
            if d is not None:
                yield d
        for _field, _env_var, d in _resolve_step_extra_input_dirs(step):
            yield d


def iter_story_events_search_dirs():
    """Directories to search for story-event .xlsx files.
    Search order (each path included at most once):
    1. Optional ``NARRATIVE_STORY_EVENTS_DIR`` env (absolute path to the folder).
    2. All pipeline-configured input/output folders (e.g. causal-rating input on Desktop).
    3. Pipeline-configured story-event output, else ``PROJECT_ROOT/data/3_story_events``.
    4. ``PROJECT_ROOT/data/3_story_events`` (Flask ``PROJECT_ROOT`` is the ``software/`` package root).
    5. ``<parent of PROJECT_ROOT>/data/3_story_events`` when data lives next to ``software/``
       (monorepo layout: ``repo/data/`` vs ``repo/software/data/``).
    """
    seen = set()
    candidates = []
    env_override = (os.environ.get('NARRATIVE_STORY_EVENTS_DIR') or '').strip()
    if env_override:
        candidates.append(Path(env_override).expanduser())
    candidates.extend(_iter_pipeline_config_io_dirs())
    primary = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
    sibling_repo_events = WORKSPACE_ROOT.parent / 'data' / '3_story_events'
    candidates.extend([primary, STORY_EVENTS_DIR, sibling_repo_events])
    for d in candidates:
        if d is None:
            continue
        try:
            key = d.resolve()
        except OSError:
            key = str(d)
        if key not in seen:
            seen.add(key)
            yield d


def get_input_dir_for_step_type(step_type):
    """Get input path for a step type from pipeline config.
    Returns Path (for directory) or Path (for file), or None.
    """
    config = get_pipeline_config()
    if config and config.get('steps'):
        for step in config['steps']:
            if step_runtime_key(step) == step_type:
                inp = step.get('inputPath', '')
                if inp:
                    return _resolve_path_from_config(inp)
    return None


def get_subject_id_from_filename(filename):
    """Extract subject ID from filename (e.g., 'subN_XXXX.txt' -> 'subN_XXXX').

    Handles the canonical ``{subject}_parsed.xlsx`` / ``{subject}_rate-recall-...``
    names AND the alternative convention with a ``story-`` prefix and hyphenated
    step tokens, e.g. ``story-alice_14_sub-3008-parsed.csv`` -> ``alice_14_sub-3008``
    and ``story-alice_14_sub-3008-recall-matched.csv`` -> ``alice_14_sub-3008``.
    """
    import re
    name = filename
    for ext in ('.txt', '.csv', '.tsv', '.xlsx', '.xls') + SUPPORTED_AUDIO_EXTENSIONS:
        if name.lower().endswith(ext):
            name = name[:-len(ext)]
            break
    # trailing user-edit (e.g. _alice-edit)
    name = re.sub(r'_\w+-edit$', '', name)
    # alternative "story-" prefix
    name = re.sub(r'^story-', '', name)
    # Step tokens / method suffixes — both underscore and hyphen forms, anchored to the
    # end. The optional ``(-...)`` tail also absorbs a ``_method`` chain
    # (e.g. ``_parsed-ollama_gemma4_e4b``, ``_spell-ollama_gemma4_e4b``) so the two
    # method variants of one subject collapse to the same id. Applied repeatedly so a
    # canonical token after a method tag is also removed.
    # The optional ``(-...)`` tail (hyphen-started) absorbs a method chain such as
    # ``-ollama_gemma4_e4b`` while staying anchored to the end, so it can't eat the
    # subject id that *precedes* an ``_events`` token in legacy verbose names.
    _suffixes = (
        r'[-_]rate-recall(-[a-zA-Z0-9_.-]+)?$',
        r'[-_]recall-matched(-[a-zA-Z0-9_.-]+)?$',
        r'[-_]events(-[a-zA-Z0-9_.-]+)?$',
        r'[-_]segmented$',
        r'[-_]segment$',
        r'[-_]parsed(-[a-zA-Z0-9_.-]+)?$',
        r'[-_]rated$',
        r'[-_]spell(-[a-zA-Z0-9_.-]+)?$',          # corrected method tag, e.g. _spell-ollama_gemma4_e4b
        r'[-_](recall-)?corrected(-[a-zA-Z0-9_.-]+)?$',
    )
    prev = None
    while prev != name:
        prev = name
        name = re.sub(r'_\w+-edit$', '', name)
        for pat in _suffixes:
            name = re.sub(pat, '', name)
    # legacy subN_YYYY pattern
    match = re.search(r'(sub\d+_\d+)', name)
    if match:
        return match.group(1)
    return name


def is_story_focused_pipeline(pipeline_config):
    """Check if pipeline is focused on story processing (vs recall processing)."""
    if not pipeline_config or not pipeline_config.get('steps'):
        return False
    
    story_step_types = {'audioTranscribe:story', 'eventSegment', 'causalRating'}
    for step in pipeline_config['steps']:
        if step_runtime_key(step) in story_step_types:
            return True
    return False


def get_story_name_from_filename(filename):
    """Extract story name from filename (e.g., 'my_story.wav' -> 'my_story').
    Handles method-suffixed filenames (e.g., my_story_events-fine.xlsx -> my_story).
    """
    import re
    # Remove common extensions
    name = filename
    for ext in ('.txt', '.csv', '.tsv', '.xlsx', '.xls') + SUPPORTED_AUDIO_EXTENSIONS:
        name = name.replace(ext, '')
    # Remove any user-edit suffix (e.g., _human-edit, _username-edit)
    name = re.sub(r'_\w+-edit', '', name)
    # Alternative "story-{name}-segmented" convention: drop the prefix + suffix.
    name = re.sub(r'^story-', '', name)
    name = re.sub(r'[-_]segmented$', '', name)
    name = re.sub(r'[-_]segment$', '', name)
    # Remove _events suffix with optional method suffix (e.g., _events-api_<model-id>)
    name = re.sub(r'_events(-[a-zA-Z0-9_.-]+)?', '', name)
    return name


def get_story_name_from_subject_id(subj_id):
    """Try to extract the story name from a subject ID.
    E.g., 'the_siren_sub-01' -> 'the_siren', 'pieman_edited_sub-03' -> 'pieman_edited'.
    Returns None if no subject suffix pattern is found.
    """
    import re
    m = re.match(r'^(.+?)_sub-?\d+$', subj_id)
    return m.group(1) if m else None


_EDITORIAL_STORY_SUFFIXES = ('_edited', '_edit', '_revised', '_clean', '_final')


def _editorial_story_stem_alternate(stem):
    """If stem ends with a common transcript rename suffix, return a shorter stem, else None.

    Segmentation output is often named from an earlier id (e.g. ``pieman_events.xlsx``) while the
    transcript in ``2_story_transcript`` may be ``pieman_edited.txt``, so the UI story id becomes
    ``pieman_edited`` and would not match ``pieman_events*.xlsx`` without this fallback.
    """
    if not stem:
        return None
    for suf in _EDITORIAL_STORY_SUFFIXES:
        if stem.endswith(suf) and len(stem) > len(suf):
            return stem[: -len(suf)]
    return None


def expand_story_event_file_bases(item_id):
    """Basenames used to find ``{base}_events*.xlsx``: id, derived story name, editorial aliases.

    Order is preserved so exact ids win over fallbacks. Subject ids that contain ``_sub-`` are not
    shortened by editorial rules; only story-style stems (e.g. ``pieman_edited``) gain alternates.
    """
    if not item_id:
        return []
    seen = set()
    out = []

    def add(b):
        if b and b not in seen:
            seen.add(b)
            out.append(b)

    add(item_id)
    add(get_story_name_from_subject_id(item_id))
    for b in list(out):
        if not b or '_sub-' in b:
            continue
        alt = _editorial_story_stem_alternate(b)
        if alt:
            add(alt)
    return out


# ------------------------------------------------------------------
# Per-step input-variant enumeration (for the launcher dropdown UI).
# ------------------------------------------------------------------

# Stream definitions: which logical input streams each step_type consumes.
# Each entry returns the inputs needed by ``enumerate_step_input_variants``:
#   - key:       short id used by the frontend and forwarded env var name
#   - label:     human-readable label for the dropdown
#   - env_var:   env var the launcher will set to the chosen variant suffix
#   - extension: file extension to scan (lowercase, with leading dot, or "" for any)
#   - tail_pat:  optional fixed tail substring required in the stem (e.g. ``_events``
#                for story-events streams) — empty string means "no constraint"
_STEP_INPUT_STREAMS = {
    'audioTranscribe:story':  [('main',          'Input',         'BATCH_INPUT_VARIANT',         'audio', '')],
    'audioTranscribe:recall': [('main',          'Input',         'BATCH_INPUT_VARIANT',         'audio', '')],
    'eventSegment':     [('main',          'Story Transcript', 'BATCH_INPUT_VARIANT',      ('.txt', '.csv', '.tsv', '.xlsx', '.xls'), '')],
    'sentenceCorrect':   [('main',          'Recall Text',   'BATCH_INPUT_VARIANT',         ('.txt', '.csv', '.tsv', '.xlsx'), '')],
    'textParsing':            [('main',          'Corrected Recall', 'BATCH_INPUT_VARIANT',      ('.txt', '.csv', '.tsv', '.xlsx', '.xls'), '')],
    'textMatching':     [
        ('parsed_recall', 'Parsed Recall', 'BATCH_INPUT_VARIANT',          ('.xlsx', '.csv', '.tsv'), '_parsed'),
        ('story_events',  'Story Events',  'BATCH_STORY_EVENTS_VARIANT',   ('.xlsx', '.csv', '.tsv'), '_events'),
    ],
    'causalRating':      [('story_events',  'Story Events',  'BATCH_STORY_EVENTS_VARIANT',  ('.xlsx', '.csv', '.tsv'), '_events')],
}

# Extra input path fields on pipeline step config (beyond inputPath / outputPath).
# Each entry: (config field name, env var passed to the batch script).
_STEP_EXTRA_INPUTS = {
    'textMatching': [
        ('storyEventsPath', 'BATCH_STORY_EVENTS_DIR'),
    ],
}


def _step_extra_input_fields(step_type):
    """Return (field, env_var) pairs for optional extra inputs on ``step_type``."""
    return _STEP_EXTRA_INPUTS.get(step_type, [])


def _resolve_step_extra_input_dirs(step):
    """Yield (field, env_var, resolved Path) for configured extra inputs on ``step``."""
    if not step:
        return
    step_type = step_runtime_key(step)
    for field, env_var in _step_extra_input_fields(step_type):
        raw = (step.get(field) or '').strip()
        if not raw:
            continue
        resolved = _resolve_path_from_config(raw)
        if resolved is not None:
            yield field, env_var, resolved


def _apply_step_path_env(env, step):
    """Set BATCH_* directory env vars from a pipeline step's configured paths."""
    input_path = step.get('inputPath', '')
    output_path = step.get('outputPath', '')
    if input_path:
        inp = _resolve_path_from_config(input_path)
        if inp is not None:
            env['BATCH_INPUT_DIR'] = str(inp)
    if output_path:
        out = _resolve_path_from_config(output_path)
        if out is not None:
            env['BATCH_OUTPUT_DIR'] = str(out)
    for _field, env_var, resolved in _resolve_step_extra_input_dirs(step):
        env[env_var] = str(resolved)


def _audio_files_for_item(root, item_id):
    """All audio-extension files in ``root`` whose stem begins with ``item_id``."""
    if not root.is_dir() or not item_id:
        return []
    out = []
    for ext in SUPPORTED_AUDIO_EXTENSIONS:
        out.extend(root.glob(f"{item_id}*{ext}"))
    return out


def _resolve_stream_dir(step_inputPath, stream_key, item_id, step=None):
    """Resolve the directory to search for a given stream.

    Most streams use the step's configured ``inputPath``. The story-events stream
    on textMatching uses ``storyEventsPath`` when set; otherwise falls back to
    ``data/3_story_events`` when the primary input is the parsed-recall directory.
    """
    if stream_key == 'story_events' and step:
        extra = (step.get('storyEventsPath') or '').strip()
        if extra:
            return _resolve_path_from_config(extra)
    if stream_key == 'story_events' and step_inputPath and 'story_events' not in step_inputPath:
        return STORY_EVENTS_DIR
    if not step_inputPath:
        return None
    return _resolve_path_from_config(step_inputPath)


def _scan_variants_for_item(stream_dir, item_id, ext, tail_pat, stream_key):
    """Return a list of variant-suffix strings present for one item in one stream.

    Variant semantics:
      - For streams with a fixed tail (e.g. ``_events`` or ``_parsed``), the variant
        is the part of the stem AFTER the tail (e.g. ``-fine`` for ``..._events-fine``).
      - For streams with no tail, the variant is the part of the stem AFTER the item_id
        (e.g. ``_spell-ollama_gemma4_e4b`` for ``{id}_spell-ollama_gemma4_e4b.txt``).
      - The empty string ``""`` represents the canonical filename.
    """
    if stream_dir is None or not stream_dir.is_dir() or not item_id:
        return []

    if stream_key == 'story_events':
        # Use basename expansion: try item_id, then derived story name + editorial aliases.
        bases = expand_story_event_file_bases(item_id)
    else:
        bases = [item_id]

    variants = set()

    def _try_collect(base):
        if ext == 'audio':
            files = _audio_files_for_item(stream_dir, base)
            for f in files:
                stem = f.stem
                # Audio variant = stem minus item_id (no _events tail). Skip if no match.
                if stem == base:
                    variants.add('')
                elif stem.startswith(base):
                    variants.add(stem[len(base):])
        else:
            exts = ext if isinstance(ext, (list, tuple)) else (ext,)
            for one_ext in exts:
                for f in stream_dir.glob(f"{base}*{one_ext}"):
                    if not f.is_file():
                        continue
                    stem = f.stem
                    if tail_pat:
                        if stem == f"{base}{tail_pat}":
                            variants.add('')
                        elif stem.startswith(f"{base}{tail_pat}"):
                            variants.add(stem[len(base) + len(tail_pat):])
                        elif stream_key != 'story_events' and stem.startswith(f"{base}_") and tail_pat in stem:
                            variants.add(stem[len(base):])
                    else:
                        if stem == base:
                            variants.add('')
                        elif stem.startswith(base):
                            variants.add(stem[len(base):])

    for b in bases:
        _try_collect(b)
        if variants:
            # Use the first base that produced any matches (matches step 5's expand-search semantics).
            break

    return sorted(variants)


def _label_for_variant(suffix, tail_pat):
    """Render a human-friendly label for a variant suffix."""
    if suffix == '':
        if tail_pat:
            return f"canonical ({tail_pat.lstrip('_')})"
        return 'canonical'
    s = suffix.lstrip('-').lstrip('_')
    return s


def enumerate_step_input_variants(step_type, item_ids, step_inputPath, step=None):
    """Compute available input variants for ``step_type``, intersected across ``item_ids``.

    Returns a dict ``{"streams": [...]}`` describing each logical input stream and the
    list of variant suffixes available for the batch. A variant is included only when
    every item_id in ``item_ids`` has a matching file with that suffix.

    The empty-string suffix represents the canonical (no-suffix) filename and is
    treated like any other variant: it's only offered when *all* items have it.
    """
    stream_defs = _STEP_INPUT_STREAMS.get(step_type, [])
    if not stream_defs or not item_ids:
        return {'streams': []}

    out_streams = []
    for stream_key, stream_label, env_var, ext, tail_pat in stream_defs:
        per_item_variants = {}  # item_id -> set[str]
        for item_id in item_ids:
            stream_dir = _resolve_stream_dir(step_inputPath, stream_key, item_id, step=step)
            found = _scan_variants_for_item(stream_dir, item_id, ext, tail_pat, stream_key)
            per_item_variants[item_id] = set(found)

        # Intersection
        if per_item_variants:
            common = set.intersection(*per_item_variants.values()) if all(per_item_variants.values()) else set()
        else:
            common = set()

        variants_out = []
        for suffix in sorted(common):
            count = sum(1 for s in per_item_variants.values() if suffix in s)
            variants_out.append({
                'suffix': suffix,
                'label': _label_for_variant(suffix, tail_pat),
                'count': count,
            })

        out_streams.append({
            'key': stream_key,
            'label': stream_label,
            'env_var': env_var,
            'variants': variants_out,
        })

    return {'streams': out_streams}


def get_all_stories(pipeline_config):
    """Get list of all available stories from story input and output directories.
    Returns stories discovered from all story-focused steps' input and output paths.
    """
    stories_set = set()
    stories_ordered = []
    
    # Check all story-focused steps to discover stories from both input and output paths
    if pipeline_config and pipeline_config.get('steps'):
        for step in pipeline_config.get('steps', []):
            step_type = step_runtime_key(step)
            if step_type in ['audioTranscribe:story', 'eventSegment', 'causalRating']:
                # Check input path
                input_path = step.get('inputPath', '')
                if input_path:
                    discovered = discover_items_from_path(input_path, step_type, is_story=True, scan_type='input')
                    for story_name in discovered:
                        if story_name and story_name not in stories_set:
                            stories_set.add(story_name)
                            stories_ordered.append(story_name)
                
                # Check output path (for existing results)
                output_path = step.get('outputPath', '')
                if output_path:
                    discovered = discover_items_from_path(output_path, step_type, is_story=True, scan_type='output')
                    for story_name in discovered:
                        if story_name and story_name not in stories_set:
                            stories_set.add(story_name)
                            stories_ordered.append(story_name)
    
    # Also check default directories if no stories found from pipeline config
    if not stories_ordered:
        # Default to story audio directory
        story_input_dir = STORY_AUDIO_DIR
        
        # Get all story files from the input directory
        if story_input_dir.exists():
            # Get audio files (all supported formats)
            audio_files = []
            for ext in SUPPORTED_AUDIO_EXTENSIONS:
                audio_files.extend(story_input_dir.glob(f'*{ext}'))
            # Get transcript files
            transcript_files = list(story_input_dir.glob('*.txt'))
            # Get event files (from story events directory)
            if STORY_EVENTS_DIR.exists():
                event_files = list(STORY_EVENTS_DIR.glob('*_events.xlsx'))
            else:
                event_files = []
            
            # Combine all files and extract story names
            all_files = audio_files + transcript_files + event_files
            # Filter out human-edit files for ordering
            regular_files = [f for f in all_files if not is_user_edit_file(f.name)]
            # Sort by filename to maintain consistent order
            regular_files.sort(key=lambda f: f.name)
            
            for file in regular_files:
                story_name = get_story_name_from_filename(file.name)
                if story_name and story_name not in stories_set:
                    stories_set.add(story_name)
                    stories_ordered.append(story_name)
    
    return stories_ordered


def discover_items_from_path(path, step_type, is_story=False, scan_type='any'):
    """Discover item IDs (subjects or stories) from files in a given path based on step type.
    
    Args:
        path: Directory path (Path object or string)
        step_type: Type of step (e.g., 'sentenceCorrect', 'textParsing', etc.)
        is_story: Whether discovering stories (True) or subjects (False)
        scan_type: 'input' when scanning an input directory, 'output' when scanning an
                   output directory, or 'any' to use patterns covering both (default).
    
    Returns:
        Set of item IDs found in the path
    """
    items = set()
    
    # Normalize path - resolve absolute (POSIX/Windows/tilde) vs project-relative.
    if isinstance(path, str):
        stripped = path.strip().rstrip('/').rstrip('\\')
        p_obj = Path(stripped).expanduser() if stripped else None
        if p_obj is None or not stripped:
            dir_path = WORKSPACE_ROOT
        elif p_obj.is_absolute():
            dir_path = p_obj
        else:
            dir_path = WORKSPACE_ROOT / stripped
    else:
        dir_path = path
    
    if not _dir_exists_cached(dir_path):
        return items
    
    # Determine file patterns based on step type and scan direction
    if step_type == 'audioTranscribe:story':
        if scan_type == 'input':
            patterns = [f'*{ext}' for ext in SUPPORTED_AUDIO_EXTENSIONS]
        elif scan_type == 'output':
            patterns = ['*.txt']
        else:
            patterns = ['*.txt'] + [f'*{ext}' for ext in SUPPORTED_AUDIO_EXTENSIONS]
        extract_func = get_story_name_from_filename if is_story else get_subject_id_from_filename
    elif step_type == 'audioTranscribe:recall':
        if scan_type == 'input':
            patterns = [f'*{ext}' for ext in SUPPORTED_AUDIO_EXTENSIONS]
        elif scan_type == 'output':
            patterns = ['*.txt']
        else:
            patterns = ['*.txt'] + [f'*{ext}' for ext in SUPPORTED_AUDIO_EXTENSIONS]
        extract_func = get_subject_id_from_filename
    elif step_type == 'eventSegment':
        if scan_type == 'input':
            patterns = ['*.txt', '*.csv', '*.tsv', '*.xlsx', '*.xls']
        elif scan_type == 'output':
            patterns = [
                '*_events.xlsx', '*_events-*.xlsx', '*_events_rule-based.xlsx', '*_events_api.xlsx',
                '*_events.csv', '*_events-*.csv', '*_events.tsv', '*_events-*.tsv',
            ]
        else:
            patterns = [
                '*.txt', '*.csv', '*.tsv', '*.xlsx', '*.xls',
                '*_events.xlsx', '*_events-*.xlsx', '*_events_rule-based.xlsx', '*_events_api.xlsx',
                '*_events.csv', '*_events-*.csv', '*_events.tsv', '*_events-*.tsv',
            ]
        extract_func = get_story_name_from_filename if is_story else get_subject_id_from_filename
    elif step_type == 'sentenceCorrect':
        patterns = ['*.txt', '*.csv', '*.tsv', '*.xlsx']
        extract_func = get_subject_id_from_filename
    elif step_type == 'textParsing':
        if scan_type == 'input':
            patterns = ['*.txt', '*.csv', '*.tsv', '*.xlsx', '*.xls']
        elif scan_type == 'output':
            patterns = ['*_parsed.xlsx', '*_parsed.csv', '*_parsed.tsv']
        else:
            patterns = ['*.txt', '*.csv', '*.tsv', '*.xlsx', '*_parsed.xlsx', '*_parsed.csv', '*_parsed.tsv']
        extract_func = get_subject_id_from_filename
    elif step_type == 'textMatching':
        if scan_type == 'input':
            patterns = [
                '*_parsed.xlsx', '*_parsed.csv', '*_parsed.tsv',
                '*_events.xlsx', '*_events-*.xlsx', '*_events.csv', '*_events-*.csv', '*_events.tsv',
            ]
        elif scan_type == 'output':
            patterns = ['*_rate-recall.xlsx', '*_rate-recall-*.xlsx', '*_rate-recall.csv', '*_rate-recall-*.csv']
        else:
            patterns = [
                '*_rate-recall.xlsx', '*_rate-recall-*.xlsx', '*_rate-recall.csv',
                '*_parsed.xlsx', '*_parsed.csv', '*_events.xlsx', '*_events.csv',
            ]
        extract_func = get_subject_id_from_filename
    elif step_type == 'causalRating':
        if scan_type == 'input':
            patterns = ['*_events.xlsx', '*_events-*.xlsx', '*_events.csv', '*_events-*.csv', '*_events.tsv']
        elif scan_type == 'output':
            patterns = ['*_causal-*.xlsx', '*_causal.xlsx', '*_causal-*.csv', '*_causal.csv']
        else:
            patterns = [
                '*_events.xlsx', '*_events-*.xlsx', '*_events.csv', '*_events-*.csv',
                '*_causal-*.xlsx', '*_causal.xlsx', '*_causal-*.csv',
            ]
        extract_func = get_story_name_from_filename if is_story else get_subject_id_from_filename
    else:
        patterns = ['*.txt', '*.xlsx']
        extract_func = get_story_name_from_filename if is_story else get_subject_id_from_filename

    # Also scan the alternative naming convention ("story-{name}-{token}.{ext}" with
    # hyphenated step tokens), so files like ``story-alice_14_sub-3008-parsed.csv`` and
    # ``story-alice_14-segmented.csv`` are discovered too. The ``*`` covers the
    # ``story-...`` prefix; the extract func strips it back to the canonical item id.
    _alt_patterns = {
        'eventSegment': ['*-segmented.xlsx', '*-segmented.csv', '*-segmented.tsv', '*-segment.xlsx', '*-segment.csv'],
        'textParsing': ['*-parsed.xlsx', '*-parsed.csv', '*-parsed.tsv'],
        'textMatching': ['*-parsed.xlsx', '*-parsed.csv', '*-parsed.tsv',
                         '*-recall-matched.xlsx', '*-recall-matched.csv', '*-recall-matched.tsv'],
        'causalRating': ['*-segmented.xlsx', '*-segmented.csv', '*-segmented.tsv'],
    }.get(step_type, [])
    for _ap in _alt_patterns:
        if _ap not in patterns:
            patterns.append(_ap)

    # Search for files matching patterns
    for pattern in patterns:
        for file in glob_files_in_dir(dir_path, pattern):
            # Skip human-edit files for discovery (they'll be found via original files)
            if not is_user_edit_file(file.name):
                # For textParsing with .txt files, make sure we're getting the right files
                # (exclude audio transcription files that might have similar patterns)
                if step_type == 'textParsing' and pattern == '*.txt':
                    # Only include .txt files that are corrected text files (not audio transcriptions)
                    # Audio transcriptions typically have patterns like "*_recall*.txt" or are in different dirs
                    if '_recall' not in file.name or file.name.startswith('sub') or file.name.startswith('int'):
                        item_id = extract_func(file.name)
                        if item_id:
                            items.add(item_id)
                else:
                    item_id = extract_func(file.name)
                    if item_id:
                        items.add(item_id)
    
    return items


def get_all_subjects(pipeline_config=None):
    """Get list of all available subjects.
    Returns subjects discovered from both default directories and pipeline-configured paths.
    
    Args:
        pipeline_config: Optional pipeline configuration to discover subjects from configured paths
    """
    subjects_set = set()
    subjects_ordered = []  # Preserve order from recall_corrected files
    
    # From corrected files - preserve file order
    if RECALL_CORRECTED_DIR.exists():
        # Get all txt files, sort by modification time to preserve original sequence
        files = list(RECALL_CORRECTED_DIR.glob('*.txt'))
        # Filter out human-edit files for ordering
        regular_files = [f for f in files if not is_user_edit_file(f.name)]
        # Sort by modification time to preserve sequence
        regular_files.sort(key=lambda f: f.stat().st_mtime)
        
        for file in regular_files:
            subj_id = get_subject_id_from_filename(file.name)
            if subj_id not in subjects_set:
                subjects_set.add(subj_id)
                subjects_ordered.append(subj_id)
    
    # Add any subjects from parsed files (canonical, csv, and the alternative
    # "story-...-parsed.csv" naming) that aren't already in the list.
    if RECALL_PARSED_DIR.exists():
        parsed_globs = ['*_parsed.xlsx', '*_parsed.csv', '*_parsed.tsv',
                        '*-parsed.xlsx', '*-parsed.csv', '*-parsed.tsv']
        for pat in parsed_globs:
            for file in RECALL_PARSED_DIR.glob(pat):
                if not is_user_edit_file(file.name):
                    subj_id = get_subject_id_from_filename(file.name)
                    if subj_id and subj_id not in subjects_set:
                        subjects_set.add(subj_id)
                        subjects_ordered.append(subj_id)

    # Add any subjects from rated files (incl. csv and "...-recall-matched.csv").
    if RECALL_RATED_DIR.exists():
        rated_globs = ['*_rate-recall.xlsx', '*_rate-recall-*.xlsx', '*_rate-recall.csv', '*_rate-recall-*.csv',
                       '*-recall-matched.xlsx', '*-recall-matched.csv', '*-recall-matched.tsv']
        for pat in rated_globs:
            for file in RECALL_RATED_DIR.glob(pat):
                if not is_user_edit_file(file.name):
                    subj_id = get_subject_id_from_filename(file.name)
                    if subj_id and subj_id not in subjects_set:
                        subjects_set.add(subj_id)
                        subjects_ordered.append(subj_id)
    
    # If pipeline is configured, also discover subjects from configured input and output paths
    if pipeline_config and pipeline_config.get('steps'):
        for step in pipeline_config['steps']:
            step_type = step_runtime_key(step)
            input_path = step.get('inputPath', '')
            output_path = step.get('outputPath', '')
            
            # Only check subject-focused steps (not story-focused)
            if step_type in ['audioTranscribe:recall', 'sentenceCorrect', 'textParsing', 'textMatching']:
                # Check input path (for steps like textParsing, input contains the source files)
                if input_path:
                    discovered = discover_items_from_path(input_path, step_type, is_story=False, scan_type='input')
                    for subj_id in discovered:
                        if subj_id not in subjects_set:
                            subjects_set.add(subj_id)
                            subjects_ordered.append(subj_id)
                
                # Check output path (for existing results)
                if output_path:
                    discovered = discover_items_from_path(output_path, step_type, is_story=False, scan_type='output')
                    for subj_id in discovered:
                        if subj_id not in subjects_set:
                            subjects_set.add(subj_id)
                            subjects_ordered.append(subj_id)
    
    return subjects_ordered


def get_subject_step_status(subj_id):
    """Get processing step status for a subject.
    Returns dict with 'step0', 'step0_audio', 'step1', 'step2', 'step3' boolean values.
    Checks both original and human-edit versions.
    """
    status = {
        'step0': False,      # Story events segmentation
        'step0_audio': False,  # Audio transcription
        'step1': False,     # Spell and grammar check
        'step2': False,     # Recall text parsing/segmenting
        'step3': False      # Recall rating
    }
    
    # Step 0: Check if story events file exists (original, user-edit, or method-suffixed)
    event_filename_base = f"{subj_id}_events"
    story_original = STORY_EVENTS_DIR / f"{event_filename_base}.xlsx"
    method_files = []
    method_files.extend(list(STORY_EVENTS_DIR.glob(f"{event_filename_base}_rule-based.xlsx")))
    method_files.extend(list(STORY_EVENTS_DIR.glob(f"{event_filename_base}_api.xlsx")))
    method_files.extend(list(STORY_EVENTS_DIR.glob(f"{event_filename_base}-*.xlsx")))
    has_method_file = len(method_files) > 0
    status['step0'] = story_original.exists() or has_any_edit_file(STORY_EVENTS_DIR, subj_id, '_events', '.xlsx') or has_method_file
    
    # Step 0 Audio: Check if audio transcription exists
    audio_dir = RECALL_AUDIO_DIR
    audio_transcribed_dir = RECALL_AUDIO_TRANSCRIBED_DIR
    # Check for audio file (try multiple patterns, all supported formats)
    audio_patterns = []
    for ext in SUPPORTED_AUDIO_EXTENSIONS:
        audio_patterns.extend([f"{subj_id}_recall*{ext}", f"{subj_id}_*{ext}", f"*{subj_id}*{ext}"])
    audio_files = []
    for pattern in audio_patterns:
        audio_files.extend(list(audio_dir.glob(pattern)))
        if audio_files:
            break
    
    if audio_transcribed_dir.exists() and len(audio_files) > 0:
        # Check for transcription file (try multiple patterns)
        transcription_patterns = [
            f"{subj_id}_recall*.txt",
            f"{subj_id}_*.txt",
            f"*{subj_id}*.txt"
        ]
        transcribed_files = []
        for pattern in transcription_patterns:
            transcribed_files.extend(list(audio_transcribed_dir.glob(pattern)))
            if transcribed_files:
                break
        status['step0_audio'] = len(transcribed_files) > 0
    else:
        status['step0_audio'] = False
    
    # Step 1: Check if corrected text exists (original, Gemma method outputs, or any user-edit)
    corrected_original = RECALL_CORRECTED_DIR / f"{subj_id}.txt"
    spell_method_files = list(RECALL_CORRECTED_DIR.glob(f"{subj_id}_spell-*.txt")) if RECALL_CORRECTED_DIR.exists() else []
    status['step1'] = (
        corrected_original.exists()
        or len(spell_method_files) > 0
        or has_any_edit_file(RECALL_CORRECTED_DIR, subj_id, '', '.txt')
    )
    
    # Step 2: Check if parsed file exists (rules or Ollama naming, original or any user-edit)
    parsed_original = RECALL_PARSED_DIR / f"{subj_id}_parsed.xlsx"
    parsed_method_files = (
        list(RECALL_PARSED_DIR.glob(f"{subj_id}_parsed-*.xlsx"))
        if RECALL_PARSED_DIR.exists()
        else []
    )
    parsed_method_files = [f for f in parsed_method_files if not is_user_edit_file(f.name)]
    status['step2'] = (
        parsed_original.exists()
        or len(parsed_method_files) > 0
        or has_any_edit_file(RECALL_PARSED_DIR, subj_id, '_parsed', '.xlsx')
    )
    
    # Step 3: Check if rated file exists (original, method-suffixed, or any user-edit)
    rated_original = RECALL_RATED_DIR / f"{subj_id}_rate-recall.xlsx"
    rated_method_files = list(RECALL_RATED_DIR.glob(f"{subj_id}_rate-recall-*.xlsx")) if RECALL_RATED_DIR.exists() else []
    rated_method_files = [f for f in rated_method_files if not is_user_edit_file(f.name)]
    status['step3'] = rated_original.exists() or has_any_edit_file(RECALL_RATED_DIR, subj_id, '_rate-recall', '.xlsx') or len(rated_method_files) > 0
    
    return status


def get_all_items_with_status():
    """Get all items (subjects or stories) with their processing step statuses based on configured pipeline.
    Discovers items from both default directories and pipeline-configured input/output paths.
    """
    pipeline_config = get_pipeline_config()
    
    # Check if pipeline is story-focused
    is_story_focused = is_story_focused_pipeline(pipeline_config)
    
    if is_story_focused:
        # Get stories - now discovers from all configured paths
        stories = get_all_stories(pipeline_config)
        result = []
        
        for story_name in stories:
            story_data = {'item_id': story_name, 'item_type': 'story'}
            if pipeline_config and pipeline_config.get('steps'):
                for step in pipeline_config['steps']:
                    # Ensure step has an id property
                    step_id = step.get('id')
                    if step_id is None:
                        print(f"Warning: Step missing 'id' field: {step}")
                        continue
                    step_key = f"step_{step_id}"
                    # Check status using step configuration with story name
                    story_data[step_key] = check_step_status(story_name, step, is_story=True)
            result.append(story_data)
        
        return result
    else:
        # Get subjects (recall-focused) - now discovers from all configured paths
        subjects = get_all_subjects(pipeline_config)
        result = []
        
        for subj_id in subjects:
            status = get_subject_step_status(subj_id)
            
            # If pipeline is configured, use it; otherwise use default mapping
            if pipeline_config and pipeline_config.get('steps'):
                subject_data = {'item_id': subj_id, 'item_type': 'subject'}
                for step in pipeline_config['steps']:
                    # Ensure step has an id property
                    step_id = step.get('id')
                    if step_id is None:
                        print(f"Warning: Step missing 'id' field: {step}")
                        continue
                    step_key = f"step_{step_id}"
                    # Check status using step configuration
                    subject_data[step_key] = check_step_status(subj_id, step, is_story=False)
                result.append(subject_data)
            else:
                # Default mapping (backward compatibility)
                result.append({
                    'item_id': subj_id,
                    'item_type': 'subject',
                    'step0': status['step0'],
                    'step0_audio': status['step0_audio'],
                    'step1': status['step1'],
                    'step2': status['step2'],
                    'step3': status['step3']
                })
        
        return result


def get_dashboard_panels():
    """Group pipeline steps into chains: when step A's output path equals step B's input path,
    they form a chain and appear in the same panel as consecutive columns.
    Each chain becomes one dashboard panel (rooted at the first step's input path).
    """
    pipeline_config = get_pipeline_config()
    steps = (pipeline_config or {}).get('steps') or []
    if not steps:
        return {'panels': []}

    def normalize_path(p):
        if not p:
            return ''
        return (str(p).strip().rstrip('/'))

    # Build chains: consecutive steps where output of step i == input of step i+1
    chains = []
    current_chain = []
    for i, step in enumerate(steps):
        if step.get('id') is None:
            continue
        if not current_chain:
            current_chain.append(step)
            continue
        prev = current_chain[-1]
        prev_out = normalize_path(prev.get('outputPath', ''))
        curr_in = normalize_path(step.get('inputPath', ''))
        if prev_out and curr_in and prev_out == curr_in:
            current_chain.append(step)
        else:
            chains.append(current_chain)
            current_chain = [step]
    if current_chain:
        chains.append(current_chain)

    panels = []
    story_step_types = {'audioTranscribe:story', 'eventSegment', 'causalRating'}

    for group_steps in chains:
        if not group_steps:
            continue
        # Root source = first step's input path (start of the chain)
        source_input_path = normalize_path(group_steps[0].get('inputPath', '')) or '(default)'
        is_story = any(step_runtime_key(s) in story_step_types for s in group_steps)
        item_type = 'story' if is_story else 'subject'
        row_label = 'Story Name' if is_story else 'Subject ID'

        # Discover items from EVERY step in the chain (each step's input and output),
        # not just the chain's root input. Otherwise an item that only has downstream
        # files — e.g. a parsed recall (``story-alice_14_sub-3008-parsed.csv``) but no
        # source recall-text — would never appear, so its textMatching step couldn't
        # be reached.
        items_set = set()
        first_step = group_steps[0]
        first_step_type = step_runtime_key(first_step)
        for s in group_steps:
            st = step_runtime_key(s)
            for path, scan in ((s.get('inputPath'), 'input'), (s.get('outputPath'), 'output')):
                path = (path or '').strip().rstrip('/')
                if path:
                    items_set.update(discover_items_from_path(path, st, is_story=is_story, scan_type=scan))
        # Story chains additionally require the story-events reference (unchanged).
        for field, _env_var, extra_dir in _resolve_step_extra_input_dirs(first_step):
            if field == 'storyEventsPath' and not is_story:
                # Subjects come from parsed recall; story-events dir is reference only.
                continue
            discovered = discover_items_from_path(extra_dir, first_step_type, is_story=is_story, scan_type='input')
            if items_set and discovered:
                items_set &= discovered
            elif discovered:
                items_set.update(discovered)

        items_ordered = sorted(items_set)

        items_with_status = []
        for item_id in items_ordered:
            item_data = {'item_id': item_id, 'item_type': item_type}
            for step in group_steps:
                step_id = step.get('id')
                if step_id is None:
                    continue
                item_data[f'step_{step_id}'] = check_step_status(item_id, step, is_story=is_story)
            items_with_status.append(item_data)

        panels.append({
            'sourceInputPath': source_input_path,
            'rowLabel': row_label,
            'itemType': item_type,
            'steps': group_steps,
            'items': items_with_status,
        })

    return {'panels': panels}


def check_story_audio_transcription(item_id, step_config, is_story=False):
    """Check if story audio transcription exists for a subject or story."""
    output_path = step_config.get('outputPath', 'output/story_audio-transcribed')
    # Handle both relative and absolute paths (POSIX, Windows, ~).
    output_dir = _resolve_path_from_config(output_path) or (PROJECT_ROOT / 'output' / output_path)
    
    if not _dir_exists_cached(output_dir):
        return False

    # Try different naming patterns
    patterns = [
        f"{item_id}*.txt",
        f"*{item_id}*.txt"
    ]

    for pattern in patterns:
        if cached_glob(output_dir, pattern):
            return True

    return False


def check_step_status(item_id, step_config, is_story=False):
    """Check status of a pipeline step based on its configuration.
    
    A step is considered "completed" (returns True) only if:
    Output files exist (already processed).
    
    Note: This function only checks for output files to determine completion status.
    Input file existence is checked separately via the input-files API endpoint.
    
    Args:
        item_id: Subject ID or story name
        step_config: Step configuration dict
        is_story: Whether item_id is a story name (True) or subject ID (False)
    
    Returns:
        True if output files exist (step is completed), False otherwise
    """
    step_type = step_config.get('type', '')
    output_path = step_config.get('outputPath', '')
    input_path = step_config.get('inputPath', '')
    
    output_dir = _resolve_path_from_config(output_path)
    input_dir = _resolve_path_from_config(input_path)

    # Flexible recognition is authoritative for these steps: detect an existing
    # output under ANY supported naming convention (subject-first, legacy verbose
    # ``{story}_events_..._rate-recall``, alternative ``story-{name}-segmented`` /
    # ``...-recall-matched``) and ANY tabular format (xlsx/xls/csv/tsv). Cross-step
    # disambiguation in the recogniser also prevents a rated/causal file from being
    # mistaken for a completed events step (the old has_any_edit_file glob did).
    if step_type in ('eventSegment', 'sentenceCorrect', 'textParsing', 'textMatching', 'causalRating'):
        try:
            from helpers.step_files import has_step_output
            return bool(output_dir) and has_step_output(output_dir, step_type, item_id, is_story=is_story)
        except Exception:
            pass  # only on unexpected error fall through to the legacy per-step checks

    if step_type == 'audioTranscribe:story':
        # For story audio, check transcription output
        if output_dir and _dir_exists_cached(output_dir):
            # Story name might be in filename (e.g., my_story.wav -> my_story.txt)
            patterns = [f"{item_id}*.txt", f"*{item_id}*.txt"]
            for pattern in patterns:
                if cached_glob(output_dir, pattern):
                    return True
        return False
    elif step_type == 'audioTranscribe:recall':
        # Check recall audio transcription
        if output_dir and _dir_exists_cached(output_dir):
            patterns = [f"{item_id}*.txt", f"*{item_id}*.txt"]
            for pattern in patterns:
                if cached_glob(output_dir, pattern):
                    return True
        return False
    elif step_type == 'eventSegment':
        # Check story events file
        if output_dir and _dir_exists_cached(output_dir):
            # Story events file naming: {story_name}_events.xlsx or {story_name}_events-{method}.xlsx
            event_file = output_dir / f"{item_id}_events.xlsx"
            method_files = cached_glob(output_dir, f"{item_id}_events-*.xlsx")
            method_files += cached_glob(output_dir, f"{item_id}_events_rule-based.xlsx")
            method_files += cached_glob(output_dir, f"{item_id}_events_api.xlsx")
            return cached_path_exists(event_file) or has_any_edit_file(output_dir, item_id, '_events', '.xlsx') or len(method_files) > 0
        return False
    elif step_type == 'sentenceCorrect':
        # Check corrected text (rule-based file or Gemma method outputs)
        if output_dir and _dir_exists_cached(output_dir):
            corrected_file = output_dir / f"{item_id}.txt"
            if cached_path_exists(corrected_file) or has_any_edit_file(output_dir, item_id, '', '.txt'):
                return True
            if cached_glob(output_dir, f"{item_id}_spell-*.txt"):
                return True

            # Also try glob patterns in case item_id doesn't match exactly
            patterns = [
                f"{item_id}.txt",
                f"{item_id}_*-edit.txt",
                f"{item_id}*.txt",
                f"*{item_id}*.txt"
            ]
            for pattern in patterns:
                matching_files = cached_glob(output_dir, pattern)
                # Filter out human-edit files if we're looking for original
                if '-edit' not in pattern:
                    matching_files = [f for f in matching_files if not is_user_edit_file(f.name)]
                # Also exclude files that are clearly not the right ones (e.g., audio transcriptions)
                matching_files = [f for f in matching_files if '_recall' not in f.name or f.name.startswith(item_id)]
                if matching_files:
                    return True
        return False
    elif step_type == 'textParsing':
        # Check parsed file (output)
        if output_dir and _dir_exists_cached(output_dir):
            # Try exact match first in main directory
            parsed_file = output_dir / f"{item_id}_parsed.xlsx"
            if cached_path_exists(parsed_file) or has_any_edit_file(output_dir, item_id, '_parsed', '.xlsx'):
                return True

            # Also check in _prev subdirectory (backup/archive location)
            prev_dir = output_dir / '_prev'
            prev_exists = _dir_exists_cached(prev_dir)
            if prev_exists:
                prev_parsed_file = prev_dir / f"{item_id}_parsed.xlsx"
                if cached_path_exists(prev_parsed_file) or has_any_edit_file(prev_dir, item_id, '_parsed', '.xlsx'):
                    return True

            # Also try glob patterns in case item_id doesn't match exactly
            patterns = [
                f"{item_id}_parsed.xlsx",
                f"{item_id}_parsed-*.xlsx",
                f"{item_id}_parsed_*-edit.xlsx",
                f"{item_id}*_parsed*.xlsx",
                f"*{item_id}*_parsed*.xlsx"
            ]
            for pattern in patterns:
                matching_files = cached_glob(output_dir, pattern)
                # Also check in _prev subdirectory
                if prev_exists:
                    matching_files = matching_files + cached_glob(prev_dir, pattern)
                # Filter out human-edit files if we're looking for original
                if '-edit' not in pattern:
                    matching_files = [f for f in matching_files if not is_user_edit_file(f.name)]
                if matching_files:
                    return True

            # Last resort: check if any _parsed.xlsx file contains the item_id
            # This handles cases where the filename format might be slightly different
            all_parsed_files = cached_glob(output_dir, '*_parsed*.xlsx')
            if prev_exists:
                all_parsed_files = all_parsed_files + cached_glob(prev_dir, '*_parsed*.xlsx')
            for file in all_parsed_files:
                if not is_user_edit_file(file.name):
                    # Extract subject ID from filename and compare
                    file_subj_id = get_subject_id_from_filename(file.name)
                    if file_subj_id == item_id:
                        return True

        # Only return True if output files exist - don't check input files here
        # Input file existence is checked separately via the input-files API endpoint
        # This function should only indicate completion (output exists), not availability (input exists)
        return False
    elif step_type == 'textMatching':
        # Check rated file (including method-suffixed variants like _rate-recall-api.xlsx)
        if output_dir and _dir_exists_cached(output_dir):
            # Try exact match first
            rated_file = output_dir / f"{item_id}_rate-recall.xlsx"
            if cached_path_exists(rated_file) or has_any_edit_file(output_dir, item_id, '_rate-recall', '.xlsx'):
                return True

            # Also try glob patterns including method-suffixed files
            patterns = [
                f"{item_id}_rate-recall.xlsx",
                f"{item_id}_rate-recall-*.xlsx",
                f"{item_id}_rate-recall_*-edit.xlsx",
                f"{item_id}*_rate-recall.xlsx",
                f"{item_id}*_rate-recall-*.xlsx",
                f"*{item_id}*_rate-recall.xlsx",
                f"*{item_id}*_rate-recall-*.xlsx"
            ]
            for pattern in patterns:
                matching_files = [f for f in cached_glob(output_dir, pattern) if not is_user_edit_file(f.name)]
                if matching_files:
                    return True
        return False
    elif step_type == 'causalRating':
        if output_dir and _dir_exists_cached(output_dir):
            patterns = [
                f"{item_id}_causal-*.xlsx",
                f"{item_id}_causal.xlsx",
                f"*{item_id}*_causal-*.xlsx",
                f"*{item_id}*_causal.xlsx",
            ]
            for pattern in patterns:
                matching_files = [f for f in cached_glob(output_dir, pattern) if not is_user_edit_file(f.name)]
                if matching_files:
                    return True
        return False
    
    return False


def _read_raw_recall_from_excel(excel_path, subj_id):
    """Extract raw recall for subj_id from an Excel file. Returns str or None."""
    try:
        for sheet_name in ['all', 'Sheet1']:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                if 'recall' not in df.columns or 'sub' not in df.columns or 'ID' not in df.columns:
                    continue
                for _, row in df.iterrows():
                    sub = row.get('sub', '')
                    sub_id = row.get('ID', '')
                    recall = row.get('recall', '')
                    if pd.isna(sub) or pd.isna(sub_id) or pd.isna(recall):
                        continue
                    sub_str = str(int(sub)) if isinstance(sub, (int, float)) else str(sub).strip()
                    sub_id_str = str(int(sub_id)) if isinstance(sub_id, (int, float)) else str(sub_id).strip()
                    filename_base = f"{sub_str}_{sub_id_str}" if sub_str.startswith('sub') else f"sub{sub_str}_{sub_id_str}"
                    if filename_base == subj_id:
                        recall_text = str(recall)
                        if recall_text.startswith('"') and recall_text.endswith('"'):
                            recall_text = recall_text[1:-1]
                        if recall_text.startswith("'") and recall_text.endswith("'"):
                            recall_text = recall_text[1:-1]
                        return recall_text
            except Exception:
                continue
    except Exception:
        pass
    return None


def get_raw_recall_text(subj_id):
    """Get raw recall text from sentenceCorrect input path (.txt files) or Excel files.
    Uses pipeline config inputPath for sentenceCorrect when configured.
    """
    # 1. Try sentenceCorrect input path
    input_path = get_input_dir_for_step_type('sentenceCorrect')
    if input_path and input_path.exists():
        if input_path.is_dir():
            # Raw input is typically subj_id.txt (spell-grammar excludes user-edit files as input)
            for candidate in [f"{subj_id}.txt"]:
                file_path = input_path / candidate
                if file_path.exists() and file_path.is_file():
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            return f.read().strip()
                    except Exception as e:
                        print(f"Error reading raw recall from {file_path}: {e}")
        elif input_path.is_file() and input_path.suffix.lower() in ('.xlsx', '.xls'):
            result = _read_raw_recall_from_excel(input_path, subj_id)
            if result is not None:
                return result

    # 2. Try default .txt directory
    default_input = PROJECT_ROOT / 'data' / '5_recall_texts'
    if default_input.exists() and default_input.is_dir():
        file_path = default_input / f"{subj_id}.txt"
        if file_path.exists():
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                print(f"Error reading raw recall from {file_path}: {e}")

    # 3. Fall back to hardcoded Excel files
    for excel_file in EXCEL_FILES:
        if os.path.exists(excel_file):
            result = _read_raw_recall_from_excel(excel_file, subj_id)
            if result is not None:
                return result

    return None


def _parse_recall_txt_file(file_path):
    """Parse a recall .txt file (corrected two-line header or plain raw text)."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        if len(lines) >= 2 and lines[0].strip().endswith('.txt'):
            text = ''.join(lines[1:]).strip()
        elif lines:
            text = ''.join(lines).strip()
        else:
            return None
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1]
        if text.startswith("'") and text.endswith("'"):
            text = text[1:-1]
        return text or None
    except Exception as e:
        print(f"Error reading recall text from {file_path}: {e}")
        return None


def _read_recall_text_from_directory(directory, subj_id):
    """Read recall text for *subj_id* from a directory of .txt files."""
    dir_path = _resolve_path_from_config(directory) if isinstance(directory, str) else directory
    if not dir_path or not Path(dir_path).exists() or not Path(dir_path).is_dir():
        return None
    dir_path = Path(dir_path)
    for pattern in (f"{subj_id}.txt", f"{subj_id}_spell-*.txt", f"{subj_id}*.txt", f"*{subj_id}*.txt"):
        matches = [f for f in dir_path.glob(pattern) if f.is_file() and not is_user_edit_file(f.name)]
        if matches:
            matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            text = _parse_recall_txt_file(matches[0])
            if text:
                return text
    return None


def _resolve_recall_text_for_manual(subj_id, input_dir=None):
    """Best-effort recall text for manual scaffold steps (textParsing, textMatching, …).

    Manual mode should not require an upstream automated step to have run first.
    Try corrected output, this step's configured input folder, raw recall inputs,
    then the conventional default folders.
    """
    text = get_corrected_text(subj_id)
    if text:
        return text
    text = _read_recall_text_from_directory(input_dir, subj_id)
    if text:
        return text
    text = get_raw_recall_text(subj_id)
    if text:
        return text
    text = _read_recall_text_from_directory(get_output_dir_for_step_type('sentenceCorrect'), subj_id)
    if text:
        return text
    return _read_recall_text_from_directory(RECALL_CORRECTED_DIR, subj_id)


def _resolve_version_stem(directory, file_version):
    """If ``file_version`` is an explicit filename stem (a method variant chosen in
    the version dropdown, not '', 'original', or a '-edit' token), return the matching
    file in ``directory`` across supported extensions, else None."""
    if not file_version or file_version == 'original' or str(file_version).endswith('-edit'):
        return None
    for ext in ('.xlsx', '.csv', '.tsv', '.xls', '.txt'):
        cand = Path(directory) / f"{file_version}{ext}"
        if cand.exists():
            return cand
    return None


def get_corrected_text(subj_id, file_version=None):
    """Get corrected recall text.
    Prioritizes user-edit version if available.
    Uses pipeline config outputPath for sentenceCorrect when configured.
    file_version: '{username}-edit', 'original', explicit stem, or None (auto-select)
    """
    output_dir = get_output_dir_for_step_type('sentenceCorrect') or RECALL_CORRECTED_DIR
    explicit = _resolve_version_stem(output_dir, file_version)
    # Determine which file to use
    if explicit is not None:
        file_path = explicit
    elif file_version and file_version.endswith('-edit'):
        file_path = output_dir / f"{subj_id}_{file_version}.txt"
    elif file_version == 'original':
        canonical = output_dir / f"{subj_id}.txt"
        if canonical.exists():
            file_path = canonical
        else:
            cands = list_corrected_recall_source_files(output_dir, subj_id)
            file_path = cands[0] if cands else canonical
    else:
        # Auto-select: prioritize user-edit, fallback to canonical rules file, then Gemma outputs (newest)
        edit_file = find_best_edit_file(output_dir, subj_id, '', '.txt')
        canonical = output_dir / f"{subj_id}.txt"
        cands = list_corrected_recall_source_files(output_dir, subj_id)
        if edit_file:
            file_path = edit_file
        elif canonical.exists():
            file_path = canonical
        elif cands:
            file_path = max(cands, key=lambda p: p.stat().st_mtime)
        else:
            file_path = canonical
    
    if not file_path.exists():
        return None
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        if len(lines) < 2:
            return None
        
        # First line is filename, rest is text
        text = '\n'.join(lines[1:]).strip()
        
        # Remove surrounding quotes if present
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1]
        if text.startswith("'") and text.endswith("'"):
            text = text[1:-1]
        
        return text
    except Exception as e:
        print(f"Error reading corrected text: {e}")
        return None


def get_parsed_texts(subj_id, file_version=None):
    """Get parsed recall texts as list of segments.
    Prioritizes user-edit version if available.
    Uses pipeline config outputPath for textParsing when configured.
    file_version: '{username}-edit', 'original', or None (auto-select)
    """
    output_dir = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
    # Non-edit parsed source files (canonical and method-suffixed), newest first
    non_edit_parsed = list_subject_parsed_source_files(output_dir, subj_id)
    canonical = output_dir / f"{subj_id}_parsed.xlsx"
    # Explicit method-variant filename stem from the version dropdown (e.g.
    # ``the_siren_sub-01_parsed-ollama_gemma4_e4b``): load that exact file.
    explicit = _resolve_version_stem(output_dir, file_version)
    # Determine which file to use
    if explicit is not None:
        file_path = explicit
    elif file_version and file_version.endswith('-edit'):
        file_path = output_dir / f"{subj_id}_parsed_{file_version}.xlsx"
    elif file_version == 'original':
        if canonical.exists():
            file_path = canonical
        elif non_edit_parsed:
            file_path = non_edit_parsed[0]
        else:
            file_path = canonical
    else:
        # Auto-select: prioritize user-edit, fallback to canonical, then method-specific outputs (newest)
        edit_file = find_best_edit_file(output_dir, subj_id, '_parsed', '.xlsx')
        if edit_file:
            file_path = edit_file
        elif canonical.exists():
            file_path = canonical
        elif non_edit_parsed:
            file_path = non_edit_parsed[0]
        else:
            file_path = canonical

    if not file_path.exists():
        # Fallback: recognise alternative-named / csv parsed files that the canonical
        # ``{subj}_parsed.xlsx`` lookups miss, e.g. ``story-alice_14_sub-3008-parsed.csv``.
        try:
            from helpers.step_files import find_step_files
            recs = find_step_files(output_dir, 'textParsing', subj_id)
            if file_version and file_version.endswith('-edit'):
                edits = [f for f in recs if is_user_edit_file(f.name)]
                recs = edits or recs
            if recs:
                file_path = recs[0]
        except Exception:
            pass

    if not file_path.exists():
        return None

    try:
        from helpers.flexible_io import read_parsed_recall_file
        df = read_parsed_recall_file(file_path)

        segments = []
        for _, row in df.iterrows():
            parsed = row.get('recall_in_temporal_order', '')
            matched = row.get('recalled_events', '')
            
            if pd.notna(parsed):
                segments.append({
                    'text': str(parsed).strip(),
                    'recalled_events': str(matched).strip() if pd.notna(matched) else ''
                })
        
        return segments
    except Exception as e:
        print(f"Error reading parsed texts: {e}")
        return None


# Per-segment "further ratings" columns, in display/export order. Must match the
# frontend RATING_KEYS list in templates/subject.html.
FURTHER_RATING_COLS = ('summary', 'error', 'confabulation', 'opinion', 'inference', 'meta')

# Benchmark CSV column names. The unrated benchmark files carry a richer schema
# than the canonical two-column rated files: confidence (per pass) + (when the
# rater has started) two per-segment progress flags. The frontend rating key
# `error` is stored under the file column `factual_error`; every other key maps
# to itself. Both passes rate confidence: the first pass writes `confidence_1`,
# the second pass writes `confidence`; the `comment` column is shared.
BENCH_CONFIDENCE_COL = 'confidence'
BENCH_CONFIDENCE1_COL = 'confidence_1'
BENCH_FIRST_PASS_COL = 'first-pass-rated'
BENCH_SECOND_PASS_COL = 'second-pass-rated'
_BENCH_RATING_FILE_COL = {'error': 'factual_error'}

# Last-accepted client save sequence per rated file path (in-memory; resets on
# server restart). Used to drop stale out-of-order benchmark saves (#6).
_BENCH_SAVE_SEQ = {}

# Per-rated-path lock so the staleness check -> read seed -> build df -> atomic
# write -> _BENCH_SAVE_SEQ update sequence runs atomically. Without it, two
# concurrent saves for the same file (debounced autosave racing the
# beforeunload beacon, or two tabs) can both pass the staleness check before
# either updates _BENCH_SAVE_SEQ, and the one that os.replace()s last wins
# regardless of which carried the newer client_seq (lost update). Matters under
# --production, where Waitress serves requests on multiple threads.
_BENCH_SAVE_LOCKS = {}
_BENCH_SAVE_LOCKS_GUARD = threading.Lock()


def _bench_save_lock(seq_key):
    """Return the (process-wide) lock for a rated path, creating it on first use."""
    with _BENCH_SAVE_LOCKS_GUARD:
        lock = _BENCH_SAVE_LOCKS.get(seq_key)
        if lock is None:
            lock = threading.Lock()
            _BENCH_SAVE_LOCKS[seq_key] = lock
        return lock


def _bench_file_col(key):
    """Map a frontend rating key to its benchmark-CSV column name."""
    return _BENCH_RATING_FILE_COL.get(key, key)


def _parse_rating_cell(value):
    """Inverse of the frontend rating cell: turn a saved "further ratings" cell
    into ``(checked, spans)``. A bare ``TRUE`` means checked with no text spans;
    quoted fragments (``'a'; 'b'``) mean checked with those spans; blank means
    unchecked."""
    text = str(value or '').strip()
    if not text or text.lower() == 'nan':
        return False, []
    if text.lower() in ('true', '1', 'yes', 'x'):
        return True, []
    frags = [m.strip() for m in re.findall(r"'([^']*)'", text)]
    frags = [f for f in frags if f]
    if not frags:
        # Legacy / unquoted content: treat the whole cell as a single fragment.
        frags = [text]
    return True, frags


# Benchmark-only highlight cells encode spans as integer offset ranges into the
# segment's recall_text, e.g. '12-47' or '12-47;50-80' (one range per highlight).
# This is intentionally NOT the shared _parse_rating_cell format above (which the
# non-benchmark span editor uses); the two paths diverge by design.
_BENCH_RANGE_CELL_RE = re.compile(r'^\s*\d+\s*-\s*\d+(\s*;\s*\d+\s*-\s*\d+)*\s*$')


def _parse_benchmark_rating_cell(value):
    """(checked, ranges) where ranges = [{'start','end'}, ...] offsets into
    recall_text. Only the 'a-b;c-d' format is understood."""
    s = str(value or '').strip()
    if not s or s.lower() == 'nan':
        return False, []
    if s.lower() in ('true', '1', 'yes', 'x'):
        return True, []
    if _BENCH_RANGE_CELL_RE.match(s):
        ranges = []
        for pair in s.split(';'):
            a, b = pair.split('-')
            start, end = int(a), int(b)
            if 0 <= start < end:
                ranges.append({'start': start, 'end': end})
        return True, ranges
    # Unrecognized non-empty cell: checked, no locatable span (don't lose it).
    return True, []


def _format_benchmark_rating_cell(checked, ranges):
    """'a-b;c-d' for ranges; 'TRUE' if checked with none; '' if unchecked."""
    if not checked:
        return ''
    pairs = [f"{int(r['start'])}-{int(r['end'])}"
             for r in (ranges or [])
             if r and r.get('start') is not None and r.get('end') is not None]
    return ';'.join(pairs) if pairs else 'TRUE'


def get_rated_texts(subj_id, file_version=None):
    """Get rated recall texts with event matches.
    Prioritizes user-edit version if available, otherwise shows original.
    Uses pipeline config outputPath for textMatching when configured.
    file_version: '{username}-edit', 'original', or None (auto-select)
    """
    output_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
    
    def _find_rated_file(subj, suffix=''):
        """Find rated file, checking method-suffixed variants if exact match missing."""
        exact = output_dir / f"{subj}_rate-recall{suffix}.xlsx"
        if exact.exists():
            return exact
        if not suffix:
            candidates = list_subject_rated_recall_source_files(output_dir, subj)
            if candidates:
                return candidates[0]
        return exact
    
    # Determine which rated file to use
    explicit_rated = _resolve_version_stem(output_dir, file_version)
    if explicit_rated is not None:
        rated_file = explicit_rated
    elif file_version and file_version.endswith('-edit'):
        rated_file = output_dir / f"{subj_id}_rate-recall_{file_version}.xlsx"
    elif file_version == 'original':
        rated_file = _find_rated_file(subj_id)
    else:
        # Auto-select: prioritize user-edit, fallback to original (including method-suffixed)
        edit_file = find_best_edit_file(output_dir, subj_id, '_rate-recall', '.xlsx')
        if edit_file:
            rated_file = edit_file
        else:
            rated_file = _find_rated_file(subj_id)

    if not rated_file.exists():
        # Fallback: recognise alternative-named / csv rated files (e.g.
        # ``story-alice_14_sub-3008-recall-matched.csv``).
        try:
            from helpers.step_files import find_step_files
            recs = find_step_files(output_dir, 'textMatching', subj_id)
            if file_version and file_version.endswith('-edit'):
                edits = [f for f in recs if is_user_edit_file(f.name)]
                recs = edits or recs
            if recs:
                rated_file = recs[0]
        except Exception:
            pass

    # If rated file doesn't exist, return None
    if not rated_file.exists():
        return None
    
    # Read rated file directly
    try:
        from helpers.flexible_io import read_parsed_recall_file, read_tabular
        df = read_parsed_recall_file(rated_file)
        # normalize_parsed_recall_df keeps only the two core columns, so read the
        # raw sheet too to recover the optional "further ratings" columns (row
        # order is preserved, so positional alignment with df is safe).
        try:
            raw_df = read_tabular(rated_file)
            raw_df = raw_df.reset_index(drop=True)
        except Exception:
            raw_df = None
        rating_cols = [c for c in FURTHER_RATING_COLS
                       if raw_df is not None and c in raw_df.columns]
        has_comment_col = raw_df is not None and 'comment' in raw_df.columns

        segments = []
        for pos, (_, row) in enumerate(df.iterrows()):
            parsed = row.get('recall_in_temporal_order', '')
            matched = row.get('recalled_events', '')
            
            if pd.notna(parsed):
                # Get matched events
                matched_str = ''
                if pd.notna(matched) and str(matched).strip() not in ['', 'nan']:
                    matched_str = str(matched).strip()
                    # Clean up - ensure integers, no floats
                    if matched_str:
                        try:
                            if ',' in matched_str:
                                # Handle comma-separated values (with or without spaces)
                                parts = []
                                for part in matched_str.split(','):
                                    part = part.strip()
                                    if part:
                                        parts.append(str(int(float(part))))
                                matched_str = ','.join(parts)
                            else:
                                matched_str = str(int(float(matched_str)))
                        except (ValueError, AttributeError):
                            # If parsing fails, keep original string
                            pass
                
                seg = {
                    'text': str(parsed).strip(),
                    'matched_event': matched_str  # Frontend expects 'matched_event', not 'recalled_events'
                }
                # Surface the optional "further ratings" columns so the checkboxes
                # (and any per-rating text fragments) reload in their saved state.
                if (rating_cols or has_comment_col) and pos < len(raw_df):
                    raw_row = raw_df.iloc[pos]
                    for rating in rating_cols:
                        checked, spans = _parse_rating_cell(raw_row.get(rating, ''))
                        seg[rating] = checked
                        seg[rating + '_spans'] = spans
                    if has_comment_col:
                        cv = raw_row.get('comment', '')
                        seg['comment'] = '' if (cv is None or str(cv).strip().lower() == 'nan') else str(cv).strip()
                segments.append(seg)
        
        print(f"Loaded {len(segments)} segments from {rated_file.name}")
        return segments
    except Exception as e:
        print(f"Error reading rated texts from {rated_file}: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_available_file_versions(subj_id, is_story=False):
    """Get available file versions for a subject.
    Returns dict with 'step1', 'step2', 'step3' lists of available versions,
    plus 'story_events' with all event segmentation output files and
    'causal' with all causal rating output files,
    and 'story_transcript' with transcript files under data/2_story_transcript.
    Uses pipeline config outputPath when configured.
    """
    versions = {
        'step1': [],
        'step2': [],
        'step3': [],
        'story_events': [],
        'causal': [],
        'story_transcript': [],
    }
    
    # Step 1: Check corrected text files (sentenceCorrect output)
    corrected_dir = get_output_dir_for_step_type('sentenceCorrect') or RECALL_CORRECTED_DIR
    original_corrected = corrected_dir / f"{subj_id}.txt"
    
    edit_versions_1 = get_all_edit_versions(corrected_dir, subj_id, '', '.txt')
    versions['step1'].extend(edit_versions_1)
    if original_corrected.exists():
        versions['step1'].append('original')
    
    # Method-variant files (e.g. ``{subj}_parsed-ollama_gemma4_e4b.xlsx``) are surfaced
    # as individually-selectable versions, keyed by their filename stem, so the same
    # subject's different processing methods can be viewed in the inspection page.
    from helpers.step_files import find_step_files

    def _method_variant_tokens(directory, step_type, canonical_stems):
        toks = []
        for f in find_step_files(directory, step_type, subj_id):
            if is_user_edit_file(f.name):
                continue
            if f.stem in canonical_stems:
                continue  # the canonical file is offered as 'original'
            toks.append(f.stem)
        # newest first already (find_step_files sorts by mtime); de-dupe, keep order
        seen = set()
        return [t for t in toks if not (t in seen or seen.add(t))]

    # Step 1: corrected
    versions['step1'].extend(_method_variant_tokens(corrected_dir, 'sentenceCorrect', {subj_id}))

    # Step 2: Check parsed files (textParsing output)
    parsed_dir = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
    original_parsed = parsed_dir / f"{subj_id}_parsed.xlsx"
    method_parsed = list_subject_parsed_source_files(parsed_dir, subj_id)

    edit_versions_2 = get_all_edit_versions(parsed_dir, subj_id, '_parsed', '.xlsx')
    versions['step2'].extend(edit_versions_2)
    if original_parsed.exists() or method_parsed:
        versions['step2'].append('original')
    versions['step2'].extend(_method_variant_tokens(parsed_dir, 'textParsing', {f"{subj_id}_parsed"}))

    # Step 3: Check rated files (textMatching output, including method-suffixed)
    rated_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
    original_rated = rated_dir / f"{subj_id}_rate-recall.xlsx"
    method_rated = list_subject_rated_recall_source_files(rated_dir, subj_id)

    edit_versions_3 = get_all_edit_versions(rated_dir, subj_id, '_rate-recall', '.xlsx')
    versions['step3'].extend(edit_versions_3)
    if original_rated.exists() or method_rated:
        versions['step3'].append('original')
    versions['step3'].extend(_method_variant_tokens(rated_dir, 'textMatching', {f"{subj_id}_rate-recall"}))
    
    # Story events: list segmentation files from configured dir and default data/3_story_events
    seen_names = set()
    event_files = []
    for events_dir in iter_story_events_search_dirs():
        if not events_dir or not events_dir.exists():
            continue
        for base in expand_story_event_file_bases(subj_id):
            for ef in events_dir.glob(f"{base}_events*.xlsx"):
                if ef.is_file() and ef.name not in seen_names:
                    seen_names.add(ef.name)
                    event_files.append(ef)
    event_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for ef in event_files:
        versions['story_events'].append({
            'filename': ef.name,
            'label': _format_event_file_label(ef.name, subj_id),
            'is_edit': is_user_edit_file(ef.name),
        })
    
    # Causal ratings: list all causal rating output files
    causal_dir = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
    if causal_dir and causal_dir.exists():
        causal_files = sorted(causal_dir.glob(f"{subj_id}_causal*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        for cf in causal_files:
            versions['causal'].append({
                'filename': cf.name,
                'label': _format_causal_file_label(cf.name, subj_id),
                'is_edit': is_user_edit_file(cf.name),
            })
    
    versions['story_transcript'] = get_story_transcript_catalog(subj_id, is_story=is_story)
    
    return versions


def _format_event_file_label(filename, item_id):
    """Create a human-readable label from an event file name."""
    stem = Path(filename).stem
    suffix = stem.replace(f'{item_id}_events', '')
    if not suffix:
        return 'Events (default)'
    suffix = suffix.lstrip('-').lstrip('_')
    if not suffix:
        return 'Events (default)'
    edit_user = extract_edit_username(filename)
    if edit_user:
        base = suffix.replace(f'_{edit_user}-edit', '').replace('_', ' ').replace('-', ' ').strip()
        return f'{base} ({edit_user} edit)' if base else f'{edit_user} (edit)'
    label_map = {
        'fine': 'Fine-grained',
        'coarse': 'Coarse-grained',
    }
    key = suffix.lower()
    if key in label_map:
        return label_map[key]
    if key.startswith('api'):
        model_part = key.replace('api_', '').replace('api', '').strip()
        return f'API ({model_part})' if model_part else 'API'
    trial_match = re.match(r'^(.+?)_trial(\d+)$', key)
    if trial_match:
        base_label = label_map.get(trial_match.group(1), trial_match.group(1).replace('_', ' ').title())
        return f'{base_label} (trial {trial_match.group(2)})'
    return suffix.replace('_', ' ').replace('-', ' ').title()


def _format_causal_file_label(filename, item_id):
    """Create a human-readable label from a causal rating file name."""
    stem = Path(filename).stem
    suffix = stem.replace(f'{item_id}_causal', '')
    if not suffix:
        return 'Causal (default)'
    suffix = suffix.lstrip('-').lstrip('_')
    if not suffix:
        return 'Causal (default)'
    edit_user = extract_edit_username(filename)
    if edit_user:
        base = suffix.replace(f'_{edit_user}-edit', '').replace('_', ' ').replace('-', ' ').strip()
        return f'{base} ({edit_user} edit)' if base else f'{edit_user} (edit)'
    label_map = {
        'manual': 'Manual',
        'linguistic': 'Linguistic',
    }
    key = suffix.lower()
    if key in label_map:
        return label_map[key]
    if key.startswith('api'):
        model_part = key.replace('api_', '').replace('api', '').strip()
        return f'API ({model_part})' if model_part else 'API'
    trial_match = re.match(r'^(.+?)_trial(\d+)$', key)
    if trial_match:
        base_label = label_map.get(trial_match.group(1), trial_match.group(1).replace('_', ' ').title())
        return f'{base_label} (trial {trial_match.group(2)})'
    return suffix.replace('_', ' ').replace('-', ' ').title()


def _story_transcript_glob_patterns(item_id, is_story=False):
    """Glob patterns for transcript files under STORY_TRANSCRIPT_DIR (order = auto-pick priority)."""
    if is_story:
        return [
            f"{item_id}.txt",
            f"{item_id}*.txt",
            f"{item_id}*.csv",
            f"{item_id}*.tsv",
            f"{item_id}*.xlsx",
            f"*{item_id}*.txt",
            f"*{item_id}*.csv",
            f"*{item_id}*.tsv",
            f"*{item_id}*.xlsx",
        ]
    return [
        f"{item_id}*.txt",
        f"{item_id}*.csv",
        f"{item_id}*.tsv",
        f"{item_id}*.xlsx",
        f"*{item_id}*.txt",
        f"*{item_id}*.csv",
        f"*{item_id}*.tsv",
        f"*{item_id}*.xlsx",
    ]


def collect_story_transcript_paths(item_id, is_story=False):
    """Unique transcript paths under STORY_TRANSCRIPT_DIR matching this item (resolved paths)."""
    transcript_dir = STORY_TRANSCRIPT_DIR
    if not transcript_dir.exists():
        return []
    seen = set()
    ordered = []
    for pattern in _story_transcript_glob_patterns(item_id, is_story):
        for fp in transcript_dir.glob(pattern):
            if not fp.is_file():
                continue
            key = fp.resolve()
            if key not in seen:
                seen.add(key)
                ordered.append(fp)
    return ordered


def _read_story_transcript_file(file_path):
    """Read full text from transcript files (.txt, .csv, .tsv, .xlsx)."""
    try:
        from helpers.flexible_io import read_document_text
        return read_document_text(file_path) or None
    except Exception as e:
        print(f"Error reading story transcript from file: {e}")
    return None


def get_story_transcript_catalog(item_id, is_story=False):
    """List selectable story transcript files for the inspect UI (newest first)."""
    paths = collect_story_transcript_paths(item_id, is_story)
    paths.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return [{'filename': p.name, 'label': p.name} for p in paths]


def get_story_transcript(item_id, is_story=False, transcript_file=None):
    """Get full story transcript (from transcript file or reconstructed from events).
    item_id: Subject ID or story name
    is_story: Whether item_id is a story name (True) or subject ID (False)
    transcript_file: If set, load this basename from STORY_TRANSCRIPT_DIR when it matches item patterns.
    """
    transcript_dir = STORY_TRANSCRIPT_DIR
    if transcript_file and transcript_dir.exists():
        raw = str(transcript_file).strip()
        if raw and '/' not in raw and '\\' not in raw and '..' not in raw:
            safe_name = Path(raw).name
            if safe_name == raw:
                allowed = {p.name for p in collect_story_transcript_paths(item_id, is_story)}
                if safe_name in allowed:
                    fp = transcript_dir / safe_name
                    if fp.is_file():
                        text = _read_story_transcript_file(fp)
                        if text:
                            return text

    # Auto: same priority as before — first pattern match, first glob result
    if transcript_dir.exists():
        for pattern in _story_transcript_glob_patterns(item_id, is_story):
            files = list(transcript_dir.glob(pattern))
            if files:
                try:
                    file_path = files[0]
                    text = _read_story_transcript_file(file_path)
                    if text:
                        return text
                except Exception as e:
                    print(f"Error reading story transcript from file: {e}")
    
    # Fallback: Try to reconstruct from events file
    events_dir = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
    if is_story:
        event_filename_base = f"{item_id}_events"
    else:
        event_filename_base = f"{item_id}_events"
    
    # Try standard file first
    file_path = events_dir / f"{event_filename_base}.xlsx"
    if not file_path.exists():
        # Try method-suffixed files (e.g., _events_rule-based.xlsx, _events_api.xlsx, _events-*.xlsx)
        method_files = []
        method_files.extend(list(events_dir.glob(f"{event_filename_base}_rule-based.xlsx")))
        method_files.extend(list(events_dir.glob(f"{event_filename_base}_api.xlsx")))
        method_files.extend(list(events_dir.glob(f"{event_filename_base}-*.xlsx")))
        if method_files:
            # Use the most recently modified one
            file_path = max(method_files, key=lambda p: p.stat().st_mtime)
    
    if file_path.exists():
        try:
            df = pd.read_excel(file_path)
            if 'story_texts' in df.columns:
                # Reconstruct full transcript from all events
                full_text = ' '.join(df['story_texts'].dropna().astype(str))
                return full_text.strip() if full_text else None
        except Exception as e:
            print(f"Error reading story transcript from events: {e}")
    
    # For subjects, try again using the story name derived from the subject ID
    if not is_story:
        story_name = get_story_name_from_subject_id(item_id)
        if story_name and story_name != item_id:
            result = get_story_transcript(story_name, is_story=True)
            if result:
                return result
    
    return None


def _read_story_events_dataframe(df):
    """Build event list from a story-events DataFrame; accepts story_texts or story_text (case-insensitive)."""
    if df is None or getattr(df, 'empty', True):
        return None
    cols_lower = {str(c).strip().lower(): c for c in df.columns}
    event_col = cols_lower.get('event')
    text_col = cols_lower.get('story_texts') or cols_lower.get('story_text')
    if not event_col or not text_col:
        return None
    events = []
    for _, row in df.iterrows():
        event_num = row.get(event_col, '')
        story_text = row.get(text_col, '')
        if pd.notna(event_num) and pd.notna(story_text):
            try:
                evn = int(event_num)
            except (TypeError, ValueError):
                try:
                    evn = int(float(event_num))
                except (TypeError, ValueError):
                    continue
            events.append({
                'event': evn,
                'text': str(story_text).strip()
            })
    return sorted(events, key=lambda x: x['event']) if events else None


def _try_read_story_events_from_paths(paths):
    """Try each path until one yields a non-empty event list (deduped by resolved path)."""
    seen = set()
    for fp in paths:
        if fp is None:
            continue
        try:
            key = fp.resolve()
        except OSError:
            key = str(fp)
        if key in seen:
            continue
        seen.add(key)
        if not fp.exists():
            continue
        try:
            from helpers.flexible_io import read_tabular, normalize_story_events_df
            df = normalize_story_events_df(read_tabular(fp))
            result = _read_story_events_dataframe(df)
            if result:
                return result
        except Exception as e:
            print(f"Skipping story events file {fp}: {e}")
    return None


def _story_events_paths_candidate_list(base_id, file_version, min_event_count, events_dir):
    """Ordered candidate paths for story events in one directory (same order as load logic)."""
    from helpers.flexible_io import STORY_EVENTS_EXTENSIONS, read_tabular, normalize_story_events_df

    event_filename_base = f"{base_id}_events"
    paths_to_try = []

    def _events_glob(pattern_stem):
        found = []
        for ext in STORY_EVENTS_EXTENSIONS:
            found.extend(events_dir.glob(f"{pattern_stem}{ext}"))
        return found

    if file_version and any(file_version.endswith(ext) for ext in STORY_EVENTS_EXTENSIONS) and '_events' in file_version:
        paths_to_try = [events_dir / Path(file_version).name]
    elif file_version and file_version.endswith('-edit'):
        paths_to_try = _events_glob(f"{event_filename_base}_{file_version}")
        if not paths_to_try:
            paths_to_try = [events_dir / f"{event_filename_base}_{file_version}.xlsx"]
    elif file_version == 'original':
        paths_to_try = _events_glob(event_filename_base)
        if not paths_to_try:
            candidates = sorted(_events_glob(f"{event_filename_base}-*"), key=lambda p: p.stat().st_mtime, reverse=True)
            candidates = [c for c in candidates if not is_user_edit_file(c.name)]
            if candidates:
                paths_to_try = [candidates[0]]
    else:
        edit_file = find_best_edit_file(events_dir, base_id, '_events', '.xlsx')
        if not edit_file:
            for ext in STORY_EVENTS_EXTENSIONS:
                edit_file = find_best_edit_file(events_dir, base_id, '_events', ext)
                if edit_file:
                    break
        method_files = _events_glob(f"{event_filename_base}-*")
        method_files.extend(_events_glob(f"{event_filename_base}_rule-based"))
        method_files.extend(_events_glob(f"{event_filename_base}_api"))
        method_files = list(dict.fromkeys(method_files))
        original_paths = _events_glob(event_filename_base)
        if edit_file:
            paths_to_try.append(edit_file)
        for f in sorted(method_files, key=lambda p: p.stat().st_mtime, reverse=True):
            paths_to_try.append(f)
        for f in original_paths:
            if f not in paths_to_try:
                paths_to_try.append(f)

        if min_event_count > 0 and paths_to_try:
            filtered = []
            for f in paths_to_try:
                try:
                    norm = normalize_story_events_df(read_tabular(f))
                    if len(norm) >= min_event_count:
                        filtered.append(f)
                except Exception:
                    continue
            if filtered:
                paths_to_try = filtered

    # Drop recall/causal pipeline files that merely embed "{story}_events-..." in
    # their names; only genuine story-events files may be loaded here.
    paths_to_try = [p for p in paths_to_try if is_story_events_filename(Path(p).name)]

    # Also recognise alternative naming conventions and tabular formats that the
    # canonical "{base}_events*" globs above miss — e.g. ``story-{name}-segmented.csv``.
    # The recogniser is story-level + cross-step disambiguated, so it won't pull in
    # recall/causal files. Appended after the canonical candidates so explicit
    # ``{base}_events`` / edit files keep priority; deduped, newest-first.
    try:
        from helpers.step_files import find_step_files
        extra = find_step_files(events_dir, 'eventSegment', base_id, is_story=True)
        for f in extra:
            if f not in paths_to_try:
                paths_to_try.append(f)
    except Exception:
        pass
    return paths_to_try


def _get_story_events_in_single_dir(base_id, file_version, min_event_count, events_dir):
    """Resolve story events for one naming base in a single directory."""
    paths_to_try = _story_events_paths_candidate_list(base_id, file_version, min_event_count, events_dir)
    return _try_read_story_events_from_paths(paths_to_try)


def get_resolved_story_events_source_path(item_id, file_version=None, is_story=False, min_event_count=0):
    """Filesystem path to the story-events .xlsx that would be loaded for this item (or None)."""
    del is_story  # same search bases as get_story_events
    if file_version and file_version.endswith('.xlsx') and '_events' in file_version:
        fn = Path(file_version).name
        for events_dir in iter_story_events_search_dirs():
            p = events_dir / fn
            if p.exists():
                return p
        return None
    for base_id in expand_story_event_file_bases(item_id):
        for events_dir in iter_story_events_search_dirs():
            for fp in _story_events_paths_candidate_list(base_id, file_version, min_event_count, events_dir):
                if _try_read_story_events_from_paths([fp]):
                    return fp
    return None


def build_story_events_prefix_for_export(item_id, story_events_filename):
    """Build '{story}_events-{granularity}' segment for export filenames (rated, etc.)."""
    story_key = get_story_name_from_subject_id(item_id)
    root = story_key if story_key else item_id
    if not story_events_filename:
        return f"{root}_events"
    stem = Path(story_events_filename).stem
    stem = re.sub(r'_(\w+-edit)$', '', stem)
    if '_events' not in stem:
        return f"{root}_events"
    left, right = stem.split('_events', 1)
    right = (right or '').lstrip('-').lstrip('_')
    display_story = story_key if story_key else (left or item_id)
    if right:
        return f"{display_story}_events-{right}"
    return f"{display_story}_events"


def resolve_rated_file_path(item_id, file_version=None):
    """Path to the rated recall .xlsx that get_rated_texts would use."""
    output_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR

    def _find_rated_file(subj, suffix=''):
        exact = output_dir / f"{subj}_rate-recall{suffix}.xlsx"
        if exact.exists():
            return exact
        if not suffix:
            candidates = list_subject_rated_recall_source_files(output_dir, subj)
            if candidates:
                return candidates[0]
        return exact

    if file_version and file_version.endswith('-edit'):
        return output_dir / f"{item_id}_rate-recall_{file_version}.xlsx"
    if file_version == 'original':
        return _find_rated_file(item_id)
    edit_file = find_best_edit_file(output_dir, item_id, '_rate-recall', '.xlsx')
    if edit_file:
        return edit_file
    return _find_rated_file(item_id)


def list_subject_rated_recall_source_files(output_dir, subj_id):
    """Rated recall .xlsx outputs for this subject (excludes *-edit files).

    Tolerates method-suffix chains from earlier steps. Matches canonical
    ``{subj}_rate-recall.xlsx``, method-suffixed ``{subj}_rate-recall-<rate-method>.xlsx``,
    and longer names that carry suffixes from spell/parse before ``_rate-recall``
    (e.g. ``{subj}_spell-<m>_parsed-<m>_rate-recall-<m>.xlsx``).
    """
    if not output_dir or not output_dir.exists():
        return []
    out = []
    seen = set()
    for p in sorted(output_dir.glob(f"{subj_id}*_rate-recall*.xlsx"),
                    key=lambda x: x.stat().st_mtime, reverse=True):
        stem = p.stem
        # Word-boundary: stem must be canonical or start with "{subj_id}_"
        if not (stem == f"{subj_id}_rate-recall" or stem.startswith(f"{subj_id}_")):
            continue
        if is_user_edit_file(p.name):
            continue
        k = p.resolve()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def extract_recall_rated_method_slug(rated_path, item_id):
    """Method / processor tag after _rate-recall (e.g. api_<vendor> -> api-<vendor>)."""
    if not rated_path:
        return 'manual'
    name = rated_path.name
    if not name.endswith('.xlsx'):
        return 'manual'
    stem = Path(name).stem
    m = re.match(r'^(?P<base>.+)_(\w+-edit)$', stem)
    if m:
        stem = m.group('base')
    marker = f"{item_id}_rate-recall"
    idx = stem.find(marker)
    if idx < 0:
        return 'manual'
    rest = stem[idx + len(marker):].lstrip('-').lstrip('_')
    if not rest:
        return 'manual'
    return rest.replace('_', '-')


def _sanitize_export_token(s, default='auto'):
    """Safe single token for filenames."""
    if not s:
        return default
    t = re.sub(r'[^\w.\-]+', '-', str(s).strip())
    return t[:120] if t else default


def list_corrected_recall_source_files(output_dir, subj_id):
    """Non-edit corrected recall .txt files: canonical ``{id}.txt`` plus ``{id}_spell-*.txt``."""
    od = Path(output_dir)
    if not od.is_dir():
        return []
    out = []
    p = od / f"{subj_id}.txt"
    if p.exists() and not is_user_edit_file(p.name):
        out.append(p)
    for f in sorted(od.glob(f"{subj_id}_spell-*.txt")):
        if not is_user_edit_file(f.name):
            out.append(f)
    return out


def list_subject_parsed_source_files(output_dir, subj_id):
    """Non-edit parsed .xlsx outputs for this subject, newest first.

    Tolerates method-suffix chains from earlier pipeline steps. Matches both the
    canonical ``{id}_parsed.xlsx`` and longer names where the parsed file carries
    method tags from the spell-correction and/or parsing step, e.g.
    ``{id}_spell-<spell-method>_parsed-<parse-method>.xlsx``.

    Uses a word-boundary guard so ``the_siren_sub-01`` does not also pick up files
    belonging to ``the_siren_sub-012``.
    """
    od = Path(output_dir) if output_dir else None
    if not od or not od.is_dir():
        return []
    out = []
    seen = set()
    from helpers.flexible_io import PARSED_RECALL_EXTENSIONS
    for ext in PARSED_RECALL_EXTENSIONS:
        for p in sorted(od.glob(f"{subj_id}*_parsed*{ext}"),
                        key=lambda x: x.stat().st_mtime, reverse=True):
            stem = p.stem
            if not (stem == f"{subj_id}_parsed" or stem.startswith(f"{subj_id}_")):
                continue
            if is_user_edit_file(p.name):
                continue
            key = p.resolve()
            if key in seen:
                continue
            seen.add(key)
            out.append(p)
    return out


def resolve_corrected_source_path(item_id, file_version=None):
    output_dir = get_output_dir_for_step_type('sentenceCorrect') or RECALL_CORRECTED_DIR
    if file_version and file_version.endswith('-edit'):
        return output_dir / f"{item_id}_{file_version}.txt"
    if file_version == 'original':
        canonical = output_dir / f"{item_id}.txt"
        if canonical.exists():
            return canonical
        cands = list_corrected_recall_source_files(output_dir, item_id)
        return cands[0] if cands else canonical
    edit_file = find_best_edit_file(output_dir, item_id, '', '.txt')
    if edit_file:
        return edit_file
    cands = list_corrected_recall_source_files(output_dir, item_id)
    if not cands:
        return output_dir / f"{item_id}.txt"
    canonical = output_dir / f"{item_id}.txt"
    if canonical.exists():
        return canonical
    return max(cands, key=lambda p: p.stat().st_mtime)


def extract_step_input_version_slug(path, item_id, step_kind):
    """Short label for which input variant is in use (corrected / parsed)."""
    del item_id, step_kind
    if not path or not path.exists():
        return 'auto'
    name = path.name
    if is_user_edit_file(name):
        u = extract_edit_username(name)
        return _sanitize_export_token(f"edited-{u}" if u else 'edited')
    return 'original'


def recall_pipeline_slug(item_id, recall_file_version):
    """Label for the recall pipeline file version (matches panel 'Recall / rated file' selection)."""
    if recall_file_version:
        return _sanitize_export_token(recall_file_version)
    p = resolve_corrected_source_path(item_id, None)
    return extract_step_input_version_slug(p, item_id, 'corrected')


def _get_story_events_for_event_base(base_id, file_version=None, min_event_count=0):
    """Resolve story events for a single naming base (subject id or story name) across all search dirs."""
    for events_dir in iter_story_events_search_dirs():
        ev = _get_story_events_in_single_dir(base_id, file_version, min_event_count, events_dir)
        if ev:
            return ev
    return None


def get_story_events(item_id, file_version=None, is_story=False, min_event_count=0):
    """Get story events for a subject or story.
    Prioritizes user-edit version if available.
    Uses pipeline config outputPath for eventSegment when configured.
    file_version: '{username}-edit', 'original', specific filename, or None (auto-select)
    is_story: Whether item_id is a story name (True) or subject ID (False)
    min_event_count: If > 0, prefer an event file with at least this many events (e.g. when loading causal pairs).
    For subjects, also tries the story name derived from the subject id (e.g. the_siren_sub-01 -> the_siren).
    For stories, also tries editorial shortenings (e.g. pieman_edited -> pieman) so event files match the transcript name.
    """
    if file_version and any(str(file_version).endswith(ext) for ext in ('.xlsx', '.csv', '.tsv', '.xls')) and '_events' in str(file_version):
        fn = Path(file_version).name
        for events_dir in iter_story_events_search_dirs():
            file_path = events_dir / fn
            if file_path.exists() and is_story_events_filename(fn):
                try:
                    from helpers.flexible_io import read_tabular, normalize_story_events_df
                    df = normalize_story_events_df(read_tabular(file_path))
                    result = _read_story_events_dataframe(df)
                    if result is not None:
                        return result
                except Exception as e:
                    print(f"Error reading story events: {e}")
        return None
    
    for base_id in expand_story_event_file_bases(item_id):
        ev = _get_story_events_for_event_base(base_id, file_version, min_event_count)
        if ev:
            return ev
    return None


def _resolve_story_events_file_for_rating(subj_id):
    """Locate a story-events file for recall matching (xlsx/csv/tsv)."""
    from helpers.flexible_io import pick_story_events_file

    events_dir = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
    story_file = pick_story_events_file(events_dir, subj_id)
    if story_file:
        return story_file
    story_name = get_story_name_from_subject_id(subj_id)
    if story_name and story_name != subj_id:
        return pick_story_events_file(events_dir, story_name)
    return None


def _auto_rate_parsed_df(subj_id, parsed_df):
    """Match parsed recall segments to story events; returns rated DataFrame or None."""
    story_file = _resolve_story_events_file_for_rating(subj_id)
    if not story_file:
        return None
    from helpers.flexible_io import order_recall_rating_columns, story_events_records_from_path

    story_events = story_events_records_from_path(story_file)
    if not story_events:
        return None

    import importlib.util
    rate_file = SCRIPTS_DIR / '5_recall-rater.py'
    spec = importlib.util.spec_from_file_location("recall_rater", rate_file)
    rate_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(rate_module)
    match_sonnet = rate_module.match_recall_to_events_sonnet
    match_test = rate_module.match_recall_to_events_test_mode
    get_client = rate_module.get_anthropic_client

    test_mode = os.getenv('ANTHROPIC_API_KEY') is None
    client = None
    if not test_mode:
        try:
            client = get_client()
        except Exception:
            test_mode = True

    if 'recall_in_temporal_order' not in parsed_df.columns:
        from helpers.flexible_io import normalize_parsed_recall_df
        parsed_df = normalize_parsed_recall_df(parsed_df)

    matched_events_list = []
    for _, row in parsed_df.iterrows():
        recall_segment = str(row.get('recall_in_temporal_order', ''))
        if not recall_segment or recall_segment.strip() == '':
            matched_events_list.append('')
            continue
        if test_mode:
            matching_events = match_test(recall_segment, story_events)
        else:
            matching_events = match_sonnet(
                client, recall_segment, story_events, DEFAULT_ANTHROPIC_RECALL_MATCH_MODEL
            )
            import time
            time.sleep(0.5)
        if matching_events:
            matched_events_list.append(','.join(str(int(e)) for e in matching_events))
        else:
            matched_events_list.append('')

    out = parsed_df.copy()
    out['recalled_events'] = matched_events_list
    return order_recall_rating_columns(out)


def _parse_causal_pairs_dataframe(df):
    """Parse long-form causal pair rows from a DataFrame (fine or coarse sheet)."""
    if df is None or getattr(df, 'empty', True):
        return []
    cols = list(df.columns)

    def _find_col(*names):
        for n in names:
            if n in cols:
                return n
        return None

    col_a = _find_col('event_A_number', 'event_A', 'event_a_number', 'event_a', 'cause', 'cause_event')
    col_b = _find_col('event_B_number', 'event_B', 'event_b_number', 'event_b', 'effect', 'effect_event')
    if not col_a and len(cols) >= 1:
        col_a = cols[0]
    if not col_b and len(cols) >= 2:
        col_b = cols[1]

    col_rating = _find_col('rating', 'rate', 'score')
    col_reasoning = _find_col('reasoning', 'reason')
    col_causal_type = _find_col('causal_type', 'type')

    pairs = []
    for _, row in df.iterrows():
        va, vb = row.get(col_a), row.get(col_b)
        if pd.isna(va) or pd.isna(vb):
            continue
        try:
            event_a = int(va)
            event_b = int(vb)
        except (ValueError, TypeError):
            continue

        vr = row.get(col_rating) if col_rating else None
        if pd.notna(vr):
            try:
                rating = min(max(0, int(vr)), 3)
            except (ValueError, TypeError):
                rating = 0
        else:
            rating = 0

        ref_val = ctx_val = sem_val = 0
        if 'referential' in cols and pd.notna(row.get('referential')):
            try:
                ref_val = int(float(row['referential']))
            except (ValueError, TypeError):
                pass
        if 'contextual' in cols and pd.notna(row.get('contextual')):
            try:
                ctx_val = int(float(row['contextual']))
            except (ValueError, TypeError):
                pass
        if 'semantic' in cols and pd.notna(row.get('semantic')):
            try:
                sem_val = int(float(row['semantic']))
            except (ValueError, TypeError):
                pass

        pairs.append({
            'event_A': event_a,
            'event_B': event_b,
            'rating': rating,
            'reasoning': str(row.get(col_reasoning, '')) if col_reasoning and pd.notna(row.get(col_reasoning, '')) else '',
            'causal_type': str(row.get(col_causal_type, '')) if col_causal_type and pd.notna(row.get(col_causal_type, '')) else '',
            'referential': ref_val,
            'contextual': ctx_val,
            'semantic': sem_val,
        })
    return pairs


# Shared by nested export, parse, and save (flags keys match UI checkboxes)
NESTED_OPTIONAL_EXPORT = [
    ('reasoning', 'coarse_reasoning', 'fine_reasoning'),
    ('causal_type', 'coarse_causal_type', 'fine_causal_type'),
    ('referential', 'coarse_referential', 'fine_referential'),
    ('contextual', 'coarse_contextual', 'fine_contextual'),
    ('semantic', 'coarse_semantic', 'fine_semantic'),
]


def _is_nested_combined_causal_sheet(df):
    """True if Sheet1 uses coarse+fine combined rows (nested manual rating export)."""
    if df is None or getattr(df, 'empty', True):
        return False
    cols = set(str(c).strip() for c in df.columns)
    markers = (
        'event_A_coarse-number', 'event_A_coarse_number', 'event_A_coarse',
        'coarse_rating',
    )
    return any(m in cols for m in markers)


def _parse_nested_combined_causal_sheet(df):
    """Parse Sheet1 with coarse columns + fine columns per row (possibly blank coarse on continuation rows)."""
    cols = list(df.columns)

    def _pick(*candidates):
        for name in candidates:
            if name in cols:
                return name
        return None

    ca_a = _pick('event_A_coarse-number', 'event_A_coarse_number')
    ca_b = _pick('event_B_coarse-number', 'event_B_coarse_number')
    fa = _pick('event_A_number')
    fb = _pick('event_B_number')

    cr = _pick('coarse_rating')
    crs = _pick('coarse_reasoning')
    fr = _pick('fine_rating', 'rating.1')
    frs = _pick('fine_reasoning', 'reasoning.1')

    if cr is None and 'rating' in cols and 'rating.1' in cols:
        cr = 'rating'
        crs = 'reasoning' if 'reasoning' in cols else None
        fr = 'rating.1'
        frs = 'reasoning.1' if 'reasoning.1' in cols else None

    pairs = []
    coarse_acc = {}

    for _, row in df.iterrows():
        if fa is not None and fb is not None:
            va, vb = row.get(fa), row.get(fb)
            if pd.notna(va) and pd.notna(vb):
                try:
                    event_a = int(va)
                    event_b = int(vb)
                except (ValueError, TypeError):
                    event_a = None
                if event_a is not None:
                    rating = 0
                    if fr is not None:
                        vr = row.get(fr)
                        if pd.notna(vr):
                            try:
                                rating = min(max(0, int(float(vr))), 3)
                            except (ValueError, TypeError):
                                rating = 0
                    if rating > 0:
                        reason = ''
                        if frs is not None and pd.notna(row.get(frs)):
                            reason = str(row.get(frs) or '')
                        fp_entry = {
                            'event_A': event_a,
                            'event_B': event_b,
                            'rating': rating,
                            'reasoning': reason,
                            'causal_type': '',
                            'referential': 0,
                            'contextual': 0,
                            'semantic': 0,
                        }
                        for fk, _cc, fc in NESTED_OPTIONAL_EXPORT:
                            if fk == 'reasoning':
                                continue
                            col = _pick(fc)
                            if col is None or not pd.notna(row.get(col)):
                                continue
                            if fk == 'causal_type':
                                fp_entry['causal_type'] = str(row.get(col) or '')
                            else:
                                try:
                                    fp_entry[fk] = int(float(row.get(col)))
                                except (ValueError, TypeError):
                                    fp_entry[fk] = 0
                        pairs.append(fp_entry)

        if ca_a is not None and ca_b is not None and cr is not None:
            vca, vcb = row.get(ca_a), row.get(ca_b)
            if pd.notna(vca) and pd.notna(vcb):
                try:
                    ca = int(float(vca))
                    cb = int(float(vcb))
                except (ValueError, TypeError):
                    ca = None
                if ca is not None:
                    vr = row.get(cr)
                    if pd.notna(vr):
                        try:
                            crating = min(max(0, int(float(vr))), 3)
                        except (ValueError, TypeError):
                            crating = 0
                    else:
                        crating = 0
                    if crating > 0:
                        key = (ca, cb)
                        reason = ''
                        if crs is not None and pd.notna(row.get(crs)):
                            reason = str(row.get(crs) or '')
                        cp_entry = {
                            'event_A': ca,
                            'event_B': cb,
                            'rating': crating,
                            'reasoning': reason,
                            'causal_type': '',
                            'referential': 0,
                            'contextual': 0,
                            'semantic': 0,
                        }
                        for fk, cc, _fc in NESTED_OPTIONAL_EXPORT:
                            if fk == 'reasoning':
                                continue
                            col = _pick(cc)
                            if col is None or not pd.notna(row.get(col)):
                                continue
                            if fk == 'causal_type':
                                cp_entry['causal_type'] = str(row.get(col) or '')
                            else:
                                try:
                                    cp_entry[fk] = int(float(row.get(col)))
                                except (ValueError, TypeError):
                                    cp_entry[fk] = 0
                        coarse_acc[key] = cp_entry

    coarse_pairs = list(coarse_acc.values())
    return pairs, coarse_pairs


def nested_combined_export_columns(flags):
    """Column order for nested causal Excel export; mirrors UI checkbox flags."""
    flags = flags or {}
    cols = ['event_A_coarse-number', 'event_B_coarse-number', 'coarse_rating']
    for fk, cc, _fc in NESTED_OPTIONAL_EXPORT:
        if flags.get(fk):
            cols.append(cc)
    cols.extend(['event_A_number', 'event_B_number', 'fine_rating'])
    for fk, _cc, fc in NESTED_OPTIONAL_EXPORT:
        if flags.get(fk):
            cols.append(fc)
    return cols


def _nested_cell_value(raw):
    if raw is None or raw == '':
        return None
    if isinstance(raw, str) and not raw.strip():
        return None
    return raw


def _nested_coarse_optional(cd, cr, flags, first_row):
    """Coarse-side optional columns for nested export (cd = coarse pair dict, cr = coarse rating)."""
    d = {}
    cd = cd or {}
    active = bool(first_row and cr > 0)
    for fk, cc, _fc in NESTED_OPTIONAL_EXPORT:
        if not flags.get(fk):
            continue
        if not active:
            d[cc] = ''
            continue
        if fk == 'reasoning':
            v = cd.get('reasoning', '') or ''
            d[cc] = v.strip() if str(v).strip() else 'manual entry'
        elif fk == 'causal_type':
            d[cc] = str(cd.get('causal_type', '') or '')
        else:
            try:
                d[cc] = int(cd.get(fk, 0) or 0)
            except (TypeError, ValueError):
                d[cc] = 0
    return d


def _nested_fine_optional(fp, fr, flags):
    """Fine-side optional columns for nested export."""
    d = {}
    fp = fp or {}
    active = fr > 0
    for fk, _cc, fc in NESTED_OPTIONAL_EXPORT:
        if not flags.get(fk):
            continue
        if not active:
            d[fc] = ''
            continue
        if fk == 'reasoning':
            v = fp.get('reasoning', '') or ''
            d[fc] = v.strip() if str(v).strip() else 'manual entry'
        elif fk == 'causal_type':
            d[fc] = str(fp.get('causal_type', '') or '')
        else:
            try:
                d[fc] = int(fp.get(fk, 0) or 0)
            except (TypeError, ValueError):
                d[fc] = 0
    return d


def _save_nested_combined_xlsx(file_path, nested_rows, columns):
    """Write nested causal export: one sheet with dynamic optional columns from flags."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(list(columns))
    for r in nested_rows:
        row_out = []
        for c in columns:
            v = _nested_cell_value(r.get(c))
            row_out.append(v)
        ws.append(row_out)
    wb.save(str(file_path))


def _build_nested_combined_from_parts(item_id, fine_pairs, coarse_pairs, flags):
    """Server-side fallback: build nested combined rows from separate fine/coarse pair lists + event files."""
    gran = get_all_event_granularities(item_id)
    if not gran:
        return []
    fine_events = gran.get('fine') or []
    coarse_events = gran.get('coarse') or []
    if not fine_events or not coarse_events:
        return []

    # Build fine event number → coarse event number mapping
    fine_num_to_coarse_num = {}
    fi = 0
    for ci, ce in enumerate(coarse_events):
        c_words = ce['text'].lower().split()
        matched = 0
        while fi < len(fine_events):
            f_words = fine_events[fi]['text'].lower().split()
            check = min(3, len(f_words))
            found = True
            for w in range(check):
                if f_words[w] not in c_words[matched:matched + len(f_words) + 5]:
                    found = False
                    break
            if found:
                fine_num_to_coarse_num[int(fine_events[fi]['event'])] = int(ce['event'])
                matched += len(f_words)
                fi += 1
            else:
                break

    # Group fine pairs by coarse pair
    groups = {}
    for fp in (fine_pairs or []):
        fa, fb = int(fp.get('event_A', 0)), int(fp.get('event_B', 0))
        ca = fine_num_to_coarse_num.get(fa)
        cb = fine_num_to_coarse_num.get(fb)
        if ca is None or cb is None:
            continue
        gk = f"{ca}_{cb}"
        groups.setdefault(gk, []).append(fp)

    # Build coarse data lookup
    coarse_data = {}
    for cp in (coarse_pairs or []):
        r = int(cp.get('rating', 0))
        if r > 0:
            ca, cb = int(cp.get('event_A', 0)), int(cp.get('event_B', 0))
            gk = f"{ca}_{cb}"
            coarse_data[gk] = cp
            groups.setdefault(gk, [])

    rows = []
    for gk in sorted(groups, key=lambda k: tuple(int(x) for x in k.split('_'))):
        ca_s, cb_s = gk.split('_')
        ca_num, cb_num = int(ca_s), int(cb_s)
        fine_list = sorted(groups[gk], key=lambda p: (int(p.get('event_A', 0)), int(p.get('event_B', 0))))
        cd = coarse_data.get(gk, {})
        cr = int(cd.get('rating', 0)) if cd else 0

        if not fine_list:
            row = {
                'event_A_coarse-number': ca_num,
                'event_B_coarse-number': cb_num,
                'coarse_rating': cr if cr > 0 else '',
            }
            row.update(_nested_coarse_optional(cd, cr, flags, True))
            row['event_A_number'] = ''
            row['event_B_number'] = ''
            row['fine_rating'] = ''
            row.update(_nested_fine_optional({}, 0, flags))
            rows.append(row)
        else:
            for idx, fp in enumerate(fine_list):
                fr = int(fp.get('rating', 0))
                row = {
                    'event_A_coarse-number': ca_num if idx == 0 else '',
                    'event_B_coarse-number': cb_num if idx == 0 else '',
                    'coarse_rating': (cr if cr > 0 else '') if idx == 0 else '',
                }
                row.update(_nested_coarse_optional(cd, cr, flags, idx == 0))
                row['event_A_number'] = int(fp.get('event_A', 0))
                row['event_B_number'] = int(fp.get('event_B', 0))
                row['fine_rating'] = fr
                row.update(_nested_fine_optional(fp, fr, flags))
                rows.append(row)
    return rows


def _merge_event_hierarchy_coarse_cells(worksheet, n_data_rows):
    """Merge coarse_event columns (A:B) for consecutive rows with the same coarse event number."""
    if n_data_rows < 1:
        return
    start = 2
    current_coarse = worksheet.cell(row=2, column=1).value
    block_start = 2
    for r in range(3, n_data_rows + 2):
        val = worksheet.cell(row=r, column=1).value
        if val != current_coarse:
            if r - 1 > block_start:
                worksheet.merge_cells(start_row=block_start, start_column=1, end_row=r - 1, end_column=1)
                worksheet.merge_cells(start_row=block_start, start_column=2, end_row=r - 1, end_column=2)
            block_start = r
            current_coarse = val
    if n_data_rows + 1 > block_start:
        worksheet.merge_cells(start_row=block_start, start_column=1, end_row=n_data_rows + 1, end_column=1)
        worksheet.merge_cells(start_row=block_start, start_column=2, end_row=n_data_rows + 1, end_column=2)


def get_causal_ratings(item_id, file_version=None):
    """Get causal rating data for a story.
    Loads the most recent causal rating file and returns it as an N×N matrix.
    Returns dict with 'matrix' (2D array), 'event_count', and 'source_file', or None.
    file_version: specific .xlsx filename, or None (auto-select).
    """
    causal_dir = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
    if not causal_dir or not causal_dir.exists():
        return None

    if file_version and file_version.endswith('.xlsx') and 'causal' in file_version:
        best_file = causal_dir / file_version
        if not best_file.exists():
            return None
    else:
        patterns = [
            f"{item_id}_causal-*.xlsx",
            f"{item_id}_causal.xlsx",
        ]
        files = []
        for pattern in patterns:
            files.extend(list(causal_dir.glob(pattern)))
        edit_files = [f for f in files if is_user_edit_file(f.name)]
        non_edit_files = [f for f in files if not is_user_edit_file(f.name)]

        best_file = None
        if edit_files:
            username = session.get('username', 'human')
            for ef in edit_files:
                if f"_{username}-edit" in ef.name:
                    best_file = ef
                    break
            if not best_file:
                best_file = max(edit_files, key=lambda p: p.stat().st_mtime)
        if not best_file and non_edit_files:
            best_file = max(non_edit_files, key=lambda p: p.stat().st_mtime)

    if not best_file:
        return None

    try:
        xl = pd.ExcelFile(best_file)
        first_name = xl.sheet_names[0]
        df = pd.read_excel(best_file, sheet_name=first_name)

        coarse_pairs = []
        if _is_nested_combined_causal_sheet(df):
            pairs, coarse_pairs = _parse_nested_combined_causal_sheet(df)
        else:
            pairs = _parse_causal_pairs_dataframe(df)
            if 'coarse_causal_pairs' in xl.sheet_names:
                df_c = pd.read_excel(best_file, sheet_name='coarse_causal_pairs')
                coarse_pairs = _parse_causal_pairs_dataframe(df_c)

        out = {
            'pairs': pairs,
            'source_file': best_file.name,
        }
        if coarse_pairs:
            out['coarse_pairs'] = coarse_pairs
        return out
    except Exception as e:
        print(f"Error reading causal ratings: {e}")
        return None


def get_all_event_granularities(item_id):
    """Get both fine-grained and coarse-grained event segmentations for a story.
    Returns dict with 'fine' and 'coarse' event lists, or None if not available.
    """
    search_dirs = [d for d in iter_story_events_search_dirs() if d and d.exists()]
    if not search_dirs:
        return None

    result = {'fine': None, 'coarse': None, 'fine_file': None, 'coarse_file': None}

    def _load_event_list(glob_pat):
        for events_dir in search_dirs:
            files = sorted(events_dir.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
            for f in files:
                if is_user_edit_file(f.name):
                    continue
                try:
                    df = pd.read_excel(f)
                    if 'event' not in df.columns or 'story_texts' not in df.columns:
                        continue
                    evs = [{'event': int(r['event']), 'text': str(r['story_texts']).strip()}
                           for _, r in df.iterrows() if pd.notna(r['event']) and pd.notna(r['story_texts'])]
                    if evs:
                        return f.name, evs
                except Exception:
                    continue
            for f in files:
                if not is_user_edit_file(f.name):
                    continue
                try:
                    df = pd.read_excel(f)
                    if 'event' not in df.columns or 'story_texts' not in df.columns:
                        continue
                    evs = [{'event': int(r['event']), 'text': str(r['story_texts']).strip()}
                           for _, r in df.iterrows() if pd.notna(r['event']) and pd.notna(r['story_texts'])]
                    if evs:
                        return f.name, evs
                except Exception:
                    continue
        return None, None

    ff, fe = _load_event_list(f"{item_id}_events-fine*.xlsx")
    if ff:
        result['fine_file'], result['fine'] = ff, fe
    cf, ce = _load_event_list(f"{item_id}_events-coarse*.xlsx")
    if cf:
        result['coarse_file'], result['coarse'] = cf, ce

    if result['fine'] and result['coarse']:
        return result
    return None


def _norm_path_for_compare(path) -> str:
    """Normalize a path string for equality / prefix checks without resolve()."""
    try:
        return os.path.normpath(str(Path(path).expanduser()))
    except (TypeError, ValueError, OSError):
        return os.path.normpath(str(path))


def _path_is_same_or_under(child, parent) -> bool:
    """True if ``child`` equals ``parent`` or is nested under it."""
    c = _norm_path_for_compare(child)
    p = _norm_path_for_compare(parent)
    if c == p:
        return True
    return c.startswith(p + os.sep) or c.startswith(p + '/')


def _looks_like_user_absolute_dir(path: Path) -> bool:
    """Heuristic for drag-dropped / typed absolute directory paths."""
    try:
        s = str(path.expanduser()).strip()
    except (TypeError, ValueError):
        return False
    if not s or s in ('/', '.', ''):
        return False
    if path.is_absolute() or s.startswith('~'):
        return True
    if len(s) >= 2 and s[1] == ':':  # Windows drive letter
        return True
    return False


def _path_is_dir_or_unstatable(path: Path) -> bool:
    """True when ``path`` is a directory, or likely one macOS TCC blocks stat on."""
    try:
        return path.is_dir()
    except PermissionError:
        return _looks_like_user_absolute_dir(path)
    except OSError:
        return False


def _resolve_path_lenient(path: Path, *, root: Path) -> Path:
    """Resolve when possible; keep the expanded path if TCC blocks resolve()."""
    try:
        return path.resolve()
    except OSError:
        return path


def _resolve_and_validate_output_path(path_str):
    """Resolve user-provided output path and accept paths inside any
    configured pipeline I/O directory.

    Accepts:
      * project-relative paths (resolved against ``PROJECT_ROOT``);
      * absolute paths under ``PROJECT_ROOT``;
      * absolute paths under any input/output directory currently configured
        in the pipeline (so a user-chosen, drag-dropped folder outside the
        project root is still trusted for save operations).

    Returns Path or None if invalid.
    """
    if not path_str or not str(path_str).strip():
        return None
    s = str(path_str).strip()
    try:
        root = WORKSPACE_ROOT.resolve()
    except OSError:
        root = WORKSPACE_ROOT
    p_in = Path(s).expanduser()
    p = _resolve_path_lenient(
        p_in if p_in.is_absolute() else (root / s),
        root=root,
    )
    if _path_is_same_or_under(p, root):
        return p

    norm_s = _norm_path_for_compare(p)

    # Allow paths that fall within any user-configured pipeline directory.
    config = get_pipeline_config() or {}
    for step in (config.get('steps') or []):
        for key in ('inputPath', 'outputPath'):
            raw_cfg = (step.get(key) or '').strip()
            if raw_cfg and _norm_path_for_compare(raw_cfg) == norm_s:
                return p
            cfg_dir = _resolve_path_from_config(raw_cfg)
            if cfg_dir is None:
                continue
            if _path_is_same_or_under(p, cfg_dir):
                return p
        for _field, _env_var, cfg_dir in _resolve_step_extra_input_dirs(step):
            if _path_is_same_or_under(p, cfg_dir):
                return p
    return None


def get_default_export_path(
    item_id,
    step,
    source_file=None,
    is_story=False,
    story_events_file=None,
    recall_file_version=None,
    story_transcript_file=None,
):
    """Get the default export path for a given step.
    Returns path as string (relative to PROJECT_ROOT when possible).
    step: 'corrected', 'parsed', 'rated', 'audio', 'story-events', or 'causal'

    story_events_file / recall_file_version mirror inspection UI query params so filenames
    encode the story-segmentation version and recall-pipeline variant in use.
    """
    edit_suffix = get_edit_suffix()

    def _rel(p):
        try:
            return _path_for_client(p)
        except ValueError:
            return str(p)

    events_path = get_resolved_story_events_source_path(item_id, story_events_file, is_story=is_story)
    events_fn = (events_path.name if events_path else None) or (
        Path(story_events_file).name if (story_events_file and str(story_events_file).endswith('.xlsx')) else None
    )
    events_prefix = build_story_events_prefix_for_export(item_id, events_fn)

    transcript_slug = 'auto'
    if story_transcript_file and str(story_transcript_file).strip():
        transcript_slug = _sanitize_export_token(Path(story_transcript_file).stem)

    if step == 'corrected':
        d = get_output_dir_for_step_type('sentenceCorrect') or RECALL_CORRECTED_DIR
        pl = recall_pipeline_slug(item_id, recall_file_version)
        stem = f"{item_id}_recall-corrected_recall-version-{pl}"
        return _rel(d / f"{stem}{edit_suffix}.txt")
    if step == 'parsed':
        d = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
        pl = recall_pipeline_slug(item_id, recall_file_version)
        stem = f"{events_prefix}_{item_id}_textParsingd_recall-version-{pl}"
        return _rel(d / f"{stem}{edit_suffix}.xlsx")
    if step == 'rated':
        d = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
        rated_p = resolve_rated_file_path(item_id, recall_file_version)
        method = extract_recall_rated_method_slug(rated_p, item_id)
        stem = f"{events_prefix}_{item_id}_rate-recall_{method}"
        return _rel(d / f"{stem}{edit_suffix}.xlsx")
    if step == 'audio':
        transcribed_dir = get_output_dir_for_step_type('audioTranscribe:recall') or RECALL_AUDIO_TRANSCRIBED_DIR
        audio_file = get_audio_file(item_id, is_story=is_story)
        if not audio_file:
            stem = f"{item_id}_audio-transcript_source-{transcript_slug}"
            return _rel(transcribed_dir / f"{stem}{edit_suffix}.txt")
        audio_path = PROJECT_ROOT / audio_file
        stem = f"{audio_path.stem}_audio-transcript_source-{transcript_slug}"
        return _rel(transcribed_dir / f"{stem}{edit_suffix}.txt")
    if step == 'story-events':
        d = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
        if source_file and source_file.endswith('.xlsx'):
            base_name = Path(source_file).stem
            base_name = re.sub(r'_(\w+-edit)$', '', base_name)
            stem = f"{base_name}_segmented_transcript-{transcript_slug}"
            return _rel(d / f"{stem}{edit_suffix}.xlsx")
        if events_fn:
            stem = f"{build_story_events_prefix_for_export(item_id, events_fn)}_segmented_transcript-{transcript_slug}"
            return _rel(d / f"{stem}{edit_suffix}.xlsx")
        return _rel(d / f"{item_id}_events_segmented_transcript-{transcript_slug}{edit_suffix}.xlsx")
    if step == 'causal':
        d = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
        if source_file and source_file.endswith('.xlsx'):
            base_name = Path(source_file).stem
            if is_user_edit_file(source_file):
                return _rel(d / source_file)
            base_name = re.sub(r'_(\w+-edit)$', '', base_name)
            stem = f"{events_prefix}_{base_name}_causal-rating"
            return _rel(d / f"{stem}{edit_suffix}.xlsx")
        stem = f"{events_prefix}_{item_id}_causal-manual"
        return _rel(d / f"{stem}{edit_suffix}.xlsx")
    return None


def get_audio_file(item_id, is_story=False):
    """Get audio file path for a subject (recall audio) or story (story audio).
    Supports wav, mp3, mp4, m4a, flac, ogg, webm, aac."""
    if is_story:
        audio_dir = STORY_AUDIO_DIR
        # For stories, try exact match first, then patterns
        patterns = []
        for ext in SUPPORTED_AUDIO_EXTENSIONS:
            patterns.extend([f"{item_id}{ext}", f"{item_id}*{ext}", f"*{item_id}*{ext}"])
    else:
        audio_dir = RECALL_AUDIO_DIR
        # Try different naming patterns
        patterns = []
        for ext in SUPPORTED_AUDIO_EXTENSIONS:
            patterns.extend([f"{item_id}_recall*{ext}", f"{item_id}_*{ext}", f"*{item_id}*{ext}"])
    
    for pattern in patterns:
        files = list(audio_dir.glob(pattern))
        if files:
            return _path_for_client(files[0])
    
    return None


def get_audio_transcription(item_id, is_story=False):
    """Get audio transcription text for a subject or story.
    Uses pipeline config outputPath for audioTranscribe:story/audioTranscribe:recall when configured.
    """
    if is_story:
        # For story audio transcription
        transcribed_dir = get_output_dir_for_step_type('audioTranscribe:story') or STORY_AUDIO_TRANSCRIBED_DIR
        if not transcribed_dir.exists():
            return None
        patterns = [
            f"{item_id}.txt",
            f"{item_id}*.txt",
            f"*{item_id}*.txt"
        ]
    else:
        # For recall audio transcription
        transcribed_dir = get_output_dir_for_step_type('audioTranscribe:recall') or RECALL_AUDIO_TRANSCRIBED_DIR
        if not transcribed_dir.exists():
            return None
        patterns = [
            f"{item_id}_recall*.txt",
            f"{item_id}_*.txt",
            f"*{item_id}*.txt"
        ]
    
    for pattern in patterns:
        files = list(transcribed_dir.glob(pattern))
        if files:
            try:
                with open(files[0], 'r', encoding='utf-8') as f:
                    return f.read().strip()
            except Exception as e:
                print(f"Error reading transcription: {e}")
                return None
    
    return None


def login_required(f):
    """No-op pass-through.

    Account/password auth was removed: the tool now identifies an editor by a
    free-text *rater name* entered on the pipeline-config page (used only to
    label exported edit files). This decorator is kept so existing
    ``@login_required`` route annotations remain valid without churn, but it no
    longer gates access.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated_function


def current_rater_name():
    """The active rater name (defaults to 'human' when none set yet)."""
    return session.get('username') or 'human'


@app.context_processor
def inject_template_globals():
    """Rater name and public feedback links for all templates."""
    return {
        'logged_in_username': session.get('username', ''),
        # Production (forced login) pins the rater to the account name, so
        # templates hide the editable rater field / change-rater controls.
        'require_auth': REQUIRE_AUTH,
        'narraters_repo_url': REPO_URL,
        'narraters_feedback_url': FEEDBACK_ISSUE_URL,
        'narraters_bug_url': BUG_ISSUE_URL,
        'narraters_discussions_url': DISCUSSIONS_URL,
        'narraters_issues_url': ISSUES_URL,
    }


# --- Rater identity (replaces account/password login) -----------------------
#
# There is no authentication anymore. An editor just types a free-text rater
# name on the pipeline-config page. It is used only to label exported edit
# files ({subj_id}_{rater}-edit.ext), so a dummy name is perfectly fine.

# Word lists for the "roll the dice" random rater name (adjective + noun).
_RATER_ADJECTIVES = [
    'brave', 'calm', 'clever', 'curious', 'eager', 'gentle', 'happy', 'jolly',
    'keen', 'lively', 'mellow', 'nimble', 'plucky', 'quiet', 'rapid', 'sunny',
    'swift', 'witty', 'zesty', 'bold', 'bright', 'cosmic', 'fuzzy', 'lucky',
]
_RATER_NOUNS = [
    'otter', 'falcon', 'maple', 'comet', 'badger', 'willow', 'panda', 'heron',
    'fox', 'lynx', 'raven', 'tiger', 'walrus', 'cedar', 'koala', 'moth',
    'newt', 'quail', 'robin', 'seal', 'wren', 'yak', 'bison', 'crane',
]


def sanitize_rater_name(raw):
    """Normalize a rater name to a filesystem/regex-safe token.

    Edit files are named ``{subj_id}_{rater}-edit.ext`` and the rater is later
    re-extracted with ``_(\\w+)-edit``. So we collapse the input to ``\\w``
    characters only. Returns '' if nothing usable remains.
    """
    token = re.sub(r'\W+', '_', (raw or '').strip()).strip('_')
    return token[:32]


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """App-level password login. Active only when REQUIRE_AUTH is on; otherwise
    there is nothing to log into, so we just pass through to the app."""
    if not REQUIRE_AUTH or current_user.is_authenticated:
        return redirect(_safe_next(request.args.get('next')) or '/')

    error = None
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''
        nxt = _safe_next(data.get('next') or request.args.get('next'))
        if username and verify_user(username, password):
            login_user(User(username))
            session['username'] = username  # default rater label for edit files
            try:
                log_user_login(username)
            except Exception as e:
                print(f"login: log_user_login failed (non-fatal): {e}")
            target = nxt or '/'
            if request.is_json:
                return jsonify({'success': True, 'redirect': target})
            return redirect(target)
        error = 'Invalid username or password.'
        if request.is_json:
            return jsonify({'success': False, 'error': error}), 401

    return render_template(
        'login.html', error=error, next=_safe_next(request.args.get('next')) or ''
    )


@app.route('/api/set-rater', methods=['POST'])
def api_set_rater():
    """Store the rater name for this session (used only to label edit files).

    In production (REQUIRE_AUTH) the rater is pinned to the logged-in account
    and cannot be changed, so any client-supplied name is ignored.
    """
    if REQUIRE_AUTH:
        pinned = current_user.id if current_user.is_authenticated else session.get('username')
        if pinned:
            session['username'] = pinned
            return jsonify({'success': True, 'rater_name': pinned})
        return jsonify({'success': False, 'error': 'Authentication required'}), 401
    try:
        data = request.get_json(silent=True) or {}
        raw = str(data.get('rater_name', ''))
        clean = sanitize_rater_name(raw)
        if not clean:
            return jsonify({
                'success': False,
                'error': 'Please enter a rater name (letters, numbers, or underscore).'
            }), 400
        session['username'] = clean
        session['logged_in'] = True  # harmless legacy flag some templates read
        try:
            log_user_login(clean)
        except Exception as e:
            print(f"set-rater: log_user_login failed (non-fatal): {e}")
        return jsonify({'success': True, 'rater_name': clean})
    except Exception as e:
        print(f"set-rater error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/random-name', methods=['GET'])
def api_random_name():
    """Return a random adjective+noun rater name for the dice button."""
    import random
    name = random.choice(_RATER_ADJECTIVES).capitalize() + random.choice(_RATER_NOUNS).capitalize()
    return jsonify({'success': True, 'name': name})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """Log out: end the Flask-Login session and clear the rater state."""
    logout_user()
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out'})


@app.route('/api/verify-auth', methods=['GET'])
def api_verify_auth():
    """Report the current rater name (kept for templates that poll it)."""
    name = session.get('username')
    if name:
        return jsonify({'success': True, 'username': name})
    return jsonify({'success': False}), 401


# ===========================================================================
# Benchmark mode (narraters serve --benchmark)
# ---------------------------------------------------------------------------
# A focused text-matching workflow: list recall files under benchmark/unrated/
# to rate, open one in the existing matching UI (subject.html step3), and save
# the result under benchmark/rated/<username>/. The benchmark CSV format is
# already understood by helpers.flexible_io and the existing matching UI, so
# these handlers just read/write the known file paths and reuse that machinery.

_BENCH_MATCHED_RE = re.compile(
    r'^(?P<sub>.+?)-recall-(?P<story>.+)-matched\.(?:csv|tsv|xlsx)$', re.IGNORECASE
)


# Benchmark items are grouped into hardcoded batches, released to raters one
# batch at a time (the admin shows/hides batches per rater on /admin; only the
# FIRST batch is visible by default). Batches also define the display order.
#
# Each batch lists matchers; an item belongs to the first batch with a matcher
# it satisfies (all keys of a matcher must match). Matcher keys:
#   'name'       — matches the story folder name OR the datasource name
#   'datasource' — datasource name (`narraters-<datasource>-matching`)
#   'story'      — story folder name
#   'sub_ids'    — list of subject ids; lets one story's subjects be split
#                  across batches, e.g. {'story': 'sirens2', 'sub_ids': [...]}
#                  in an early batch and {'story': 'sirens2'} in a later one.
# Items matching no batch fall into an implicit trailing 'other' batch.
BENCHMARK_BATCHES = [
    {'name': 'sirens2',          'match': [{'name': 'sirens2'}]},
    {'name': 'georgiou',         'match': [{'name': 'georgiou'}]},
    {'name': 'flashfiction',     'match': [{'name': 'flashfiction'}]},
    {'name': 'memsearch',        'match': [{'name': 'memsearch'}]},
    {'name': 'monthiversary',    'match': [{'name': 'monthiversary'}]},
    {'name': 'sherlock',         'match': [{'name': 'sherlock'}]},
    {'name': 'alice',            'match': [{'name': 'alice'}]},
    {'name': 'emomem',           'match': [{'name': 'emomem'}]},
    {'name': 'eternal_sunshine', 'match': [{'name': 'eternal_sunshine'}]},
]
BENCHMARK_OTHER_BATCH = 'other'


def _benchmark_matcher_hits(matcher, item):
    """Whether one batch matcher matches one scanned item (all keys must hold)."""
    if 'name' in matcher and matcher['name'] not in (item['story'], item['datasource']):
        return False
    if 'datasource' in matcher and matcher['datasource'] != item['datasource']:
        return False
    if 'story' in matcher and matcher['story'] != item['story']:
        return False
    if 'sub_ids' in matcher and item['sub_id'] not in matcher['sub_ids']:
        return False
    return True


def _benchmark_batch_for(item):
    """Name of the first batch in BENCHMARK_BATCHES that matches ``item``."""
    for batch in BENCHMARK_BATCHES:
        if any(_benchmark_matcher_hits(m, item) for m in batch['match']):
            return batch['name']
    return BENCHMARK_OTHER_BATCH


def _benchmark_batch_index(batch_name):
    """Position of a batch in the display order ('other' sorts last)."""
    for i, batch in enumerate(BENCHMARK_BATCHES):
        if batch['name'] == batch_name:
            return i
    return len(BENCHMARK_BATCHES)


def _benchmark_batch_names(items=None):
    """Batch names in display order; 'other' appended only when some scanned
    item actually falls into it. Pass ``items`` to reuse an existing scan."""
    names = [b['name'] for b in BENCHMARK_BATCHES]
    if items is None:
        items = _benchmark_scan()
    if any(it['batch'] == BENCHMARK_OTHER_BATCH for it in items):
        names.append(BENCHMARK_OTHER_BATCH)
    return names


def _benchmark_datasource_name(ds_dir_name):
    """`narraters-monthiversary-matching` -> `monthiversary`."""
    name = ds_dir_name
    if name.startswith('narraters-'):
        name = name[len('narraters-'):]
    if name.endswith('-matching'):
        name = name[:-len('-matching')]
    return name or ds_dir_name


def _benchmark_find_segmented(story_dir):
    """Locate the `story-<story>-segmented.*` file in a story folder."""
    for pattern in ('story-*-segmented.*', '*-segmented.*'):
        hits = sorted(p for p in story_dir.glob(pattern) if p.is_file())
        if hits:
            return hits[0]
    return None


def _benchmark_scan():
    """Discover all recall files to rate under benchmark/unrated/.

    Layout: unrated/<ds_dir>/<story>/matches/<sub>-recall-<story>-matched.csv
    Returns a list of item dicts (id, ds_dir, datasource, story, sub_id, batch,
    matched_file, segmented_file), ordered by BENCHMARK_BATCHES, then by
    datasource, story, sub_id.
    """
    items = []
    if not BENCHMARK_UNRATED_DIR.is_dir():
        return items
    for ds_dir in sorted(p for p in BENCHMARK_UNRATED_DIR.iterdir() if p.is_dir()):
        datasource = _benchmark_datasource_name(ds_dir.name)
        for story_dir in sorted(p for p in ds_dir.iterdir() if p.is_dir()):
            matches_dir = story_dir / 'matches'
            if not matches_dir.is_dir():
                continue
            segmented = _benchmark_find_segmented(story_dir)
            for mf in sorted(p for p in matches_dir.iterdir() if p.is_file()):
                m = _BENCH_MATCHED_RE.match(mf.name)
                if not m:
                    continue
                sub_id = m.group('sub')
                story = story_dir.name
                item = {
                    'id': f"bench__{ds_dir.name}__{story}__{sub_id}",
                    'ds_dir': ds_dir.name,
                    'datasource': datasource,
                    'story': story,
                    'sub_id': sub_id,
                    'matched_file': mf,
                    'segmented_file': segmented,
                }
                item['batch'] = _benchmark_batch_for(item)
                items.append(item)

    items.sort(key=lambda it: (_benchmark_batch_index(it['batch']),
                               it['datasource'], it['story'], it['sub_id']))
    return items


def _benchmark_item_for_id(slug):
    """Resolve a benchmark slug (bench__<ds>__<story>__<sub>) to its item dict."""
    if not slug or not str(slug).startswith('bench__'):
        return None
    for item in _benchmark_scan():
        if item['id'] == slug:
            return item
    return None


def _benchmark_username():
    """Rater name for the rated/ folder + greeting. Uses the logged-in name when
    present (hosting), else defaults to the OS login name for local use."""
    name = session.get('username')
    if name:
        return name
    import getpass
    try:
        raw = getpass.getuser()
    except Exception:
        raw = 'human'
    clean = sanitize_rater_name(raw) or 'human'
    session['username'] = clean
    return clean


def _benchmark_username_error():
    """Block reading/writing benchmark data if the active rater name isn't
    filesystem-safe (the name is used as a directory + filename prefix).

    Returns a Flask ``(response, status)`` error tuple to surface, or ``None``
    when the name is safe. Defense in depth: account creation already rejects
    unsafe names (`narraters users add`), so this only catches legacy accounts
    created before that check existed."""
    username = _benchmark_username()
    if sanitize_rater_name(username) != username:
        return jsonify({
            'error': 'Your account name contains characters that are not allowed '
                     'for saving ratings. Please take a screenshot of this message '
                     'and email it to the admin. Do not continue rating.',
            'error_code': 'bad_username',
        }), 409
    return None


def _benchmark_active_pass(username):
    """Which pass this rater is locked to (1 = matching, 2 = rating). Admin-set
    via `narraters users second-pass/first-pass <name>`; defaults to 1."""
    try:
        return get_benchmark_pass(username)
    except Exception:
        return 1


def _benchmark_default_batch_visible(batch_name):
    """Server default for batches the admin has not explicitly toggled: only
    the first hardcoded batch is visible."""
    return bool(BENCHMARK_BATCHES) and batch_name == BENCHMARK_BATCHES[0]['name']


def _benchmark_visible_batches(username, items=None):
    """Set of batch names ``username`` may see. Admin-set overrides come from
    benchmark_batches.json; unset batches use the default (first batch only).
    The admin account sees every batch."""
    names = _benchmark_batch_names(items)
    if _is_admin():
        return set(names)
    try:
        overrides = get_batch_visibility(username)
    except Exception:
        overrides = {}
    return {n for n in names
            if overrides.get(n, _benchmark_default_batch_visible(n))}


def _benchmark_item_blocked(item):
    """Whether the current rater may not open/save this item because its batch
    is hidden for them (direct-URL guard behind the overview filter)."""
    username = _benchmark_username()
    return item['batch'] not in _benchmark_visible_batches(username)


def _benchmark_rated_path(item, username, pass_no=1):
    """Destination for a rater's edited matches:
    rated/<user>/<user>-<datasource>-matching/<story>/matches/<same-filename>.
    The two passes live in separate files: pass 1 keeps the unrated filename,
    pass 2 appends `_second-pass` before the extension."""
    name = item['matched_file'].name
    if pass_no == 2:
        p = Path(name)
        name = f"{p.stem}_second-pass{p.suffix}"
    return (
        BENCHMARK_RATED_DIR / username
        / f"{username}-{item['datasource']}-matching"
        / item['story'] / 'matches' / name
    )


def _benchmark_pass_seed(item, username, pass_no):
    """The file a rater's view/save of ``pass_no`` starts from: that pass's own
    rated file if present; for the second pass, otherwise the first-pass rated
    file (so pass 2 opens pre-populated with the rater's first-pass work);
    otherwise the unrated source file."""
    rated_path = _benchmark_rated_path(item, username, pass_no)
    if rated_path.exists():
        return rated_path
    if pass_no == 2:
        first_path = _benchmark_rated_path(item, username, 1)
        if first_path.exists():
            return first_path
    return item['matched_file']


def _benchmark_read_segments(path):
    """Read a matched CSV into the `rated_texts` shape the matching UI expects
    (mirrors get_rated_texts: text, matched_event, further-rating cols, comment)."""
    from helpers.flexible_io import read_parsed_recall_file, read_tabular
    df = read_parsed_recall_file(path)
    try:
        raw_df = read_tabular(path).reset_index(drop=True)
    except Exception:
        raw_df = None
    # (frontend rating key, file column) pairs actually present in the file.
    rating_cols = [(k, _bench_file_col(k)) for k in FURTHER_RATING_COLS
                   if raw_df is not None and _bench_file_col(k) in raw_df.columns]
    has_comment_col = raw_df is not None and 'comment' in raw_df.columns
    has_conf_col = raw_df is not None and BENCH_CONFIDENCE_COL in raw_df.columns
    has_conf1_col = raw_df is not None and BENCH_CONFIDENCE1_COL in raw_df.columns
    has_first_col = raw_df is not None and BENCH_FIRST_PASS_COL in raw_df.columns
    has_second_col = raw_df is not None and BENCH_SECOND_PASS_COL in raw_df.columns

    def _truthy_cell(v):
        return str(v).strip().lower() in ('true', '1', 'yes', 'x')

    segments = []
    for pos, (_, row) in enumerate(df.iterrows()):
        parsed = row.get('recall_in_temporal_order', '')
        matched = row.get('recalled_events', '')
        if not pd.notna(parsed):
            continue
        matched_str = ''
        if pd.notna(matched) and str(matched).strip() not in ['', 'nan']:
            matched_str = str(matched).strip()
            try:
                if ',' in matched_str:
                    parts = [str(int(float(p.strip()))) for p in matched_str.split(',') if p.strip()]
                    matched_str = ','.join(parts)
                else:
                    matched_str = str(int(float(matched_str)))
            except (ValueError, AttributeError):
                pass
        # No .strip(): highlight offsets index into the verbatim recall_text, and
        # recall_in_temporal_order is a verbatim copy of it (flexible_io.py).
        seg = {'text': str(parsed), 'matched_event': matched_str}
        if raw_df is not None and pos < len(raw_df):
            raw_row = raw_df.iloc[pos]
            for key, filecol in rating_cols:
                checked, ranges = _parse_benchmark_rating_cell(raw_row.get(filecol, ''))
                seg[key] = checked
                seg[key + '_ranges'] = ranges
            if has_comment_col:
                cv = raw_row.get('comment', '')
                seg['comment'] = '' if (cv is None or str(cv).strip().lower() == 'nan') else str(cv).strip()
            if has_conf_col:
                cv = raw_row.get(BENCH_CONFIDENCE_COL, '')
                try:
                    seg['confidence'] = int(float(str(cv).strip()))
                except (ValueError, TypeError):
                    seg['confidence'] = ''
            if has_conf1_col:
                cv = raw_row.get(BENCH_CONFIDENCE1_COL, '')
                try:
                    seg['confidence_1'] = int(float(str(cv).strip()))
                except (ValueError, TypeError):
                    seg['confidence_1'] = ''
            if has_first_col:
                seg['first_pass'] = _truthy_cell(raw_row.get(BENCH_FIRST_PASS_COL, ''))
            if has_second_col:
                seg['second_pass'] = _truthy_cell(raw_row.get(BENCH_SECOND_PASS_COL, ''))
        segments.append(seg)
    return segments


def _benchmark_story_events(path):
    """Read a segmented story CSV into the `story_events` list shape ({event, text})."""
    if not path or not Path(path).exists():
        return []
    try:
        from helpers.flexible_io import read_tabular, normalize_story_events_df
        df = normalize_story_events_df(read_tabular(path))
        return _read_story_events_dataframe(df) or []
    except Exception as e:
        print(f"benchmark: error reading story events {path}: {e}")
        return []


def _benchmark_subject_payload(item):
    """Build the /api/subject payload for a benchmark item (matching UI, step3)."""
    username = _benchmark_username()
    active_pass = _benchmark_active_pass(username)
    rated_path = _benchmark_rated_path(item, username, active_pass)
    source = _benchmark_pass_seed(item, username, active_pass)
    try:
        rated_texts = _benchmark_read_segments(source)
    except Exception as e:
        print(f"benchmark: error reading segments {source}: {e}")
        rated_texts = []
    return {
        'subject_id': item['id'],
        'username': username,
        'benchmark': True,
        'benchmark_pass': active_pass,
        'benchmark_meta': {
            'datasource': item['datasource'],
            'story': item['story'],
            'sub_id': item['sub_id'],
        },
        'raw_recall': '',
        'corrected_text': None,
        'parsed_texts': [],
        'rated_texts': rated_texts,
        'story_events': _benchmark_story_events(item['segmented_file']),
        'story_transcript': None,
        'audio_file': None,
        'audio_transcription': None,
        'causal_ratings': None,
        'available_versions': {
            'step1': [], 'step2': [], 'step3': [],
            'story_events': [], 'causal': [], 'story_transcript': [],
        },
        'event_granularities': [],
        'default_export_paths': {
            'corrected': '', 'parsed': '',
            'rated': _path_for_client(rated_path),
            'audio': '', 'story-events': '', 'causal': '',
        },
    }


def _benchmark_save_rated(item, data):
    """Save edited matches for a benchmark item into benchmark/rated/<user>/."""
    err = _benchmark_username_error()
    if err:
        return err

    segments = data.get('segments', [])
    if not segments:
        return jsonify({'error': 'No segments provided'}), 400

    # The active pass is resolved server-side, so a client can only ever write
    # the pass file it is locked to.
    username = _benchmark_username()
    active_pass = _benchmark_active_pass(username)
    rated_path = _benchmark_rated_path(item, username, active_pass)
    rated_path.parent.mkdir(parents=True, exist_ok=True)

    # Drop stale writes: a debounced save built from an older snapshot must not
    # land after a newer one (e.g. the beforeunload beacon). client_seq is a
    # monotonic-ish client timestamp; only strictly-newer writes are accepted.
    seq_key = str(rated_path)
    client_seq = data.get('client_seq')
    if client_seq is not None:
        try:
            client_seq = float(client_seq)
        except (TypeError, ValueError):
            client_seq = None
    # Hold a per-rated-path lock across the whole staleness-check -> read ->
    # build -> atomic-write -> seq-update sequence so concurrent saves for the
    # same file can't both pass the staleness check and lose an update (#3).
    with _bench_save_lock(seq_key):
        if client_seq is not None and client_seq <= _BENCH_SAVE_SEQ.get(seq_key, float('-inf')):
            return jsonify({'success': True, 'skipped': 'stale'})

        from helpers.flexible_io import read_parsed_recall_file, read_tabular
        # For the second pass's first save this seeds from the first-pass rated
        # file (pre-populating it); afterwards the two pass files stay separate.
        seed_file = _benchmark_pass_seed(item, username, active_pass)
        df = read_parsed_recall_file(seed_file)
        if 'recall_in_temporal_order' not in df.columns:
            return jsonify({'error': 'Invalid file format'}), 400

        # Raw read of the seed file (all original columns, before normalization
        # collapses df to recalled_events + recall_in_temporal_order). Used below
        # to carry through extra/admin columns that aren't part of df (#1).
        try:
            raw_seed = read_tabular(seed_file)
        except Exception:
            raw_seed = None

        # Abort (without writing) unless the rater's segment indices are exactly
        # the file's row set {0..len(df)-1} — no gaps, no duplicates. A mismatch
        # means either the unrated file changed under the rater (aligning by
        # integer index would silently drop/misplace data) or the client sent a
        # partial/duplicated index list (a duplicate index would silently
        # overwrite another row's data without raising).
        indices = [s.get('index') for s in segments]
        if (any(not isinstance(i, int) for i in indices)
                or sorted(indices) != list(range(len(df)))):
            return jsonify({
                'error': 'The recall file changed since you opened it, so your latest '
                         'edits were NOT saved. Please take a screenshot of this message '
                         'and email it to the admin.',
                'error_code': 'row_mismatch',
            }), 409

        # Preserve the original benchmark CSV schema (recall_segment, recall_text,
        # story_segments, summary, …) so the rated file matches the unrated shape.
        preserve_schema = None
        sch = _capture_tabular_schema(seed_file)
        if sch and not (set(c.lower() for c in sch['columns']) <= {'recalled_events', 'recall_in_temporal_order'}):
            preserve_schema = sch

        # Apply matched events.
        for segment in segments:
            idx = segment.get('index')
            matched_event = segment.get('matched_event', segment.get('recalled_events', '')).strip()
            if idx is None or not (0 <= idx < len(df)):
                continue
            if matched_event:
                try:
                    parts = [str(int(float(p.strip()))) for p in matched_event.split(',') if p.strip()]
                    matched_event = ', '.join(parts) if parts else ''
                except (ValueError, AttributeError):
                    matched_event = matched_event.strip()
            else:
                matched_event = ''
            df.at[idx, 'recalled_events'] = matched_event

        # Per-segment "further ratings" + comment (only when the toggle is on).
        rating_cols = FURTHER_RATING_COLS
        further_ratings = bool(data.get('further_ratings'))

        def _format_conf(value):
            try:
                return str(int(float(str(value).strip())))
            except (ValueError, TypeError):
                return ''

        if further_ratings:
            rating_by_idx = {r: {} for r in rating_cols}
            comment_by_idx = {}
            conf_by_idx = {}
            conf1_by_idx = {}
            first_by_idx = {}
            second_by_idx = {}
            for segment in segments:
                idx = segment.get('index')
                for r in rating_cols:
                    rating_by_idx[r][idx] = _format_benchmark_rating_cell(bool(segment.get(r)), segment.get(r + '_ranges'))
                comment_by_idx[idx] = str(segment.get('comment') or '').strip()
                conf_by_idx[idx] = _format_conf(segment.get(BENCH_CONFIDENCE_COL))
                conf1_by_idx[idx] = _format_conf(segment.get(BENCH_CONFIDENCE1_COL))
                first_by_idx[idx] = 'TRUE' if segment.get(BENCH_FIRST_PASS_COL) else ''
                second_by_idx[idx] = 'TRUE' if segment.get(BENCH_SECOND_PASS_COL) else ''
            # Rating cats are stored under their benchmark-CSV column name (error -> factual_error).
            for r in rating_cols:
                df[_bench_file_col(r)] = [rating_by_idx[r].get(i, '') for i in range(len(df))]
            df['comment'] = [comment_by_idx.get(i, '') for i in range(len(df))]
            df[BENCH_CONFIDENCE_COL] = [conf_by_idx.get(i, '') for i in range(len(df))]
            df[BENCH_CONFIDENCE1_COL] = [conf1_by_idx.get(i, '') for i in range(len(df))]
            df[BENCH_FIRST_PASS_COL] = [first_by_idx.get(i, '') for i in range(len(df))]
            df[BENCH_SECOND_PASS_COL] = [second_by_idx.get(i, '') for i in range(len(df))]
        else:
            drop = [_bench_file_col(r) for r in rating_cols if _bench_file_col(r) in df.columns]
            for extra in ('comment', BENCH_CONFIDENCE_COL, BENCH_CONFIDENCE1_COL, BENCH_FIRST_PASS_COL, BENCH_SECOND_PASS_COL):
                if extra in df.columns:
                    drop.append(extra)
            if drop:
                df = df.drop(columns=drop)

        for col in df.columns:
            df[col] = df[col].fillna('').astype(str).replace('nan', '')
        df = df.fillna('')

        if 'recalled_events' in df.columns and 'recall_in_temporal_order' in df.columns:
            cols = ['recalled_events', 'recall_in_temporal_order']
            for r in FURTHER_RATING_COLS:
                if _bench_file_col(r) in df.columns:
                    cols.append(_bench_file_col(r))
            for extra in ('comment', BENCH_CONFIDENCE_COL, BENCH_CONFIDENCE1_COL, BENCH_FIRST_PASS_COL, BENCH_SECOND_PASS_COL):
                if extra in df.columns:
                    cols.append(extra)
            df = df[cols]

        out_ext = rated_path.suffix.lower()
        csv_sep = '\t' if out_ext == '.tsv' else ','
        if preserve_schema:
            out_df = pd.DataFrame()
            for col in preserve_schema['columns']:
                if col == preserve_schema.get('index_col'):
                    out_df[col] = list(range(len(df)))
                elif col == preserve_schema.get('recall_col'):
                    out_df[col] = df['recall_in_temporal_order'].values
                elif col == preserve_schema.get('matches_col'):
                    out_df[col] = df['recalled_events'].values
                elif col in df.columns:
                    out_df[col] = df[col].values
                elif (raw_seed is not None and col in raw_seed.columns
                      and len(raw_seed) == len(df)):
                    # Carry through extra source columns (e.g. admin/reference
                    # metadata) that normalization dropped from df, so they
                    # aren't silently blanked on save (#1).
                    out_df[col] = (raw_seed[col].fillna('').astype(str)
                                   .replace('nan', '').values)
                else:
                    out_df[col] = [''] * len(df)
            # Append any rater-input column the unrated schema didn't include (rating
            # categories, comment, confidence, progress flags) so nothing the rater
            # entered is silently dropped on files whose header lacked that column.
            rater_extras = ([_bench_file_col(r) for r in FURTHER_RATING_COLS]
                            + ['comment', BENCH_CONFIDENCE_COL, BENCH_CONFIDENCE1_COL,
                               BENCH_FIRST_PASS_COL, BENCH_SECOND_PASS_COL])
            for extra in rater_extras:
                if extra not in out_df.columns and extra in df.columns:
                    out_df[extra] = df[extra].values
            df = out_df

        # Atomic write: build the file under a sibling temp name, then os.replace so a
        # crash / killed beforeunload-beacon can never leave a truncated rated file.
        tmp_path = rated_path.with_name(rated_path.name + '.tmp')
        try:
            df.to_csv(tmp_path, index=False, sep=csv_sep, encoding='utf-8', na_rep='')
            os.replace(tmp_path, rated_path)
        except Exception:
            try:
                tmp_path.unlink()
            except OSError:
                pass
            raise
        if client_seq is not None:
            _BENCH_SAVE_SEQ[seq_key] = client_seq

    # Copy the segmented story alongside so each rated datasource is self-contained.
    seg_src = item.get('segmented_file')
    if seg_src and Path(seg_src).exists():
        dest_seg = rated_path.parent.parent / Path(seg_src).name
        if not dest_seg.exists():
            try:
                shutil.copy2(seg_src, dest_seg)
            except Exception as e:
                print(f"benchmark: failed to copy segmented story: {e}")

    log_user_edit(username, 'save-rated', item['id'], rated_path.name)
    return jsonify({
        'success': True,
        'message': 'Rated events saved successfully',
        'updated_data': {'rated_texts': _benchmark_read_segments(rated_path)},
    })


def _benchmark_pass_col_state(rated_path, col):
    """Completion of one pass, from its tracking column in one rated file:
    'complete' (every non-blank segment done), 'partial' (some but not all
    done), or 'none' (nothing done / missing file / unreadable or absent
    column)."""
    if not rated_path.exists():
        return 'none'
    try:
        from helpers.flexible_io import read_tabular, read_parsed_recall_file
        raw = read_tabular(rated_path).reset_index(drop=True)
        if len(raw) == 0 or col not in raw.columns:
            return 'none'

        def _truthy(v):
            return str(v).strip().lower() in ('true', '1', 'yes', 'x')

        flags = raw[col].map(_truthy).reset_index(drop=True)

        # Blank recall rows are kept for display but are never rated; exclude them
        # from the completion check so an item that's fully rated apart from
        # blank/trailing rows can still count as done.
        try:
            parsed = read_parsed_recall_file(rated_path).reset_index(drop=True)
            text = parsed['recall_in_temporal_order'].astype(str).str.strip()
            mask = (text != '') & (text.str.lower() != 'nan')
        except Exception:
            mask = None
        if mask is not None and len(mask) == len(flags):
            if not mask.any():
                return 'none'
            flags = flags[mask]

        if flags.all():
            return 'complete'
        return 'partial' if flags.any() else 'none'
    except Exception as e:
        print(f"benchmark: error reading status {rated_path}: {e}")
        return 'none'


def _benchmark_pass_status(item, username):
    """Per-pass completion of a rater's work on a benchmark item. Each pass is
    read from its own rated file (pass 1 = unrated filename, pass 2 =
    `_second-pass` suffix). Returns {'exists', 'first', 'second'}: 'exists' is
    whether either rated file is present; 'first'/'second' are each one of
    'complete', 'partial', or 'none' (see _benchmark_pass_col_state)."""
    first_path = _benchmark_rated_path(item, username, 1)
    second_path = _benchmark_rated_path(item, username, 2)
    return {
        'exists': first_path.exists() or second_path.exists(),
        'first': _benchmark_pass_col_state(first_path, BENCH_FIRST_PASS_COL),
        'second': _benchmark_pass_col_state(second_path, BENCH_SECOND_PASS_COL),
    }


def _benchmark_rated_status(item, username):
    """Overall progress string for a benchmark item: 'not_rated' (no file),
    'rated' (every segment done in both passes), or 'in_progress' otherwise."""
    st = _benchmark_pass_status(item, username)
    if not st['exists']:
        return 'not_rated'
    return 'rated' if (st['first'] == 'complete' and st['second'] == 'complete') else 'in_progress'


@app.route('/benchmark')
@login_required
def benchmark_overview():
    """Benchmark overview: the recall files to rate (and which are done)."""
    if _is_admin():
        return redirect('/admin')
    return render_template('benchmark.html', username=_benchmark_username(),
                           require_auth=REQUIRE_AUTH, version=__version__)


@app.route('/api/benchmark/files')
def api_benchmark_files():
    """List benchmark recall files with their rated status for the overview.
    Only items in batches visible to the current rater are returned."""
    username = _benchmark_username()
    items = _benchmark_scan()
    visible = _benchmark_visible_batches(username, items)
    files = []
    for item in items:
        if item['batch'] not in visible:
            continue
        st = _benchmark_pass_status(item, username)
        rated = st['first'] == 'complete' and st['second'] == 'complete'
        status = 'not_rated' if not st['exists'] else ('rated' if rated else 'in_progress')
        files.append({
            'id': item['id'],
            'datasource': item['datasource'],
            'story': item['story'],
            'sub_id': item['sub_id'],
            'batch': item['batch'],
            'status': status,
            'rated': rated,
            'first_pass': st['first'],
            'second_pass': st['second'],
        })
    return jsonify({'username': username, 'files': files,
                    'active_pass': _benchmark_active_pass(username)})


# --- Admin panel ---------------------------------------------------------
#
# User management for the account named ADMIN_USERNAME ("admin"): add/remove
# raters, change passwords, and switch a rater between pass 1 and pass 2.
# Every endpoint enforces the admin identity server-side (@admin_required);
# account logic lives in narraters.accounts. Removing an account never touches
# the user's rated data or their benchmark_passes.json entry.

@app.route('/admin')
@login_required
def admin_page():
    """Admin panel page; non-admins are sent back to the overview."""
    if not _is_admin():
        return redirect('/benchmark')
    return render_template('admin.html', username=ADMIN_USERNAME,
                           require_auth=REQUIRE_AUTH, version=__version__)


@app.route('/api/admin/users')
@admin_required
def api_admin_list_users():
    """All accounts with their created date and active benchmark pass."""
    users = []
    for name, record in sorted(load_users().items()):
        created = record.get('created', '') if isinstance(record, dict) else ''
        users.append({
            'username': name,
            'created': created,
            'pass': get_benchmark_pass(name),
            'is_admin': name == ADMIN_USERNAME,
        })
    return jsonify({'success': True, 'users': users})


@app.route('/api/admin/users', methods=['POST'])
@admin_required
def api_admin_add_user():
    """Create an account: {username, password}."""
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not is_safe_username(username):
        return jsonify({'success': False,
                        'error': 'Invalid username: use letters, digits, and underscores only (max 32).'}), 400
    if not password.strip():
        return jsonify({'success': False, 'error': 'Password must not be empty.'}), 400
    if not add_user(username, password):
        if username in load_users():
            return jsonify({'success': False, 'error': f'User "{username}" already exists.'}), 409
        return jsonify({'success': False, 'error': 'Could not save the users file.'}), 500
    return jsonify({'success': True})


@app.route('/api/admin/users/<username>', methods=['DELETE'])
@admin_required
def api_admin_remove_user(username):
    """Delete an account. The user's rated data files are left untouched."""
    if not is_safe_username(username):
        return jsonify({'success': False, 'error': 'Invalid username.'}), 400
    if username == ADMIN_USERNAME:
        return jsonify({'success': False, 'error': 'Cannot remove the admin account.'}), 400
    if not remove_user(username):
        if username not in load_users():
            return jsonify({'success': False, 'error': f'User "{username}" not found.'}), 404
        return jsonify({'success': False, 'error': 'Could not save the users file.'}), 500
    return jsonify({'success': True})


@app.route('/api/admin/users/<username>/password', methods=['POST'])
@admin_required
def api_admin_set_password(username):
    """Change an account's password: {password}."""
    if not is_safe_username(username):
        return jsonify({'success': False, 'error': 'Invalid username.'}), 400
    data = request.get_json(silent=True) or {}
    password = data.get('password') or ''
    if not password.strip():
        return jsonify({'success': False, 'error': 'Password must not be empty.'}), 400
    if not set_password(username, password):
        if username not in load_users():
            return jsonify({'success': False, 'error': f'User "{username}" not found.'}), 404
        return jsonify({'success': False, 'error': 'Could not save the users file.'}), 500
    return jsonify({'success': True})


@app.route('/api/admin/users/<username>/pass', methods=['POST'])
@admin_required
def api_admin_set_pass(username):
    """Lock a rater to benchmark pass 1 or 2: {pass: 1|2}."""
    if not is_safe_username(username):
        return jsonify({'success': False, 'error': 'Invalid username.'}), 400
    data = request.get_json(silent=True) or {}
    pass_no = data.get('pass')
    if isinstance(pass_no, bool) or pass_no not in (1, 2):
        return jsonify({'success': False, 'error': 'pass must be 1 or 2.'}), 400
    if not set_benchmark_pass(username, pass_no):
        return jsonify({'success': False, 'error': 'Could not save the passes file.'}), 500
    return jsonify({'success': True, 'pass': pass_no})


@app.route('/api/admin/batches')
@admin_required
def api_admin_batches():
    """Batch overview for the admin page: the batches in order (with item
    counts) and, per non-admin account, each batch's visibility and that
    rater's per-pass progress within it."""
    items = _benchmark_scan()
    batch_names = _benchmark_batch_names(items)
    totals = {name: 0 for name in batch_names}
    for item in items:
        totals[item['batch']] += 1

    users = []
    for name in sorted(load_users()):
        if name == ADMIN_USERNAME:
            continue
        overrides = get_batch_visibility(name)
        progress = {bn: {'visible': overrides.get(bn, _benchmark_default_batch_visible(bn)),
                         'total': totals[bn],
                         'first_done': 0, 'first_partial': 0,
                         'second_done': 0, 'second_partial': 0}
                    for bn in batch_names}
        for item in items:
            st = _benchmark_pass_status(item, name)
            cell = progress[item['batch']]
            if st['first'] == 'complete':
                cell['first_done'] += 1
            elif st['first'] == 'partial':
                cell['first_partial'] += 1
            if st['second'] == 'complete':
                cell['second_done'] += 1
            elif st['second'] == 'partial':
                cell['second_partial'] += 1
        users.append({'username': name, 'pass': get_benchmark_pass(name),
                      'batches': progress})

    return jsonify({'success': True,
                    'batches': [{'name': bn, 'total_items': totals[bn]}
                                for bn in batch_names],
                    'users': users})


@app.route('/api/admin/users/<username>/batches', methods=['POST'])
@admin_required
def api_admin_set_batch(username):
    """Show or hide one benchmark batch for a rater: {batch, visible}."""
    if not is_safe_username(username):
        return jsonify({'success': False, 'error': 'Invalid username.'}), 400
    if username not in load_users():
        return jsonify({'success': False, 'error': f'User "{username}" not found.'}), 404
    data = request.get_json(silent=True) or {}
    batch = data.get('batch')
    visible = data.get('visible')
    if batch not in _benchmark_batch_names():
        return jsonify({'success': False, 'error': 'Unknown batch.'}), 400
    if not isinstance(visible, bool):
        return jsonify({'success': False, 'error': 'visible must be true or false.'}), 400
    if not set_batch_visible(username, batch, visible):
        return jsonify({'success': False, 'error': 'Could not save the batches file.'}), 500
    return jsonify({'success': True, 'batch': batch, 'visible': visible})


@app.route('/')
def index():
    """Main dashboard. Requires a configured pipeline; rater name is optional
    here (it is collected on the pipeline-config page)."""
    if BENCHMARK_MODE:
        return redirect('/benchmark')
    pipeline_file = _pipeline_config_path()
    if not pipeline_file.exists():
        print("No pipeline config found, redirecting to pipeline configuration page")
        return redirect('/pipeline-config')

    # Dashboard loads data via /api/dashboard/panels; avoid duplicate filesystem scan here.
    return render_template('index.html')


@app.route('/pipeline-config')
@login_required
def pipeline_config():
    """Pipeline configuration page."""
    if BENCHMARK_MODE:
        return redirect('/benchmark')
    print("Rendering pipeline configuration page")
    try:
        template_path = PACKAGE_ROOT / 'templates' / 'pipeline-config.html'
        if not template_path.exists():
            return f"Error: Template not found at {template_path}", 500
        return render_template('pipeline-config.html')
    except Exception as e:
        print(f"Error rendering pipeline-config.html: {e}")
        import traceback
        traceback.print_exc()
        return f"Error loading pipeline configuration page: {e}", 500


@app.route('/api/subjects')
def api_subjects():
    """API endpoint to get all subjects."""
    subjects = get_all_subjects(get_pipeline_config())
    return jsonify(subjects)


@app.route('/api/subjects/status')
def api_subjects_status():
    """API endpoint to get all subjects with their processing step statuses."""
    items_with_status = get_all_items_with_status()
    return jsonify(items_with_status)


@app.route('/api/dashboard/panels')
def api_dashboard_panels():
    """API endpoint for dashboard data grouped by source (input path).
    Each panel has steps that share the same input path, and items (rows) discovered from that source.
    """
    data = get_dashboard_panels()
    return jsonify(data)


def _parse_file_version_query_args():
    """Split query params: story event .xlsx vs recall pipeline vs causal rating .xlsx (legacy file_version supported)."""
    story_events_file = (request.args.get('story_events_file') or '').strip() or None
    recall_fv = (request.args.get('recall_file_version') or '').strip() or None
    story_transcript_file = (request.args.get('story_transcript_file') or '').strip() or None
    causal_rating_file = (request.args.get('causal_rating_file') or '').strip() or None
    legacy = (request.args.get('file_version') or '').strip() or None
    if legacy and request.args.get('story_events_file') is None and request.args.get('recall_file_version') is None:
        if legacy.endswith('.xlsx') and '_events' in legacy:
            story_events_file = legacy
        elif legacy.endswith('.xlsx') and 'causal' in legacy:
            causal_rating_file = legacy
        else:
            recall_fv = legacy
    return story_events_file, recall_fv, story_transcript_file, causal_rating_file


@app.route('/api/subject/<subj_id>')
def api_subject(subj_id):
    """API endpoint to get all data for a subject."""
    if BENCHMARK_MODE:
        item = _benchmark_item_for_id(subj_id)
        if item:
            err = _benchmark_username_error()
            if err:
                return err
            if _benchmark_item_blocked(item):
                return jsonify({
                    'error': 'This item is not available to you (its batch is '
                             'hidden). Please return to the overview.',
                    'error_code': 'batch_hidden',
                }), 403
            return jsonify(_benchmark_subject_payload(item))
        # In benchmark mode we must NOT fall through to the generic pipeline
        # lookups with a bench__… slug (wrong tree + wrong format, yielding an
        # empty/nonsensical payload). The unrated file was likely moved/renamed
        # under the rater; fail loudly, mirroring the save endpoint.
        print(f"benchmark: read failed — could not resolve item for slug {subj_id!r}")
        return jsonify({
            'error': 'Could not locate this benchmark item. The recall file may '
                     'have been moved or renamed — please return to the overview '
                     'and pick a file again.',
            'error_code': 'item_unresolved',
        }), 409
    story_events_file, recall_fv, story_transcript_file, causal_rating_file = _parse_file_version_query_args()
    if not causal_rating_file and recall_fv and str(recall_fv).endswith('.xlsx') and 'causal' in str(recall_fv):
        causal_rating_file, recall_fv = recall_fv, None

    cr = get_causal_ratings(subj_id, causal_rating_file)
    source_file = (cr.get('source_file', '') or '') if cr else ''
    
    story_events = get_story_events(subj_id, story_events_file, is_story=False)
    story_transcript = get_story_transcript(subj_id, is_story=False, transcript_file=story_transcript_file)
    
    # Fallback transcript: try the story name derived from the subject ID
    story_name = get_story_name_from_subject_id(subj_id)
    if story_name and story_name != subj_id and not story_transcript:
        story_transcript = get_story_transcript(story_name, is_story=True)
    
    data = {
        'subject_id': subj_id,
        'username': session.get('username', 'human'),
        'raw_recall': get_raw_recall_text(subj_id),
        'corrected_text': get_corrected_text(subj_id, recall_fv),
        'parsed_texts': get_parsed_texts(subj_id, recall_fv),
        'rated_texts': get_rated_texts(subj_id, recall_fv),
        'story_events': story_events,
        'story_transcript': story_transcript,
        'audio_file': get_audio_file(subj_id, is_story=False),
        'audio_transcription': get_audio_transcription(subj_id, is_story=False),
        'causal_ratings': cr,
        'available_versions': get_available_file_versions(subj_id, is_story=False),
        'event_granularities': get_all_event_granularities(subj_id),
        'default_export_paths': {
            'corrected': get_default_export_path(
                subj_id, 'corrected', is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'parsed': get_default_export_path(
                subj_id, 'parsed', is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'rated': get_default_export_path(
                subj_id, 'rated', is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'audio': get_default_export_path(
                subj_id, 'audio', is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'story-events': get_default_export_path(
                subj_id, 'story-events', is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'causal': get_default_export_path(
                subj_id, 'causal', source_file=source_file, is_story=False,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
        },
    }
    
    return jsonify(data)


@app.route('/api/story/<story_name>')
def api_story(story_name):
    """API endpoint to get all data for a story."""
    story_events_file, recall_fv, story_transcript_file, causal_rating_file = _parse_file_version_query_args()
    if not causal_rating_file and recall_fv and str(recall_fv).endswith('.xlsx') and 'causal' in str(recall_fv):
        causal_rating_file, recall_fv = recall_fv, None

    min_events = 0
    cr = get_causal_ratings(story_name, causal_rating_file)
    if causal_rating_file and 'causal' in str(causal_rating_file) and cr and cr.get('pairs'):
        for p in cr['pairs']:
            min_events = max(min_events, p.get('event_A', 0), p.get('event_B', 0))
    
    source_file = (cr.get('source_file', '') or '') if cr else ''
    data = {
        'subject_id': story_name,
        'story_name': story_name,
        'username': session.get('username', 'human'),
        'story_events': get_story_events(story_name, story_events_file, is_story=True, min_event_count=min_events),
        'story_transcript': get_story_transcript(story_name, is_story=True, transcript_file=story_transcript_file),
        'audio_file': get_audio_file(story_name, is_story=True),
        'audio_transcription': get_audio_transcription(story_name, is_story=True),
        'causal_ratings': cr,
        'available_versions': get_available_file_versions(story_name, is_story=True),
        'event_granularities': get_all_event_granularities(story_name),
        'default_export_paths': {
            'corrected': get_default_export_path(
                story_name, 'corrected', is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'parsed': get_default_export_path(
                story_name, 'parsed', is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'rated': get_default_export_path(
                story_name, 'rated', is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'audio': get_default_export_path(
                story_name, 'audio', is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'story-events': get_default_export_path(
                story_name, 'story-events', is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
            'causal': get_default_export_path(
                story_name, 'causal', source_file=source_file, is_story=True,
                story_events_file=story_events_file, recall_file_version=recall_fv,
                story_transcript_file=story_transcript_file,
            ),
        },
    }
    
    return jsonify(data)


@app.route('/api/subject/<subj_id>/versions')
def api_subject_versions(subj_id):
    """API endpoint to get available file versions for a subject."""
    versions = get_available_file_versions(subj_id)
    return jsonify(versions)


@app.route('/subject/<subj_id>')
@login_required
def subject_view(subj_id):
    """View page for a specific subject."""
    step = request.args.get('step', None)
    if BENCHMARK_MODE:
        item = _benchmark_item_for_id(subj_id)
        if item:
            if _benchmark_item_blocked(item):
                return redirect('/benchmark')
            return render_template(
                'subject.html', subject_id=subj_id, step=step,
                benchmark=True,
                benchmark_meta={
                    'datasource': item['datasource'],
                    'story': item['story'],
                    'sub_id': item['sub_id'],
                },
                benchmark_file=item['matched_file'].relative_to(BENCHMARK_UNRATED_DIR).as_posix(),
            )
    return render_template('subject.html', subject_id=subj_id, step=step)


@app.route('/story/<story_name>')
@login_required
def story_view(story_name):
    """View page for a specific story."""
    step = request.args.get('step', None)
    return render_template('subject.html', subject_id=story_name, step=step, is_story=True)


@app.route('/story/<story_name>/step0')
@login_required
def story_step0(story_name):
    """View page for story step 0: Story events segmentation."""
    return render_template('subject.html', subject_id=story_name, step='step0', is_story=True)


@app.route('/story/<story_name>/step0_story_audio')
@login_required
def story_step0_audio(story_name):
    """View page for story step 0 audio: Story audio transcription."""
    return render_template('subject.html', subject_id=story_name, step='step0_story_audio', is_story=True)


@app.route('/story/<story_name>/step0_causal')
@login_required
def story_step0_causal(story_name):
    """View page for story causal rating."""
    return render_template('subject.html', subject_id=story_name, step='step0_causal', is_story=True)


@app.route('/subject/<subj_id>/step0_causal')
@login_required
def subject_step0_causal(subj_id):
    """View page for subject causal rating."""
    return render_template('subject.html', subject_id=subj_id, step='step0_causal')


@app.route('/subject/<subj_id>/step1')
def subject_step1(subj_id):
    """View page for step 1: Spell and grammar check."""
    return render_template('subject.html', subject_id=subj_id, step='step1')


@app.route('/subject/<subj_id>/step2')
def subject_step2(subj_id):
    """View page for step 2: Recall text parsing/segmenting."""
    return render_template('subject.html', subject_id=subj_id, step='step2')


@app.route('/subject/<subj_id>/step0')
def subject_step0(subj_id):
    """View page for step 0: Story events segmentation."""
    return render_template('subject.html', subject_id=subj_id, step='step0')


@app.route('/subject/<subj_id>/step0_audio')
def subject_step0_audio(subj_id):
    """View page for step 0 audio: Audio transcription."""
    return render_template('subject.html', subject_id=subj_id, step='step0_audio')


@app.route('/subject/<subj_id>/step0_story_audio')
def subject_step0_story_audio(subj_id):
    """View page for step 0 story audio: Story audio transcription."""
    return render_template('subject.html', subject_id=subj_id, step='step0_story_audio')


@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files (JS, CSS, audio, etc.) from project static folder.

    Uses Flask's send_from_directory which rejects path-traversal attempts
    (``..`` segments and absolute paths).
    """
    from flask import send_from_directory
    from werkzeug.exceptions import NotFound

    try:
        return send_from_directory(str(PACKAGE_ROOT / 'static'), filename)
    except NotFound:
        return "File not found", 404


@app.route('/favicon.ico')
def favicon():
    """Serve the favicon at the conventional root path that browsers request
    automatically, in addition to the explicit <link> tags in the templates."""
    from flask import send_from_directory
    return send_from_directory(str(PACKAGE_ROOT / 'static' / 'favicon'), 'favicon.ico')


@app.route('/subject/<subj_id>/step3')
def subject_step3(subj_id):
    """View page for step 3: Recall rating (match to story events)."""
    return render_template('subject.html', subject_id=subj_id, step='step3')


@app.route('/api/subject/<subj_id>/matrix-plot', methods=['GET', 'POST'])
def get_matrix_plot(subj_id):
    """Generate matrix plot visualization for recall rating step."""
    return get_matrix_plot_impl(subj_id, is_story=False)


@app.route('/api/story/<story_name>/matrix-plot', methods=['GET', 'POST'])
def get_story_matrix_plot(story_name):
    """Generate matrix plot visualization for story recall rating step."""
    return get_matrix_plot_impl(story_name, is_story=True)


def get_matrix_plot_impl(item_id, is_story=False):
    """Generate matrix plot visualization for recall rating step."""
    try:
        import sys
        import io
        import base64
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        
        story_events_file, recall_fv, _story_tf, _causal_rf = _parse_file_version_query_args()
        
        # Check if POST request with current ratings (for real-time updates)
        current_ratings = None
        if request.method == 'POST' and request.is_json:
            data = request.get_json()
            current_ratings = data.get('ratings', None)
        
        # Get rated texts and story events
        if is_story:
            # For stories, we need to get rated texts from subjects that recalled this story
            # For now, return error as story-level matrix plot may need different logic
            return jsonify({'error': 'Matrix plot for stories not yet implemented'}), 404
        else:
            # Use current ratings if provided (for real-time updates), otherwise get from file
            if current_ratings:
                rated_texts = current_ratings
            else:
                rated_texts = get_rated_texts(item_id, recall_fv)
            story_events = get_story_events(item_id, story_events_file, is_story=False)
        
        if not rated_texts or not story_events:
            return jsonify({'error': 'Missing rated texts or story events'}), 404
        
        # Create DataFrame from rated texts
        import pandas as pd
        try:
            # Ensure we have a list of dictionaries
            if not isinstance(rated_texts, list):
                return jsonify({'error': 'Invalid ratings format: expected list'}), 400
            
            # Validate and normalize the data
            normalized_ratings = []
            for i, rating in enumerate(rated_texts):
                if not isinstance(rating, dict):
                    return jsonify({'error': f'Invalid rating format at index {i}: expected dict'}), 400
                
                # Ensure matched_event is a string (handle None, empty, or other types)
                matched_event = rating.get('matched_event', '')
                if matched_event is None:
                    matched_event = ''
                else:
                    matched_event = str(matched_event).strip()
                
                normalized_ratings.append({
                    'text': str(rating.get('text', '')),
                    'matched_event': matched_event
                })
            
            rated_df = pd.DataFrame(normalized_ratings)
            
            # Rename 'matched_event' to 'recalled_events' for compatibility with plot function
            if 'matched_event' in rated_df.columns:
                rated_df['recalled_events'] = rated_df['matched_event']
            else:
                return jsonify({'error': 'Missing matched_event column in ratings data'}), 400
            
            # Get number of events
            num_events = len(story_events)
            
            if num_events == 0:
                return jsonify({'error': 'No story events available'}), 400
            
            # Import plot function
            sys.path.insert(0, str(PACKAGE_ROOT))
            sys.path.insert(0, str(PACKAGE_ROOT / 'helpers'))
            from plot_matrix_comparison import create_matrix_from_ratings
            
            # Create matrix from ratings
            model_matrix = create_matrix_from_ratings(rated_df, num_events)
            
            if model_matrix is None:
                return jsonify({'error': 'Failed to create matrix from ratings data'}), 500
                
        except Exception as df_error:
            print(f"Error processing ratings data: {df_error}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Error processing ratings data: {str(df_error)}'}), 500
        
        # Create plot (model only, no human comparison)
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
        import matplotlib.pyplot as plt
        import numpy as np
        
        fig, ax = plt.subplots(1, 1, figsize=(12, 8))
        
        # Convert to binary for display
        binary_matrix = (model_matrix > 0).astype(int)
        
        im = ax.imshow(binary_matrix, aspect='auto', cmap='viridis', vmin=0, vmax=1, interpolation='nearest')
        ax.set_title(f'Recall Segment vs Story Events Matrix - {item_id}', 
                     fontsize=14, fontweight='bold')
        ax.set_xlabel('Recall Segment', fontsize=12)
        ax.set_ylabel('Story Event', fontsize=12)
        
        # Set ticks
        max_events, max_segments = binary_matrix.shape
        if max_events <= 50:
            ax.set_yticks(range(max_events))
            ax.set_yticklabels(range(1, max_events + 1))
        else:
            step = max(1, max_events // 20)
            ax.set_yticks(range(0, max_events, step))
            ax.set_yticklabels(range(1, max_events + 1, step))
        
        if max_segments <= 50:
            ax.set_xticks(range(max_segments))
            ax.set_xticklabels(range(1, max_segments + 1), rotation=45, ha='right')
        else:
            step = max(1, max_segments // 20)
            ax.set_xticks(range(0, max_segments, step))
            ax.set_xticklabels(range(1, max_segments + 1, step), rotation=45, ha='right')
        
        plt.colorbar(im, ax=ax, label='Match (1=Yes, 0=No)')
        plt.tight_layout()
        
        # Convert plot to base64 image
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        # Encode image as base64
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()
        
        return jsonify({
            'success': True,
            'image': f'data:image/png;base64,{img_base64}'
        })
        
    except Exception as e:
        print(f"Error generating matrix plot: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/subject/<subj_id>/export-default-path', methods=['GET'])
def get_subject_export_default_path(subj_id):
    """Return the default export path for a given step."""
    step = request.args.get('step', '')
    source_file = request.args.get('source_file', '') or None
    story_events_file, recall_fv, story_transcript_file, _causal_rf = _parse_file_version_query_args()
    path = get_default_export_path(
        subj_id, step, source_file=source_file, is_story=False,
        story_events_file=story_events_file, recall_file_version=recall_fv,
        story_transcript_file=story_transcript_file,
    )
    if not path:
        return jsonify({'error': f'Unknown step: {step}'}), 400
    return jsonify({'path': path, 'username': session.get('username', 'human')})


@app.route('/api/story/<story_name>/export-default-path', methods=['GET'])
def get_story_export_default_path(story_name):
    """Return the default export path for causal matrix (story context)."""
    step = request.args.get('step', 'causal')
    source_file = request.args.get('source_file', '') or None
    story_events_file, recall_fv, story_transcript_file, _causal_rf = _parse_file_version_query_args()
    path = get_default_export_path(
        story_name, step, source_file=source_file, is_story=True,
        story_events_file=story_events_file, recall_file_version=recall_fv,
        story_transcript_file=story_transcript_file,
    )
    if not path:
        return jsonify({'error': f'Unknown step: {step}'}), 400
    return jsonify({'path': path, 'username': session.get('username', 'human')})


@app.route('/api/subject/<subj_id>/save', methods=['POST'])
def save_corrected_text(subj_id):
    """Save corrected text to file."""
    try:
        # Get JSON data
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        corrected_text = data.get('corrected_text', '')
        
        # Handle None or empty text
        if corrected_text is None:
            corrected_text = ''
        
        # Use custom output path if provided and valid
        output_path_str = data.get('output_path', '')
        file_path = _resolve_and_validate_output_path(output_path_str)
        if file_path is None:
            # Use pipeline output path when configured
            output_dir = get_output_dir_for_step_type('sentenceCorrect') or RECALL_CORRECTED_DIR
            output_dir.mkdir(parents=True, exist_ok=True)
            edit_suffix = get_edit_suffix()
            file_path = output_dir / f"{subj_id}{edit_suffix}.txt"
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)

        # Header line is the on-disk filename so downstream tools can re-derive
        # the subject id from the file alone.
        header_filename = file_path.name

        # Write the corrected text
        # Format: first line is filename, rest is text
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"{header_filename}\n")
                f.write(str(corrected_text))
                f.flush()  # Ensure data is written
            
            # Verify file was written
            if not file_path.exists():
                return jsonify({'error': 'File was not created. Check directory permissions.'}), 500
            
            print(f"Saved corrected text to {file_path} ({len(corrected_text)} characters)")
            log_user_edit(session.get('username', 'unknown'), 'save-corrected', subj_id, file_path.name)
            
            # FORWARD CASCADE: Tab 1 → Tab 2 → Tab 3
            # Automatically re-parse the corrected text and update parsed file
            try:
                import importlib.util
                import sys
                sys.path.insert(0, str(SCRIPTS_DIR))
                sys.path.insert(0, str(PACKAGE_ROOT))
                parse_file = SCRIPTS_DIR / '4_parse-texts.py'
                spec = importlib.util.spec_from_file_location("parse_texts", parse_file)
                parse_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(parse_module)
                parse_text_func = parse_module.parse_text
                create_dataframe_func = parse_module.create_dataframe
                
                # Parse the corrected text
                parsed_units = parse_text_func(corrected_text)
                
                # Create dataframe
                parsed_df = create_dataframe_func(parsed_units)
                
                # Save to parsed file with _{username}-edit suffix (use pipeline output path)
                parsed_dir = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
                parsed_dir.mkdir(parents=True, exist_ok=True)
                parsed_file = parsed_dir / f"{subj_id}_parsed{edit_suffix}.xlsx"
                
                parsed_df.to_excel(parsed_file, index=False, engine='openpyxl', na_rep='')
                print(f"Auto-updated parsed file: {parsed_file} ({len(parsed_units)} segments)")
                
                # Automatically re-rate the parsed segments
                try:
                    rated_df = _auto_rate_parsed_df(subj_id, parsed_df)
                    if rated_df is not None:
                        rated_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
                        rated_dir.mkdir(parents=True, exist_ok=True)
                        rated_file = rated_dir / f"{subj_id}_rate-recall.xlsx"
                        rated_df.to_excel(rated_file, index=False, engine='openpyxl', na_rep='')
                        matched_count = sum(1 for v in rated_df['recalled_events'] if str(v).strip())
                        print(f"Auto-updated rated file: {rated_file} ({matched_count} segments matched)")
                    else:
                        print(f"Warning: Story events file not found for {subj_id}, skipping auto-rating")
                except Exception as rate_error:
                    print(f"Warning: Could not auto-update rated file: {rate_error}")
                    import traceback
                    traceback.print_exc()
                    # Don't fail the save if rating fails
                
            except Exception as parse_error:
                print(f"Warning: Could not auto-update parsed file: {parse_error}")
                # Don't fail the save if parsing fails
            
            # Get updated data to return to client
            updated_parsed = get_parsed_texts(subj_id)
            updated_rated = get_rated_texts(subj_id)
            
            return jsonify({
                'success': True, 
                'message': 'Text saved successfully. Parsed and rated files updated automatically.',
                'updated_data': {
                    'corrected_text': corrected_text,
                    'parsed_texts': updated_parsed,
                    'rated_texts': updated_rated
                }
            })
            
        except PermissionError as perm_error:
            print(f"Permission error saving corrected text: {perm_error}")
            return jsonify({'error': 'Permission denied. Check file and directory permissions.'}), 500
        except IOError as io_error:
            error_msg = str(io_error)
            print(f"IO Error saving corrected text: {io_error}")
            if 'locked' in error_msg.lower() or 'busy' in error_msg.lower():
                return jsonify({'error': 'File is locked or in use. Please close any programs using this file.'}), 500
            return jsonify({'error': f'File write error: {error_msg}'}), 500
        except Exception as write_error:
            print(f"Error writing file: {write_error}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Write error: {str(write_error)}'}), 500
            
    except Exception as e:
        print(f"Error saving corrected text: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500


def _capture_tabular_schema(path):
    """Read an existing tabular file's header and classify column roles, so a save can
    write back in the SAME shape (preserving e.g. ``recall_segment, recall_text,
    story_segments, summary…``) instead of collapsing to the canonical two columns.
    Returns {'columns', 'index_col', 'recall_col', 'matches_col'} or None.
    """
    try:
        from helpers.flexible_io import (
            read_tabular, detect_parsed_recall_columns, _norm_name, _NON_RATING_COL_NAMES,
        )
        raw = read_tabular(path)
        cols = [str(c) for c in raw.columns]
        det = detect_parsed_recall_columns(raw)
        recall_col = det[1] if det else None
        matches_col = det[0] if (det and det[0] != det[1]) else None
        index_col = None
        _comment_rating = {"comment", "comments", "note", "notes",
                           "summary", "error", "confabulation", "inference", "opinion", "meta"}
        for c in cols:
            if c in (recall_col, matches_col):
                continue
            nm = _norm_name(c)
            if nm in _NON_RATING_COL_NAMES and nm not in _comment_rating:
                index_col = c
                break
        return {'columns': cols, 'index_col': index_col, 'recall_col': recall_col, 'matches_col': matches_col}
    except Exception:
        return None


@app.route('/api/subject/<subj_id>/save-rated', methods=['POST'])
def save_rated_events(subj_id):
    """Save edited matched events for rated texts."""
    try:
        data = request.get_json()

        if BENCHMARK_MODE:
            item = _benchmark_item_for_id(subj_id)
            if item:
                if _benchmark_item_blocked(item):
                    return jsonify({
                        'error': 'This item is not available to you (its batch '
                                 'is hidden) — your changes were not saved.',
                        'error_code': 'batch_hidden',
                    }), 403
                return _benchmark_save_rated(item, data)
            # In benchmark mode we must NOT fall through to the generic pipeline
            # save (wrong tree + wrong format). The unrated file was likely
            # moved/renamed under the rater; fail loudly so nothing is misrouted.
            print(f"benchmark: save failed — could not resolve item for slug {subj_id!r}")
            return jsonify({
                'error': 'Could not locate this benchmark item to save. The recall '
                         'file may have been moved or renamed — your latest changes '
                         'were not saved.',
                'error_code': 'item_unresolved',
            }), 409

        segments = data.get('segments', [])

        if not segments:
            return jsonify({'error': 'No segments provided'}), 400
        
        # Use custom output path if provided and valid
        output_path_str = data.get('output_path', '')
        file_path = _resolve_and_validate_output_path(output_path_str)
        if file_path is None:
            rated_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
            rated_dir.mkdir(parents=True, exist_ok=True)
            edit_suffix = get_edit_suffix()
            file_path = rated_dir / f"{subj_id}_rate-recall{edit_suffix}.xlsx"
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)
        
        from helpers.flexible_io import read_parsed_recall_file

        # Read existing file if it exists, otherwise read from original rated file to preserve existing ratings
        seed_file = None
        if file_path.exists():
            df = read_parsed_recall_file(file_path)
            seed_file = file_path
            print(f"Reading from existing user-edit file: {file_path.name}")
        else:
            # Try to read from original rated file first (to preserve existing ratings)
            orig_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
            original_rated = orig_dir / f"{subj_id}_rate-recall.xlsx"
            if not original_rated.exists():
                candidates = list_subject_rated_recall_source_files(orig_dir, subj_id)
                if candidates:
                    original_rated = candidates[0]
            if not original_rated.exists():
                # Recognise alternative-named / csv rated files (story-...-recall-matched.csv)
                try:
                    from helpers.step_files import find_step_files
                    recs = find_step_files(orig_dir, 'textMatching', subj_id)
                    if recs:
                        original_rated = recs[0]
                except Exception:
                    pass
            if original_rated.exists():
                df = read_parsed_recall_file(original_rated)
                seed_file = original_rated
                print(f"Reading from original rated file: {original_rated.name}")
            else:
                # Fallback: create from parsed file (prioritize user-edit, then canonical,
                # then any method-suffixed parsed output)
                parsed_dir = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
                parsed_edit = find_best_edit_file(parsed_dir, subj_id, '_parsed', '.xlsx')
                parsed_original = parsed_dir / f"{subj_id}_parsed.xlsx"
                if parsed_edit:
                    parsed_file = parsed_edit
                elif parsed_original.exists():
                    parsed_file = parsed_original
                else:
                    method_cands = list_subject_parsed_source_files(parsed_dir, subj_id)
                    parsed_file = method_cands[0] if method_cands else parsed_original
                if not parsed_file.exists():
                    # Recognise alternative-named / csv parsed files (story-...-parsed.csv)
                    try:
                        from helpers.step_files import find_step_files
                        precs = find_step_files(parsed_dir, 'textParsing', subj_id)
                        if precs:
                            parsed_file = precs[0]
                    except Exception:
                        pass

                if parsed_file.exists():
                    df = read_parsed_recall_file(parsed_file)
                    seed_file = parsed_file
                    print(f"Creating from parsed file: {parsed_file.name}")
                else:
                    return jsonify({'error': 'Rated file not found and cannot be created from parsed file'}), 404
        
        if 'recall_in_temporal_order' not in df.columns:
            return jsonify({'error': 'Invalid file format'}), 400

        # If the export target is a CSV/TSV and we seeded from an existing file with a
        # non-canonical schema (e.g. recall_segment, recall_text, story_segments, …),
        # remember that schema so the save writes back in the same shape rather than
        # collapsing to recalled_events/recall_in_temporal_order.
        preserve_schema = None
        if str(file_path).lower().endswith(('.csv', '.tsv')) and seed_file is not None:
            sch = _capture_tabular_schema(seed_file)
            if sch and not (set(c.lower() for c in sch['columns']) <= {'recalled_events', 'recall_in_temporal_order'}):
                preserve_schema = sch

        # Update matched events based on segment indices
        updated_count = 0
        print(f"Received {len(segments)} segments to update")
        print(f"Dataframe has {len(df)} rows")
        
        for segment in segments:
            idx = segment.get('index')
            # Frontend sends 'matched_event', but also check 'recalled_events' for compatibility
            matched_event = segment.get('matched_event', segment.get('recalled_events', '')).strip()
            
            print(f"Processing segment {idx}: matched_event='{matched_event}'")
            
            if idx is not None and 0 <= idx < len(df):
                # Get current value for comparison
                current_value = str(df.at[idx, 'recalled_events']).strip() if pd.notna(df.at[idx, 'recalled_events']) else ''
                print(f"  Current value at index {idx}: '{current_value}'")
                
                # Ensure event values are formatted as integers (comma-separated if multiple)
                if matched_event and matched_event.strip():
                    # Parse and validate event numbers
                    try:
                        event_parts = [str(int(float(part.strip()))) for part in matched_event.split(',') if part.strip()]
                        matched_event = ', '.join(event_parts) if event_parts else ''
                    except (ValueError, AttributeError) as e:
                        # If parsing fails, keep original but log warning
                        print(f"Warning: Could not parse event value '{matched_event}' for segment {idx}: {e}")
                        matched_event = matched_event.strip()  # Keep original if parsing fails
                else:
                    matched_event = ''  # Ensure empty string for blank values
                
                # Update the dataframe
                df.at[idx, 'recalled_events'] = matched_event
                print(f"  Updated index {idx} to: '{matched_event}' (type: {type(matched_event)})")
                updated_count += 1
            else:
                print(f"Warning: Invalid segment index {idx} (dataframe has {len(df)} rows)")
        
        print(f"Updated {updated_count} segments out of {len(segments)} provided")

        # Per-segment "further ratings" (summary | error | confabulation). Each
        # checkbox becomes its own column, added only when the rater enabled the
        # "Further ratings" toggle so ordinary matching files are left unchanged;
        # stale columns are dropped when the toggle is off.
        rating_cols = FURTHER_RATING_COLS
        further_ratings = bool(data.get('further_ratings'))

        def _format_rating_cell(checked, spans):
            """Cell value for a rating: the quoted text fragments selected for it
            (``'frag a'; 'frag b'``) when present, else ``TRUE`` when only the box
            is checked, else blank."""
            if not checked:
                return ''
            frags = [str(s).strip() for s in (spans or []) if str(s).strip()]
            if frags:
                return '; '.join("'" + f.replace("'", "’") + "'" for f in frags)
            return 'TRUE'

        if further_ratings:
            rating_by_idx = {r: {} for r in rating_cols}
            comment_by_idx = {}
            for segment in segments:
                idx = segment.get('index')
                for r in rating_cols:
                    rating_by_idx[r][idx] = _format_rating_cell(
                        bool(segment.get(r)), segment.get(r + '_spans'))
                comment_by_idx[idx] = str(segment.get('comment') or '').strip()
            for r in rating_cols:
                df[r] = [rating_by_idx[r].get(i, '') for i in range(len(df))]
            # Free-text per-segment comment (its own column).
            df['comment'] = [comment_by_idx.get(i, '') for i in range(len(df))]
        else:
            drop = [r for r in rating_cols if r in df.columns]
            if 'comment' in df.columns:
                drop.append('comment')
            if drop:
                df = df.drop(columns=drop)

        # Clean data before saving - ensure all columns are properly formatted
        # IMPORTANT: Do this AFTER updates to preserve the values we just set
        for col in df.columns:
            if col == 'recall_in_temporal_order':
                df[col] = df[col].fillna('').astype(str)
            elif col == 'recalled_events':
                # For recalled_events, preserve the values we just set
                # Convert to string, but preserve actual values (including empty strings)
                df[col] = df[col].astype(str)
                # Replace 'nan' string with empty string (but keep actual values like '5', '1,2', etc.)
                df[col] = df[col].replace('nan', '')
                # Fill NaN values with empty string
                df[col] = df[col].fillna('')
                # Clean up: only strip whitespace, preserve all actual values
                # Don't convert valid values to empty string
                df[col] = df[col].apply(lambda x: '' if (pd.isna(x) or str(x).strip().lower() in ['nan', 'none']) else str(x).strip())
            else:
                df[col] = df[col].fillna('').astype(str)
        
        # Final check: ensure no NaN values remain (but preserve our updates)
        df = df.fillna('')
        
        # Debug: Print a few sample values before saving
        print("Sample values before saving:")
        for i in range(min(3, len(df))):
            print(f"  Row {i}: recalled_events='{df.iloc[i].get('recalled_events', 'N/A')}'")
        
        # Save updated dataframe
        try:
            # Close any existing file handles first
            import gc
            gc.collect()
            
            # Ensure proper column order (matched events first, then recall text,
            # then any optional "further ratings" columns).
            if 'recalled_events' in df.columns and 'recall_in_temporal_order' in df.columns:
                cols = ['recalled_events', 'recall_in_temporal_order']
                for r in FURTHER_RATING_COLS + ('comment',):
                    if r in df.columns:
                        cols.append(r)
                df = df[cols]
            
            # Try to save
            print(f"Attempting to save to: {file_path}")
            print(f"Dataframe shape: {df.shape}, columns: {list(df.columns)}")
            
            # Save the file - use explicit file writing to ensure it's saved.
            # Keep the format matching the chosen output extension: a .csv/.tsv export
            # stays a real CSV/TSV (writing xlsx bytes to a .csv name would corrupt it).
            import os
            out_ext = file_path.suffix.lower()
            is_csv_out = out_ext in ('.csv', '.tsv')
            csv_sep = '\t' if out_ext == '.tsv' else ','

            if is_csv_out and preserve_schema:
                # Re-emit in the original CSV schema (column names + order preserved).
                sch = preserve_schema
                out_df = pd.DataFrame()
                for col in sch['columns']:
                    if col == sch.get('index_col'):
                        out_df[col] = list(range(len(df)))
                    elif col == sch.get('recall_col'):
                        out_df[col] = df['recall_in_temporal_order'].values
                    elif col == sch.get('matches_col'):
                        out_df[col] = df['recalled_events'].values
                    elif col in df.columns:
                        out_df[col] = df[col].values
                    else:
                        out_df[col] = [''] * len(df)
                df = out_df

            if is_csv_out:
                temp_path = file_path.with_name(file_path.name + '.tmp')
                df.to_csv(temp_path, index=False, sep=csv_sep, encoding='utf-8', na_rep='')
            else:
                temp_path = file_path.with_suffix('.tmp.xlsx')
                df.to_excel(temp_path, index=False, engine='openpyxl', na_rep='')

            # Verify temp file was created
            if not temp_path.exists():
                raise Exception(f"Temporary file was not created: {temp_path}")

            # Replace the original file with the temp file
            if file_path.exists():
                os.remove(file_path)
            os.rename(temp_path, file_path)

            # Verify file was created and has content
            if not file_path.exists():
                raise Exception(f"File was not created after rename: {file_path}")

            # Verify the saved data by reading it back
            verify_df = pd.read_csv(file_path, sep=csv_sep) if is_csv_out else pd.read_excel(file_path)
            print(f"Verification: File exists, {len(verify_df)} rows saved")
            
            # Check if our updates are in the file
            for segment in segments:
                idx = segment.get('index')
                if idx is not None and 0 <= idx < len(verify_df):
                    saved_value = str(verify_df.iloc[idx].get('recalled_events', '')).strip()
                    expected_value = segment.get('matched_event', '').strip()
                    if expected_value:
                        # Format expected value for comparison
                        try:
                            event_parts = [str(int(float(part.strip()))) for part in expected_value.split(',') if part.strip()]
                            expected_formatted = ', '.join(event_parts)
                        except:
                            expected_formatted = expected_value
                        print(f"  Segment {idx}: Expected '{expected_formatted}', Saved '{saved_value}'")
                        if saved_value != expected_formatted:
                            print(f"  WARNING: Mismatch at index {idx}!")
                        else:
                            print(f"  ✓ Segment {idx} saved correctly")
            
        except Exception as save_error:
            error_msg = str(save_error)
            print(f"ERROR during Excel save: {save_error}")
            import traceback
            traceback.print_exc()
            
            # Check if file is locked
            if 'locked' in error_msg.lower() or 'disturbed' in error_msg.lower() or 'body' in error_msg.lower():
                return jsonify({'error': 'Excel file is locked or in use. Please close the file in Excel or any other program and try again.'}), 500
            
            # Try saving as CSV as fallback
            try:
                csv_path = file_path.with_suffix('.csv')
                df.to_csv(csv_path, index=False, encoding='utf-8', na_rep='')
                print(f"Saved as CSV fallback: {csv_path}")
                return jsonify({'error': f'Failed to save as Excel (file may be locked), saved as CSV instead: {csv_path.name}'}), 500
            except Exception as csv_error:
                print(f"CSV fallback also failed: {csv_error}")
                return jsonify({'error': f'Failed to save: {error_msg}'}), 500
        
        print(f"SUCCESS: Saved {len(df)} segments with matched events to {file_path}")
        log_user_edit(session.get('username', 'unknown'), 'save-rated', subj_id, file_path.name)
        
        # Return updated data so frontend can update UI without reloading
        updated_segments = []
        for idx, row in df.iterrows():
            parsed_text = str(row.get('recall_in_temporal_order', '')).strip()
            matched = str(row.get('recalled_events', '')).strip() if pd.notna(row.get('recalled_events')) else ''
            updated_segments.append({
                'text': parsed_text,
                'matched_event': matched
            })
        
        return jsonify({
            'success': True, 
            'message': 'Rated events saved successfully',
            'updated_data': {
                'rated_texts': updated_segments
            }
        })
    except Exception as e:
        print(f"Error saving rated events: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/subject/<subj_id>/save-parsed', methods=['POST'])
def save_parsed_segments(subj_id):
    """Save edited parsed segments."""
    try:
        data = request.get_json()
        segments = data.get('segments', [])
        
        if not segments:
            return jsonify({'error': 'No segments provided'}), 400
        
        # Use custom output path if provided and valid
        output_path_str = data.get('output_path', '')
        file_path = _resolve_and_validate_output_path(output_path_str)
        if file_path is None:
            parsed_dir = get_output_dir_for_step_type('textParsing') or RECALL_PARSED_DIR
            parsed_dir.mkdir(parents=True, exist_ok=True)
            edit_suffix = get_edit_suffix()
            file_path = parsed_dir / f"{subj_id}_parsed{edit_suffix}.xlsx"
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Read existing file if it exists, otherwise create new structure
        if file_path.exists():
            from helpers.flexible_io import read_parsed_recall_file
            read_parsed_recall_file(file_path)
        else:
            # Create new dataframe with required columns
            df = pd.DataFrame(columns=['recalled_events', 'recall_in_temporal_order'])
        
        # Sort segments by index to maintain order
        segments_sorted = sorted(segments, key=lambda x: x.get('index', 0))
        
        # Build NEW dataframe from the edited segments
        # This handles merged/split/edited segments properly
        # When the segmenter is in 1-5 mode each segment carries a
        # boundary_strength (strength of the boundary that ends it); write it to
        # a third column only when present.
        any_strength = any('boundary_strength' in seg for seg in segments_sorted)
        new_segments = []
        for seg in segments_sorted:
            text = seg.get('text', '').strip()
            if text:  # Only include non-empty segments
                row = {
                    'recall_in_temporal_order': text,
                    'recalled_events': ''  # Clear old matches - will be re-rated
                }
                if any_strength:
                    bs = seg.get('boundary_strength', '')
                    row['boundary_strength'] = '' if bs is None else str(bs)
                new_segments.append(row)
        
        # Create new dataframe with the updated segments
        new_df = pd.DataFrame(new_segments)
        
        # Ensure recalled_events column exists
        if 'recalled_events' not in new_df.columns:
            new_df['recalled_events'] = ''
        
        # Reset index to ensure sequential numbering (0, 1, 2, ...)
        new_df = new_df.reset_index(drop=True)
        
        # Ensure correct column order to match pipeline output format
        if 'recalled_events' in new_df.columns and 'recall_in_temporal_order' in new_df.columns:
            cols = ['recalled_events', 'recall_in_temporal_order']
            if any_strength and 'boundary_strength' in new_df.columns:
                cols.append('boundary_strength')
            new_df = new_df[cols]
        
        # Clean data before saving - ensure all columns are strings and handle NaN
        # Convert all columns to string type to avoid type issues
        for col in new_df.columns:
            if col == 'recall_in_temporal_order':
                new_df[col] = new_df[col].fillna('').astype(str)
            elif col == 'recalled_events':
                new_df[col] = new_df[col].fillna('').astype(str)
                new_df[col] = new_df[col].replace('nan', '')
            else:
                new_df[col] = new_df[col].fillna('').astype(str)
        
        # Replace any remaining NaN values
        new_df = new_df.fillna('')
        
        # Save updated dataframe
        try:
            # Close any existing file handles first
            import gc
            gc.collect()
            
            # Try to save
            new_df.to_excel(file_path, index=False, engine='openpyxl', na_rep='')
        except Exception as save_error:
            error_msg = str(save_error)
            print(f"Error during Excel save: {save_error}")
            
            # Check if file is locked
            if 'locked' in error_msg.lower() or 'disturbed' in error_msg.lower() or 'body' in error_msg.lower():
                return jsonify({'error': 'Excel file is locked or in use. Please close the file in Excel or any other program and try again.'}), 500
            
            # Try saving as CSV as fallback
            try:
                csv_path = file_path.with_suffix('.csv')
                new_df.to_csv(csv_path, index=False, encoding='utf-8', na_rep='')
                return jsonify({'error': f'Failed to save as Excel (file may be locked), saved as CSV instead: {csv_path.name}'}), 500
            except Exception as csv_error:
                return jsonify({'error': f'Failed to save: {error_msg}'}), 500
        
        print(f"Saved {len(new_df)} segments to {file_path}")
        log_user_edit(session.get('username', 'unknown'), 'save-parsed', subj_id, file_path.name)
        print(f"Note: Segment count changed from {len(df) if 'df' in locals() else 0} to {len(new_df)}")
        
        # FORWARD CASCADE: Tab 2 → Tab 3 (does NOT update Tab 1)
        try:
            # boundary_strength is a segmentation-only column; keep it out of the rated file.
            rated_df = _auto_rate_parsed_df(subj_id, new_df.drop(columns=['boundary_strength'], errors='ignore'))
            if rated_df is not None:
                rated_dir = get_output_dir_for_step_type('textMatching') or RECALL_RATED_DIR
                rated_dir.mkdir(parents=True, exist_ok=True)
                rated_file = rated_dir / f"{subj_id}_rate-recall.xlsx"
                rated_df.to_excel(rated_file, index=False, engine='openpyxl', na_rep='')
                matched_count = sum(1 for v in rated_df['recalled_events'] if str(v).strip())
                print(f"Auto-updated rated file: {rated_file} with {len(rated_df)} segments ({matched_count} matched)")
            else:
                print(f"Warning: Story events file not found for {subj_id}, skipping auto-rating")
        except Exception as rate_error:
            print(f"Warning: Could not auto-update rated file: {rate_error}")
            import traceback
            traceback.print_exc()
            # Don't fail the save if rating fails
        
        # Get updated data to return to client
        updated_parsed = get_parsed_texts(subj_id)
        updated_rated = get_rated_texts(subj_id)
        
        return jsonify({
            'success': True, 
            'message': f'Parsed segments saved successfully ({len(new_df)} segments). Rated file updated automatically.',
            'segment_count': len(new_df),
            'updated_data': {
                'parsed_texts': updated_parsed,
                'rated_texts': updated_rated
            }
        })
    except Exception as e:
        print(f"Error saving parsed segments: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/subject/<subj_id>/save-story-events', methods=['POST'])
def save_story_events(subj_id):
    """Save edited story events segmentation."""
    try:
        data = request.json
        events = data.get('events', [])
        
        if not events:
            return jsonify({'error': 'No events provided'}), 400
        
        # Use custom output path if provided and valid
        output_path_str = data.get('output_path', '')
        file_path = _resolve_and_validate_output_path(output_path_str)
        if file_path is None:
            events_dir = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
            events_dir.mkdir(parents=True, exist_ok=True)
            edit_suffix = get_edit_suffix()
            source_file = data.get('source_file', '')
            if source_file and source_file.endswith('.xlsx'):
                base_name = Path(source_file).stem
                file_path = events_dir / f"{base_name}{edit_suffix}.xlsx"
            else:
                file_path = events_dir / f"{subj_id}_events{edit_suffix}.xlsx"
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create dataframe from events. In 1-5 mode each event carries a
        # boundary_strength (strength of the boundary that ends it); it is added
        # as a third column after the original columns are preserved below.
        any_strength = any('boundary_strength' in event for event in events)
        events_data = []
        for event in events:
            events_data.append({
                'event': event.get('event', ''),
                'story_texts': event.get('text', '')
            })

        new_df = pd.DataFrame(events_data)
        strength_values = None
        if any_strength:
            strength_values = []
            for event in events:
                bs = event.get('boundary_strength', '')
                strength_values.append('' if bs is None else str(bs))
        
        # Try to preserve other columns from original file if it exists
        events_dir = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
        original_file = events_dir / f"{subj_id}_events.xlsx"
        if not original_file.exists():
            candidates = sorted(events_dir.glob(f"{subj_id}_events-*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
            candidates = [c for c in candidates if not is_user_edit_file(c.name)]
            if candidates:
                original_file = candidates[0]
        if original_file.exists():
            try:
                original_df = pd.read_excel(original_file)
                # Merge with original to preserve other columns
                for col in original_df.columns:
                    if col not in new_df.columns:
                        new_df[col] = None
                # Reorder columns to match original
                new_df = new_df.reindex(columns=original_df.columns, fill_value='')
            except Exception as e:
                print(f"Warning: Could not read original file to preserve columns: {e}")

        # Append the segmentation strength column last (after original-column
        # preservation, which would otherwise drop a new column via reindex).
        if strength_values is not None:
            new_df['boundary_strength'] = strength_values

        # Save to file
        try:
            new_df.to_excel(file_path, index=False, engine='openpyxl')
        except Exception as e:
            error_msg = str(e)
            print(f"Error saving Excel file: {error_msg}")
            # Try CSV as fallback
            try:
                csv_path = file_path.with_suffix('.csv')
                new_df.to_csv(csv_path, index=False, encoding='utf-8', na_rep='')
                return jsonify({'error': f'Failed to save as Excel (file may be locked), saved as CSV instead: {csv_path.name}'}), 500
            except Exception as csv_error:
                return jsonify({'error': f'Failed to save: {error_msg}'}), 500
        
        print(f"Saved {len(new_df)} story events to {file_path}")
        log_user_edit(session.get('username', 'unknown'), 'save-story-events', subj_id, file_path.name)
        return jsonify({'success': True, 'message': f'Saved {len(new_df)} events', 'events': events})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/story/<story_name>/save-causal-matrix', methods=['POST'])
@app.route('/api/subject/<story_name>/save-causal-matrix', methods=['POST'])
def save_causal_matrix(story_name):
    """Save causal rating matrix data from the manual entry UI."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        pairs = data.get('pairs', [])
        flags = data.get('flags', {})

        # Use custom output path if provided and valid (user's path takes precedence)
        output_path_str = str(data.get('output_path') or '').strip()
        file_path = None
        if output_path_str:
            file_path = _resolve_and_validate_output_path(output_path_str)
            if file_path:
                print(f"save_causal_matrix: using user output_path={output_path_str!r} -> {file_path}")
            # If path is filename-only (no slashes), resolve relative to causal output dir
            if file_path is None and '/' not in output_path_str and '\\' not in output_path_str:
                causal_dir = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
                try:
                    rel = _path_for_client(causal_dir)
                except ValueError:
                    rel = 'output/causal_rated'
                file_path = _resolve_and_validate_output_path(f"{rel}/{output_path_str}".replace('//', '/'))
            # Fallback: use user's filename in default causal dir (preserves custom names like ...edit999.xlsx)
            if file_path is None and output_path_str:
                causal_dir = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
                filename = Path(output_path_str).name
                if filename and (filename.endswith('.xlsx') or filename.endswith('.csv')):
                    candidate = causal_dir / filename
                    # Accept candidates that resolve under PROJECT_ROOT, OR under the
                    # user-configured causal output dir (which may itself be absolute
                    # and outside PROJECT_ROOT — e.g. ~/Desktop/runs/...).
                    causal_dir_resolved = causal_dir.resolve()
                    try:
                        candidate.resolve().relative_to(PROJECT_ROOT)
                        file_path = candidate
                    except ValueError:
                        try:
                            candidate.resolve().relative_to(causal_dir_resolved)
                            file_path = candidate
                        except ValueError:
                            pass
            if file_path is None and output_path_str:
                print(f"save_causal_matrix: REJECTED output_path={output_path_str!r}")
                return jsonify({'error': f'Invalid output path (must be under project root): {output_path_str[:80]}'}), 400
        else:
            print(f"save_causal_matrix: no output_path in request, using default")
        if file_path is None:
            causal_dir = get_output_dir_for_step_type('causalRating') or CAUSAL_RATED_DIR
            causal_dir.mkdir(parents=True, exist_ok=True)
            edit_suffix = get_edit_suffix()
            source_file = data.get('source_file', '')
            if source_file and source_file.endswith('.xlsx'):
                base_name = Path(source_file).stem
                if is_user_edit_file(source_file):
                    file_path = causal_dir / source_file
                else:
                    file_path = causal_dir / f"{base_name}{edit_suffix}.xlsx"
            else:
                file_path = causal_dir / f"{story_name}_causal-manual{edit_suffix}.xlsx"
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)

        def _pair_rows(pair_list):
            out = []
            for pair in pair_list:
                rating = int(pair.get('rating', 0))
                if rating > 0:
                    row = {
                        'event_A_number': int(pair['event_A']),
                        'event_B_number': int(pair['event_B']),
                        'rating': rating,
                        'reasoning': pair.get('reasoning', '') or 'manual entry',
                    }
                    if flags.get('causal_type'):
                        row['causal_type'] = pair.get('causal_type', '')
                    if flags.get('referential'):
                        row['referential'] = int(pair.get('referential', 0))
                    if flags.get('contextual'):
                        row['contextual'] = int(pair.get('contextual', 0))
                    if flags.get('semantic'):
                        row['semantic'] = int(pair.get('semantic', 0))
                    out.append(row)
            return out

        _coarse_in_req = data.get('coarse_pairs')
        is_nested = (
            bool(data.get('is_nested'))
            or ('nested_combined_rows' in data)
            or (isinstance(_coarse_in_req, list) and len(_coarse_in_req) > 0)
        )
        if is_nested:
            nested_combined = data.get('nested_combined_rows') or []
            if not isinstance(nested_combined, list):
                nested_combined = []
            if not nested_combined:
                coarse_in = data.get('coarse_pairs') or []
                nested_combined = _build_nested_combined_from_parts(
                    story_name, pairs, coarse_in, flags
                )
                print(f"Built nested combined rows server-side ({len(nested_combined)} rows) for {story_name}")
            export_cols = nested_combined_export_columns(flags)
            norm = []
            for r in nested_combined:
                norm.append({c: r.get(c) for c in export_cols})
            try:
                _save_nested_combined_xlsx(file_path, norm, export_cols)
            except Exception as e:
                csv_path = file_path.with_suffix('.csv')
                pd.DataFrame(norm).reindex(columns=export_cols).to_csv(csv_path, index=False, encoding='utf-8', na_rep='')
                return jsonify({'error': f'Failed to save as Excel, saved as CSV: {csv_path.name}'}), 500
            n_nonzero = sum(
                1 for r in norm
                if int(float(r.get('fine_rating') or 0)) > 0
            )
            try:
                rel_path = _path_for_client(file_path)
            except ValueError:
                rel_path = str(file_path)
            print(f"Saved nested combined causal sheet to {rel_path} ({len(norm)} rows, {n_nonzero} fine pairs)")
            log_user_edit(session.get('username', 'unknown'), 'save-causal-matrix', story_name, file_path.name)
            return jsonify({
                'success': True,
                'message': f'Saved {n_nonzero} fine causal pairs (nested combined format)',
                'file': file_path.name,
                'path_used': rel_path,
            })

        rows = _pair_rows(pairs)
        coarse_pairs_in = data.get('coarse_pairs') or []
        coarse_rows = _pair_rows(coarse_pairs_in)
        event_hierarchy = data.get('event_hierarchy') or []

        columns = ['event_A_number', 'event_B_number', 'rating', 'reasoning']
        if flags.get('causal_type'):
            columns.append('causal_type')
        if flags.get('referential'):
            columns.append('referential')
        if flags.get('contextual'):
            columns.append('contextual')
        if flags.get('semantic'):
            columns.append('semantic')

        if rows:
            df = pd.DataFrame(rows)
            df = df[[c for c in columns if c in df.columns]]
        else:
            df = pd.DataFrame(columns=columns)

        try:
            if coarse_rows or event_hierarchy:
                h_columns = ['coarse_event_number', 'coarse_event_text', 'fine_event_number', 'fine_event_text']
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='')
                    if coarse_rows:
                        cdf = pd.DataFrame(coarse_rows)
                        cdf = cdf[[c for c in columns if c in cdf.columns]]
                        cdf.to_excel(writer, sheet_name='coarse_causal_pairs', index=False, na_rep='')
                    if event_hierarchy:
                        hdf = pd.DataFrame(event_hierarchy)
                        for c in h_columns:
                            if c not in hdf.columns:
                                hdf[c] = ''
                        hdf = hdf[h_columns]
                        hdf.to_excel(writer, sheet_name='event_hierarchy', index=False, na_rep='')
                        ws = writer.sheets['event_hierarchy']
                        _merge_event_hierarchy_coarse_cells(ws, len(hdf))
            else:
                df.to_excel(file_path, index=False, engine='openpyxl', na_rep='')
        except Exception as e:
            csv_path = file_path.with_suffix('.csv')
            df.to_csv(csv_path, index=False, encoding='utf-8', na_rep='')
            return jsonify({'error': f'Failed to save as Excel, saved as CSV: {csv_path.name}'}), 500

        n_nonzero = len(rows)
        try:
            rel_path = _path_for_client(file_path)
        except ValueError:
            rel_path = str(file_path)
        print(f"Saved {n_nonzero} causal pairs to {rel_path} (requested: {repr(output_path_str) if output_path_str else 'default'})")
        log_user_edit(session.get('username', 'unknown'), 'save-causal-matrix', story_name, file_path.name)
        return jsonify({'success': True, 'message': f'Saved {n_nonzero} causal pairs', 'file': file_path.name, 'path_used': rel_path})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/subject/<subj_id>/audio')
def serve_subject_audio(subj_id):
    """Serve the recall audio file for a subject."""
    audio_file = get_audio_file(subj_id, is_story=False)
    if not audio_file:
        return "Audio file not found", 404
    path = PROJECT_ROOT / audio_file
    if not path.exists() or not path.is_file():
        return "Audio file not found", 404
    mimetype = _get_audio_mimetype(path)
    return send_file(str(path), mimetype=mimetype, as_attachment=False)


def _get_audio_mimetype(path):
    """Return MIME type for audio file based on extension."""
    ext = Path(path).suffix.lower()
    mime_map = {
        '.mp3': 'audio/mpeg',
        '.wav': 'audio/wav',
        '.mp4': 'audio/mp4',
        '.m4a': 'audio/mp4',
        '.flac': 'audio/flac',
        '.ogg': 'audio/ogg',
        '.webm': 'audio/webm',
        '.aac': 'audio/aac',
    }
    return mime_map.get(ext, 'application/octet-stream')


@app.route('/api/story/<story_name>/audio')
def serve_story_audio(story_name):
    """Serve the story audio file."""
    audio_file = get_audio_file(story_name, is_story=True)
    if not audio_file:
        return "Audio file not found", 404
    path = PROJECT_ROOT / audio_file
    if not path.exists() or not path.is_file():
        return "Audio file not found", 404
    mimetype = _get_audio_mimetype(path)
    return send_file(str(path), mimetype=mimetype, as_attachment=False)


@app.route('/api/subject/<subj_id>/save-audio-transcription', methods=['POST'])
def save_audio_transcription(subj_id):
    """Save edited audio transcription."""
    try:
        data = request.json
        transcription = data.get('transcription', '')
        
        # Use custom output path if provided and valid
        output_path_str = data.get('output_path', '')
        file_path = _resolve_and_validate_output_path(output_path_str)
        if file_path is None:
            transcribed_dir = get_output_dir_for_step_type('audioTranscribe:recall') or RECALL_AUDIO_TRANSCRIBED_DIR
            transcribed_dir.mkdir(parents=True, exist_ok=True)
            audio_file = get_audio_file(subj_id)
            if not audio_file:
                return jsonify({'error': 'Audio file not found'}), 404
            audio_path = PROJECT_ROOT / audio_file
            edit_suffix = get_edit_suffix()
            transcription_filename = audio_path.stem + edit_suffix + '.txt'
            file_path = transcribed_dir / transcription_filename
        else:
            file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save transcription
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(transcription)
        
        print(f"Saved audio transcription to {file_path}")
        log_user_edit(session.get('username', 'unknown'), 'save-audio-transcription', subj_id, file_path.name)
        return jsonify({'success': True, 'message': 'Transcription saved', 'transcription': transcription})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _safe_browse_dir(path_str: str, *, walk_up_to_dir: bool = False) -> Path | None:
    """Resolve a path string to a real directory for the folder picker.

    Accepts project-relative paths, POSIX/Windows absolute paths, and tilde
    expansion. Returns ``None`` if the path cannot be resolved to an
    existing directory.

    The picker is intentionally permissive about paths outside the project
    root because, after ``pip install``, ``PROJECT_ROOT`` may be inside
    ``site-packages`` and the user's actual data lives elsewhere on the
    filesystem.

    When macOS privacy (TCC) blocks ``stat``/``resolve`` on Desktop, Dropbox,
    etc., an absolute path from drag-and-drop is still accepted so the user
    can select it even when subfolders cannot be listed.

    When ``walk_up_to_dir`` is true, if the resolved path points at a file
    (or doesn't exist), walk up its parent chain until a real directory is
    found. This makes drag-and-drop forgiving: if the user drops a file
    from inside a folder, we use the folder.
    """
    try:
        root = WORKSPACE_ROOT.resolve()
    except OSError:
        root = WORKSPACE_ROOT
    raw = (path_str or '').strip()
    if not raw:
        return root if _path_is_dir_or_unstatable(root) else root
    p_obj = Path(raw).expanduser()
    if p_obj.is_absolute():
        target = _resolve_path_lenient(p_obj, root=root)
    else:
        rel = raw.replace('\\', '/').strip('/')
        target = _resolve_path_lenient(root / rel if rel else root, root=root)

    if _path_is_dir_or_unstatable(target):
        return target
    if not walk_up_to_dir:
        if p_obj.is_absolute() and _looks_like_user_absolute_dir(p_obj):
            return p_obj
        return None
    # Walk up parent chain looking for the nearest existing directory.
    cur = target
    for _ in range(64):  # paranoia cap; filesystem trees are not infinite
        parent = cur.parent
        if parent == cur:
            break
        if _path_is_dir_or_unstatable(parent):
            return parent
        cur = parent
    if p_obj.is_absolute() and _looks_like_user_absolute_dir(p_obj):
        return p_obj
    return None


# Backwards-compatible alias for older callers.
_safe_project_dir = _safe_browse_dir


def _resolve_desktop_path():
    """Best-effort Desktop folder for the folder picker shortcuts."""
    try:
        home = Path.home()
    except RuntimeError:
        return None
    candidates = [
        home / 'Desktop',
        home / 'OneDrive' / 'Desktop',
        home / 'Library' / 'Mobile Documents' / 'com~apple~CloudDocs' / 'Desktop',
    ]
    for c in candidates:
        try:
            if c.is_dir():
                return c
        except OSError:
            continue
    return home / 'Desktop'


def _list_picker_subdirs(directory):
    """List immediate subdirectories for the folder picker.

    Returns ``(folder_names, warning_message)``. ``warning_message`` is set when
    the directory exists but cannot be read (common on macOS Desktop without
    Full Disk Access).
    """
    if directory is None:
        return [], None
    try:
        target = Path(directory)
    except (TypeError, ValueError):
        return [], 'Invalid folder path'
    try:
        if not target.is_dir():
            if _looks_like_user_absolute_dir(target):
                # TCC may block stat while the user still drag-dropped a valid folder.
                pass
            else:
                return [], 'Not a folder'
    except PermissionError:
        pass
    except OSError as e:
        return [], str(e)
    folders = []
    try:
        with os.scandir(target) as it:
            for entry in it:
                if entry.name.startswith('.'):
                    continue
                try:
                    if entry.is_dir(follow_symlinks=False):
                        folders.append(entry.name)
                except OSError:
                    continue
    except PermissionError:
        return [], (
            'Permission denied reading this folder. On macOS, grant Full Disk Access '
            'to Terminal (or Python) in System Settings → Privacy & Security, or paste '
            'the folder path into the text box below.'
        )
    except (FileNotFoundError, NotADirectoryError) as e:
        return [], str(e)
    except OSError as e:
        return [], str(e)
    return sorted(folders), None


def _relative_project_path(directory: Path) -> str:
    """Render a folder path for the picker UI.

    Returns a project-relative POSIX path when ``directory`` is under
    ``PROJECT_ROOT`` (so the common case stays clean); otherwise returns the
    absolute path so external folders survive a round trip through the UI.
    """
    root = WORKSPACE_ROOT.resolve()
    try:
        directory = directory.resolve()
    except OSError:
        return str(directory)
    if directory == root:
        return ''
    try:
        return directory.relative_to(root).as_posix()
    except ValueError:
        return directory.as_posix()


@app.route('/api/browse-folders', methods=['GET'])
def browse_folders():
    """List subdirectories for the pipeline folder picker.

    Accepts a ``path`` query parameter that may be project-relative, an
    absolute POSIX/Windows path, or use ``~`` for the user's home directory.
    """
    try:
        rel = request.args.get('path', '')
        # Drag-and-drop senders set ?derive_dir=1 so a dropped file (or a
        # slightly-off path) auto-resolves to its closest existing parent.
        derive = request.args.get('derive_dir', '').lower() in ('1', 'true', 'yes')
        target = _safe_browse_dir(rel, walk_up_to_dir=derive)
        if target is None:
            return jsonify({'success': False, 'error': 'Invalid or missing folder path'}), 400

        root = WORKSPACE_ROOT.resolve()
        folders, list_warning = _list_picker_subdirs(target)
        folder_entries = []
        for name in folders:
            child = target / name
            folder_entries.append({
                'name': name,
                'path': _relative_project_path(child),
                'absolute': str(child),
            })

        parent = None
        target_parent = target.parent
        if target_parent != target:
            parent = _relative_project_path(target_parent)

        try:
            home_path = str(Path.home())
        except RuntimeError:
            home_path = ''

        desktop_path = _resolve_desktop_path()
        try:
            desktop_str = str(desktop_path.resolve()) if desktop_path else ''
        except OSError:
            desktop_str = str(desktop_path) if desktop_path else ''

        return jsonify({
            'success': True,
            'current': _relative_project_path(target),
            'currentAbsolute': str(target),
            'parent': parent,
            'folders': folders,
            'folderEntries': folder_entries,
            'warning': list_warning,
            'projectRoot': str(root),
            'home': home_path,
            'desktop': desktop_str,
            'pythonExecutable': sys.executable,
            'fdaHint': (
                'If folders stay empty after granting Full Disk Access to Terminal, also add '
                f'this Python binary in System Settings → Privacy & Security → Full Disk Access: {sys.executable}'
            ),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/find-folder-by-name', methods=['GET'])
def find_folder_by_name():
    """Locate a folder by name in common user/data locations.

    When the browser hides the absolute path of a drag-and-dropped folder
    (Safari and some Chromium configurations strip ``file://`` URIs out of
    ``DataTransfer``), the client falls back to this endpoint with just the
    folder name. We probe the usual places a researcher would keep input
    data — home, Desktop, Documents, Downloads, common cloud-sync dirs, the
    project root and its siblings, plus the direct subdirectories of home
    — and return every match.

    Returns:
        {
            "success": true,
            "name": "<folder name>",
            "matches": [ {"path": "/abs/path", "label": "Desktop / remote3"} , ... ]
        }
    """
    try:
        raw_name = (request.args.get('name') or '').strip()
        # The optional ``hints`` param lets the client narrow the candidate
        # set when it has read directory entries via webkitGetAsEntry — a
        # comma-separated list of immediate-child entry names. We pick the
        # match with the most overlap when several share the name.
        raw_hints = (request.args.get('hints') or '').strip()
        if not raw_name:
            return jsonify({'success': False, 'error': 'Missing folder name'}), 400
        # Reject anything that smells like a path so the endpoint can't be
        # tricked into traversing the filesystem from a non-name input.
        bad_chars = ('/', '\\', '..', '\0', '\n', '\r')
        if any(ch in raw_name for ch in bad_chars):
            return jsonify({'success': False, 'error': 'Invalid folder name'}), 400
        if raw_name in ('.', '..'):
            return jsonify({'success': False, 'error': 'Invalid folder name'}), 400

        try:
            home = Path.home()
        except RuntimeError:
            home = None

        matches = []
        seen = set()

        def consider(parent, source_label):
            """If ``parent / raw_name`` is an existing directory, record it."""
            try:
                if not parent or not parent.is_dir():
                    return
            except OSError:
                return
            candidate = parent / raw_name
            try:
                if not candidate.is_dir():
                    return
                full = candidate.resolve()
            except OSError:
                return
            key = str(full)
            if key in seen:
                return
            seen.add(key)
            try:
                rel_to_home = full.relative_to(home).as_posix() if home else str(full)
            except (ValueError, AttributeError):
                rel_to_home = str(full)
            matches.append({
                'path': str(full),
                'label': source_label,
                'displayPath': str(full),
                'shortPath': rel_to_home if home and str(full).startswith(str(home)) else str(full),
            })

        # 1) Well-known user directories.
        if home:
            for sub, label in [
                ('', 'Home'),
                ('Desktop', 'Desktop'),
                ('Documents', 'Documents'),
                ('Downloads', 'Downloads'),
                ('Movies', 'Movies'),
                ('Music', 'Music'),
                ('Pictures', 'Pictures'),
                ('Public', 'Public'),
                ('Dropbox', 'Dropbox'),
                ('OneDrive', 'OneDrive'),
                ('Google Drive', 'Google Drive'),
                ('iCloud Drive', 'iCloud Drive'),
                ('Library/Mobile Documents/com~apple~CloudDocs', 'iCloud Drive'),
            ]:
                consider((home / sub) if sub else home, label)

        # 2) Project root + immediate sibling (handy for cloned-repo installs
        #    where the user's data folder lives next to the project).
        #    Skip the sibling scan when we're running out of site-packages
        #    (after ``pip install``), because that "parent" is full of other
        #    Python packages and would generate noisy false matches.
        consider(WORKSPACE_ROOT, 'project root')
        project_root_str = str(WORKSPACE_ROOT)
        if WORKSPACE_ROOT != PACKAGE_ROOT:
            consider(PACKAGE_ROOT, 'package root')
        if 'site-packages' not in project_root_str and 'dist-packages' not in project_root_str:
            consider(PROJECT_ROOT.parent, 'project parent')

        # 3) Direct subdirectories of home (catches things like
        #    ~/research_data/remote3 without an expensive deep scan).
        SKIP_AT_HOME = {
            'Library', 'Applications', 'opt', 'anaconda3', 'miniconda3',
            'venv', '.venv', 'node_modules', '__pycache__', '.cache',
        }
        if home and home.is_dir():
            count = 0
            try:
                with os.scandir(home) as it:
                    for entry in it:
                        if count >= 250:
                            break
                        if entry.name.startswith('.') or entry.name in SKIP_AT_HOME:
                            continue
                        try:
                            if not entry.is_dir(follow_symlinks=False):
                                continue
                        except OSError:
                            continue
                        count += 1
                        consider(Path(entry.path), entry.name)
            except (PermissionError, OSError):
                pass

        # 4) One level deeper inside Desktop / Documents / Downloads and the
        #    most common cloud-sync dirs — research data often lives in
        #    e.g. ~/Desktop/2026_data/remote3 or ~/Dropbox/research/remote3.
        DEEPER_BASES = []
        if home:
            for sub in ('Desktop', 'Documents', 'Downloads',
                        'Dropbox', 'OneDrive', 'Google Drive',
                        'Library/Mobile Documents/com~apple~CloudDocs'):
                DEEPER_BASES.append(home / sub)
        for base in DEEPER_BASES:
            try:
                if not base.is_dir():
                    continue
            except OSError:
                continue
            count = 0
            try:
                with os.scandir(base) as it:
                    for entry in it:
                        if count >= 200:
                            break
                        if entry.name.startswith('.'):
                            continue
                        try:
                            if not entry.is_dir(follow_symlinks=False):
                                continue
                        except OSError:
                            continue
                        count += 1
                        consider(Path(entry.path), f"{base.name}/{entry.name}")
            except (PermissionError, OSError):
                pass

        # 5) Re-rank by hint overlap if the client supplied directory entry
        #    names from webkitGetAsEntry.
        if raw_hints and matches:
            hints = [h.strip() for h in raw_hints.split(',') if h.strip()][:50]
            hints_set = set(hints)
            if hints_set:
                def overlap(match):
                    try:
                        with os.scandir(match['path']) as it:
                            names = {e.name for e in it}
                        return len(hints_set & names)
                    except (PermissionError, OSError):
                        return 0
                matches.sort(key=overlap, reverse=True)

        return jsonify({
            'success': True,
            'name': raw_name,
            'matches': matches,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pipeline/save', methods=['POST'])
def save_pipeline():
    """Save pipeline configuration."""
    try:
        data = request.json
        steps = data.get('steps', [])
        
        if not steps:
            return jsonify({'success': False, 'error': 'No steps provided'}), 400
        
        normalized_steps = [normalize_pipeline_step(s) for s in steps]
        pipeline_config = {
            'steps': normalized_steps,
            'created_at': datetime.now().isoformat(),
        }
        
        pipeline_file = _pipeline_config_path()
        ACCOUNT_DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(pipeline_file, 'w', encoding='utf-8') as f:
            json.dump(pipeline_config, f, indent=2)
        
        print(f"Pipeline configuration saved with {len(steps)} steps")
        return jsonify({'success': True, 'message': f'Pipeline saved with {len(steps)} steps'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/event-segment-options', methods=['GET'])
def get_event_segment_options():
    """Return available models and prompt versions for story event segmentation."""
    try:
        import importlib.util
        seg_file = SCRIPTS_DIR / '2_story-event-segment.py'
        spec = importlib.util.spec_from_file_location("event_seg", seg_file)
        seg_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(seg_module)

        models = []
        for mid, info in seg_module.SUPPORTED_MODELS.items():
            provider = info['provider']
            if provider == 'openai':
                key_env = 'OPENAI_API_KEY'
                has_key = bool(os.getenv(key_env))
            elif provider == 'ollama':
                key_env = 'OLLAMA_HOST'
                has_key = True
            else:
                key_env = 'ANTHROPIC_API_KEY'
                has_key = bool(os.getenv(key_env))
            models.append({
                'id': mid,
                'label': info['label'],
                'provider': provider,
                'has_api_key': has_key,
                'key_env_var': key_env,
            })

        prompts = seg_module.list_event_segment_prompts()
        prompt_list = []
        for p in prompts:
            prompt_list.append({
                'filename': p,
                'label': p.replace('.txt', '').replace('_', ' ').title(),
            })

        return jsonify({
            'success': True,
            'models': models,
            'prompts': prompt_list,
            'default_model': seg_module.DEFAULT_MODEL,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/causal-rating-options', methods=['GET'])
def get_causal_rating_options():
    """Return available models and prompt versions for causal rating."""
    try:
        import importlib.util
        causal_file = SCRIPTS_DIR / '6_causal-rater.py'
        spec = importlib.util.spec_from_file_location("causal_rater", causal_file)
        causal_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(causal_module)

        models = []
        for mid, info in causal_module.SUPPORTED_MODELS.items():
            provider = info['provider']
            key_env = 'OPENAI_API_KEY' if provider == 'openai' else 'ANTHROPIC_API_KEY'
            has_key = bool(os.getenv(key_env))
            models.append({
                'id': mid,
                'label': info['label'],
                'provider': provider,
                'has_api_key': has_key,
                'key_env_var': key_env,
            })

        prompts = causal_module.list_causal_rating_prompts()
        prompt_list = []
        for p in prompts:
            prompt_list.append({
                'filename': p,
                'label': p.replace('.txt', '').replace('_', ' ').title(),
            })

        return jsonify({
            'success': True,
            'models': models,
            'prompts': prompt_list,
            'default_model': causal_module.DEFAULT_MODEL,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/recall-rating-options', methods=['GET'])
def get_recall_rating_options():
    """Return available models and prompt versions for recall-to-event matching (step 5)."""
    try:
        import importlib.util
        rate_file = SCRIPTS_DIR / '5_recall-rater.py'
        spec = importlib.util.spec_from_file_location("recall_rater_opts", rate_file)
        rate_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(rate_module)

        models = []
        for mid, info in rate_module.SUPPORTED_MODELS.items():
            provider = info['provider']
            key_env = 'OPENAI_API_KEY' if provider == 'openai' else 'ANTHROPIC_API_KEY'
            has_key = bool(os.getenv(key_env))
            models.append({
                'id': mid,
                'label': info['label'],
                'provider': provider,
                'has_api_key': has_key,
                'key_env_var': key_env,
            })

        prompts = rate_module.list_recall_rating_prompts()
        prompt_list = [
            {'filename': p, 'label': p.replace('.txt', '').replace('_', ' ').title()}
            for p in prompts
        ]

        return jsonify({
            'success': True,
            'models': models,
            'prompts': prompt_list,
            'default_model': rate_module.DEFAULT_MODEL,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pipeline/load', methods=['GET'])
def load_pipeline():
    """Load pipeline configuration."""
    try:
        pipeline_file = _pipeline_config_path()
        if not pipeline_file.exists():
            return jsonify({'success': False, 'error': 'No pipeline configuration found', 'pipeline': None})
        
        with open(pipeline_file, 'r', encoding='utf-8') as f:
            pipeline_config = json.load(f)
        
        return jsonify({'success': True, 'pipeline': pipeline_config})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/process-files', methods=['POST'])
def process_files():
    """Process uploaded files for a specific step and item."""
    try:
        from werkzeug.utils import secure_filename
        
        item_id = request.form.get('item_id')
        step_type = request.form.get('step_type')
        step_index = request.form.get('step_index')
        input_path = request.form.get('input_path')
        output_path = request.form.get('output_path')
        
        if not all([item_id, step_type, input_path, output_path]):
            return jsonify({'success': False, 'error': 'Missing required parameters'}), 400
        
        # Get uploaded files
        uploaded_files = []
        for key in request.files:
            if key.startswith('file_'):
                file = request.files[key]
                if file.filename:
                    uploaded_files.append(file)
        
        if not uploaded_files:
            return jsonify({'success': False, 'error': 'No files uploaded'}), 400
        
        input_dir = _resolve_path_from_config(input_path) or (PROJECT_ROOT / input_path)
        input_dir.mkdir(parents=True, exist_ok=True)

        output_dir = _resolve_path_from_config(output_path) or (PROJECT_ROOT / output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save uploaded files to input directory
        saved_files = []
        for file in uploaded_files:
            filename = secure_filename(file.filename)
            # Optionally prefix with item_id if needed
            file_path = input_dir / filename
            file.save(str(file_path))
            saved_files.append(_path_for_client(file_path))
        
        # Process files using the appropriate script
        script_map = {
            'audioTranscribe:story': '1_audio-transcribe.py',
            'audioTranscribe:recall': '1_audio-transcribe.py',
            'eventSegment': '2_story-event-segment.py',
            'sentenceCorrect': '3_spell-grammar-correct.py',
            'textParsing': '4_parse-texts.py',
            'textMatching': '5_recall-rater.py',
            'causalRating': '6_causal-rater.py'
        }
        
        script_name = script_map.get(step_type)
        if not script_name:
            return jsonify({'success': False, 'error': f'Processing not available for step type: {step_type}'}), 400
        
        script_path = SCRIPTS_DIR / script_name
        if not script_path.exists():
            return jsonify({'success': False, 'error': f'Script not found: {script_name}'}), 404
        
        # Run the script with proper single-file processing
        import subprocess
        import sys
        import os
        
        env = os.environ.copy()
        env['BATCH_STEP_TYPE'] = step_type
        env['BATCH_ITEM_ID'] = item_id
        
        # For single file processing, call scripts with appropriate arguments
        cmd = [sys.executable, str(script_path)]
        
        if step_type == 'eventSegment':
            # For story event segmentation, process each uploaded file individually
            # Each file should be a transcript file
            processed_count = 0
            errors = []
            
            for saved_file in saved_files:
                file_path = PROJECT_ROOT / saved_file
                if not file_path.exists():
                    errors.append(f"File not found: {saved_file}")
                    continue
                
                # Use item_id as story name (from web interface) or extract from filename
                story_name = item_id if item_id else file_path.stem.replace('_story', '').replace('_transcript', '')
                
                # Set the story name in environment so script can use it
                env_with_story = env.copy()
                env_with_story['BATCH_ITEM_ID'] = story_name
                
                # Call script with --input argument for single file
                single_cmd = cmd + ['--input', str(file_path), '--output', str(output_dir)]
                result = subprocess.run(
                    single_cmd,
                    cwd=str(WORKSPACE_ROOT),
                    env=env_with_story,
                    capture_output=True,
                    text=True,
                    timeout=3600
                )
                
                if result.returncode == 0:
                    processed_count += 1
                else:
                    errors.append(f"Failed to process {saved_file}: {result.stderr}")
            
            if processed_count > 0:
                return jsonify({
                    'success': True,
                    'message': f'Processed {processed_count} file(s)',
                    'files': saved_files,
                    'errors': errors if errors else None
                })
            else:
                return jsonify({
                    'success': False,
                    'error': f'Failed to process files: {"; ".join(errors)}'
                }), 500
                
        elif step_type in ['audioTranscribe:story', 'audioTranscribe:recall']:
            # For audio transcription, set environment variables for batch processing
            # The script will process all files in the input directory
            env['BATCH_INPUT_DIR'] = str(input_dir)
            env['BATCH_OUTPUT_DIR'] = str(output_dir)
            
            result = subprocess.run(
                cmd,
                cwd=str(WORKSPACE_ROOT),
                env=env,
                capture_output=True,
                text=True,
                timeout=3600
            )
            
            if result.returncode == 0:
                return jsonify({
                    'success': True,
                    'message': f'Processed {len(saved_files)} file(s)',
                    'files': saved_files,
                    'output': result.stdout
                })
            else:
                return jsonify({
                    'success': False,
                    'error': f'Processing failed: {result.stderr}',
                    'output': result.stdout
                }), 500
        else:
            # For other step types, use batch processing with environment variables
            env['BATCH_INPUT_DIR'] = str(input_dir)
            env['BATCH_OUTPUT_DIR'] = str(output_dir)
            
            result = subprocess.run(
                cmd,
                cwd=str(WORKSPACE_ROOT),
                env=env,
                capture_output=True,
                text=True,
                timeout=3600
            )
            
            if result.returncode == 0:
                return jsonify({
                    'success': True,
                    'message': f'Processed {len(saved_files)} file(s)',
                    'files': saved_files,
                    'output': result.stdout
                })
            else:
                return jsonify({
                    'success': False,
                    'error': f'Processing failed: {result.stderr}',
                    'output': result.stdout
                }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'Processing timed out'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/status/<item_id>', methods=['GET'])
def get_item_status(item_id):
    """Get current processing status for a specific item."""
    try:
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        
        # Determine if story or subject
        is_story = False
        stories = get_all_stories(pipeline_config)
        if item_id in stories:
            is_story = True
        
        # Get status for all steps
        status = {}
        for step in pipeline_config['steps']:
            step_id = step.get('id')
            if step_id is None:
                continue
            step_key = f"step_{step_id}"
            status[step_key] = check_step_status(item_id, step, is_story=is_story)
        
        return jsonify({
            'success': True,
            'item_id': item_id,
            'is_story': is_story,
            'status': status
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Test route to verify routing works
@app.route('/api/test-input-files', methods=['GET'])
def test_input_files_route():
    """Test route to verify routing is working."""
    return jsonify({'success': True, 'message': 'Route is working', 'cwd': os.getcwd(), 'project_root': str(PROJECT_ROOT)})

@app.route('/api/item/<item_id>/step/<int:step_index>/input-files', methods=['GET'])
def get_input_files(item_id, step_index):
    """Get input files for a specific item and step."""
    print(f"DEBUG: get_input_files ROUTE HIT with item_id={item_id}, step_index={step_index}")
    print(f"DEBUG: Current working directory: {os.getcwd()}")
    print(f"DEBUG: PROJECT_ROOT: {PROJECT_ROOT}")
    print(f"DEBUG: PROJECT_ROOT exists: {PROJECT_ROOT.exists()}")
    
    # Immediately return a test response to verify route is working
    # Remove this after confirming route works
    # return jsonify({'success': True, 'test': True, 'item_id': item_id, 'step_index': step_index})
    
    try:
        print(f"DEBUG: get_input_files called with item_id={item_id}, step_index={step_index}")
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            print(f"DEBUG: No pipeline config found")
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        
        if step_index < 0 or step_index >= len(pipeline_config['steps']):
            print(f"DEBUG: Invalid step_index {step_index}, pipeline has {len(pipeline_config['steps'])} steps")
            return jsonify({'success': False, 'error': 'Invalid step index'}), 400
        
        step = pipeline_config['steps'][step_index]
        step_type = step_runtime_key(step)
        input_path = step.get('inputPath', '')
        
        print(f"DEBUG: step_type={step_type}, input_path={input_path}")
        
        if not input_path:
            print(f"DEBUG: No input path configured")
            return jsonify({'success': True, 'files': []})
        
        input_path = input_path.rstrip('/')
        input_path_obj = _resolve_path_from_config(input_path) or (PROJECT_ROOT / input_path)
        
        print(f"DEBUG: Resolved input_path_obj={input_path_obj}, exists={input_path_obj.exists()}")
        
        if not input_path_obj.exists():
            print(f"DEBUG: Input path does not exist")
            return jsonify({'success': True, 'files': []})
        
        # Check if input_path is a file or directory
        if input_path_obj.is_file():
            # If it's a file, check if it contains data for this item
            print(f"DEBUG: Input path is a file: {input_path_obj}")
            
            # For sentenceCorrect with Excel input, the file contains multiple subjects
            # So if the file exists, it's available for all items
            if step_type == 'sentenceCorrect' and input_path_obj.suffix.lower() in ['.xlsx', '.xls']:
                print(f"DEBUG: Excel file input for sentenceCorrect, returning file for all items")
                return jsonify({
                    'success': True,
                    'files': [{
                        'name': input_path_obj.name,
                        'path': _path_for_client(input_path_obj),
                        'size': input_path_obj.stat().st_size
                    }]
                })
            
            if input_path_obj.suffix.lower() in ['.xlsx', '.xls']:
                # For Excel files, check if they contain data for this item
                try:
                    import pandas as pd
                    df = pd.read_excel(input_path_obj, sheet_name=None)
                    for sheet_name, sheet_df in df.items():
                        # Check for various column name variations
                        sub_col = None
                        id_col = None
                        for col in sheet_df.columns:
                            col_lower = str(col).lower().strip()
                            if col_lower in ['sub', 'subject', 'subj']:
                                sub_col = col
                            if col_lower in ['id', 'subject_id', 'subj_id']:
                                id_col = col
                        
                        if sub_col and id_col:
                            # Check if item_id matches any row
                            for _, row in sheet_df.iterrows():
                                sub = row.get(sub_col, '')
                                sub_id = row.get(id_col, '')
                                if pd.notna(sub) and pd.notna(sub_id):
                                    try:
                                        # Handle different data types more carefully
                                        if isinstance(sub, (int, float)):
                                            sub_str = str(int(sub))
                                        else:
                                            sub_str = str(sub).strip()
                                        
                                        if isinstance(sub_id, (int, float)):
                                            sub_id_str = str(int(sub_id))
                                        else:
                                            sub_id_str = str(sub_id).strip()
                                        
                                        if not sub_str.startswith('sub'):
                                            sub_str = f"sub{sub_str}"
                                        
                                        filename_base = f"{sub_str}_{sub_id_str}"
                                        if filename_base == item_id:
                                            return jsonify({
                                                'success': True,
                                                'files': [{
                                                    'name': input_path_obj.name,
                                                    'path': _path_for_client(input_path_obj),
                                                    'size': input_path_obj.stat().st_size
                                                }]
                                            })
                                    except Exception as e:
                                        print(f"DEBUG: Error processing row for Excel matching: {e}")
                                        continue
                except Exception as e:
                    print(f"Warning: Could not read Excel file {input_path_obj}: {e}")
                    import traceback
                    traceback.print_exc()
                    # For sentenceCorrect, if we can't read the Excel, still return it
                    # since it likely contains the data
                    if step_type == 'sentenceCorrect':
                        return jsonify({
                            'success': True,
                            'files': [{
                                'name': input_path_obj.name,
                                'path': _path_for_client(input_path_obj),
                                'size': input_path_obj.stat().st_size
                            }]
                    })
                    # For other types, return empty if we can't verify
                    return jsonify({'success': True, 'files': []})
            else:
                # For other file types, check if filename matches item_id or return it anyway
                file_stem = input_path_obj.stem
                if item_id in file_stem or file_stem in item_id or step_type == 'sentenceCorrect':
                    # For sentenceCorrect, return the file even if item_id doesn't match
                    # (the file might contain multiple subjects)
                    return jsonify({
                        'success': True,
                        'files': [{
                            'name': input_path_obj.name,
                            'path': _path_for_client(input_path_obj),
                            'size': input_path_obj.stat().st_size
                        }]
                    })
                else:
                    return jsonify({'success': True, 'files': []})
        
        # If it's a directory, proceed with directory-based file discovery
        input_dir = input_path_obj

        # Flexible recognition for alternative-named / csv inputs. A step's input is the
        # previous step's OUTPUT, so recognise by that step's tokens (handles e.g.
        # ``story-alice_14_sub-3008-parsed.csv`` as the textMatching input — issue: the
        # input checkbox showed unchecked despite the file existing).
        _input_recog_type = {
            'textMatching': 'textParsing',     # parsed recall
            'textParsing': 'sentenceCorrect',  # corrected text
            'causalRating': 'eventSegment',    # story events
        }.get(step_type)
        if _input_recog_type:
            try:
                from helpers.step_files import find_step_files
                recs = find_step_files(input_dir, _input_recog_type, item_id,
                                       is_story=(step_type == 'causalRating'))
                if recs:
                    return jsonify({'success': True, 'files': [
                        {'name': f.name, 'path': _path_for_client(f), 'size': f.stat().st_size}
                        for f in recs]})
            except Exception as _e:
                print(f"DEBUG: recognizer input check failed: {_e}")

        # Determine file patterns based on step type and item_id
        files = []
        found_files = set()
        
        if step_type == 'sentenceCorrect' and input_dir.is_dir():
            exact_txt = input_dir / f"{item_id}.txt"
            if exact_txt.is_file():
                return jsonify({
                    'success': True,
                    'files': [{
                        'name': exact_txt.name,
                        'path': _path_for_client(exact_txt),
                        'size': exact_txt.stat().st_size
                    }]
                })
        
        # First, do a quick direct check for the most common case (exact filename match)
        # This is faster and more reliable than pattern matching
        if step_type == 'textParsing':
            # Check for exact match first
            exact_file = input_dir / f"{item_id}.txt"
            if exact_file.exists() and exact_file.is_file():
                print(f"DEBUG: Found exact match: {exact_file.name}")
                try:
                    files.append({
                        'name': exact_file.name,
                        'path': _path_for_client(exact_file),
                        'size': exact_file.stat().st_size
                    })
                    found_files.add(exact_file.name)
                except Exception as e:
                    print(f"DEBUG: Error getting file stats for {exact_file}: {e}")
                    files.append({
                        'name': exact_file.name,
                        'path': _path_for_client(exact_file),
                        'size': 0
                    })
                    found_files.add(exact_file.name)
                # If we found the exact file, we can return early (but still check for other patterns in case there are multiple)
                # Actually, let's continue to check patterns in case there are variations, but we've already found the main one
        
        # For some steps, we need to check multiple locations or use more flexible patterns
        if step_type == 'audioTranscribe:story':
            patterns = []
            for ext in SUPPORTED_AUDIO_EXTENSIONS:
                patterns.extend([f"{item_id}{ext}", f"{item_id}*{ext}", f"*{item_id}*{ext}"])
        elif step_type == 'audioTranscribe:recall':
            patterns = []
            for ext in SUPPORTED_AUDIO_EXTENSIONS:
                patterns.extend([
                    f"{item_id}_recall*{ext}", f"{item_id}_*{ext}",
                    f"*{item_id}*recall*{ext}", f"*{item_id}*{ext}"
                ])
        elif step_type == 'eventSegment':
            patterns = [
                f"{item_id}.txt", f"{item_id}*.txt",
                f"*{item_id}*.txt",
                f"{item_id}.xlsx", f"{item_id}*.xlsx"
            ]
        elif step_type == 'sentenceCorrect':
            # For sentenceCorrect, input might be from Excel files or raw text
            # Check the configured input path first
            patterns = [
                f"{item_id}.txt", f"{item_id}*.txt",
                f"*{item_id}*.txt"
            ]
            # Also check Excel files that might contain raw recall text
            patterns.extend([f"*{item_id}*.xlsx"])
        elif step_type == 'textParsing':
            # For textParsing, input is corrected text files (from previous step output)
            # Try multiple patterns to catch various naming conventions
            patterns = [
                f"{item_id}.txt",  # Exact match first (most common case)
                f"{item_id}*.txt",  # With any suffix
                f"*{item_id}*.txt",  # With any prefix/suffix
            ]
            # Exclude audio transcription files (but be more specific)
            exclude_patterns = ['_recall_audio', '_recall_transcribed', '_audio_transcribed']
        elif step_type == 'textMatching':
            # For textMatching, input is parsed files
            # Try multiple patterns to catch various naming conventions
            patterns = [
                f"{item_id}_parsed.xlsx",
                f"{item_id}*_parsed.xlsx",
                f"*{item_id}*_parsed.xlsx",
                # Also try without underscores
                f"{item_id.replace('_', '')}_parsed.xlsx",
                f"*{item_id.replace('_', '')}*_parsed.xlsx",
                # Try variations
                f"{item_id}_parsed*.xlsx",
                f"*{item_id}*parsed*.xlsx"
            ]
        elif step_type == 'causalRating':
            patterns = [
                f"{item_id}_events.xlsx",
                f"{item_id}_events-*.xlsx",
                f"*{item_id}*_events*.xlsx",
            ]
        else:
            patterns = [f"{item_id}*", f"*{item_id}*"]
        
        # Search for files matching patterns (textMatching may read textParsing output from _prev/)
        input_glob_roots = [input_dir]
        if step_type == 'textMatching':
            _prev_in = input_dir / '_prev'
            if _prev_in.is_dir():
                input_glob_roots.append(_prev_in)

        print(f"DEBUG: Searching in directories: {input_glob_roots}")
        print(f"DEBUG: Primary input exists: {input_dir.exists()}")
        print(f"DEBUG: Item ID: {item_id}")
        print(f"DEBUG: Patterns to search: {patterns}")
        
        for _root in input_glob_roots:
            if _root.exists() and _root.is_dir():
                all_files = list(_root.iterdir())
                print(f"DEBUG: Files in {_root} ({len(all_files)}): {[f.name for f in all_files[:10]]}")
        
        for search_root in input_glob_roots:
            if not search_root.exists() or not search_root.is_dir():
                continue
            for pattern in patterns:
                try:
                    matching_files = list(search_root.glob(pattern))
                    print(f"DEBUG: Pattern '{pattern}' in {search_root} matched {len(matching_files)} files: {[f.name for f in matching_files[:5]]}")

                    for file_path in matching_files:
                        if file_path.is_file() and file_path.name not in found_files:
                            # Skip human-edit files for input listing unless this step accepts them as primary input
                            # (textParsing "complete" includes *_parsed_*-edit.xlsx; textMatching must list the same)
                            if is_user_edit_file(file_path.name) and step_type != 'textMatching':
                                continue

                            # For textParsing, exclude audio transcription files
                            if step_type == 'textParsing':
                                file_lower = file_path.name.lower()
                                if any(exclude in file_lower for exclude in ['_recall_audio', '_recall_transcribed', '_audio_transcribed', '_transcribed_audio']):
                                    print(f"DEBUG: Excluding audio transcription file: {file_path.name}")
                                    continue

                            found_files.add(file_path.name)
                            try:
                                files.append({
                                    'name': file_path.name,
                                    'path': _path_for_client(file_path),
                                    'size': file_path.stat().st_size
                                })
                                print(f"DEBUG: Added file: {file_path.name}")
                            except Exception:
                                files.append({
                                    'name': file_path.name,
                                    'path': _path_for_client(file_path),
                                    'size': 0
                                })
                                print(f"DEBUG: Added file (no stats): {file_path.name}")
                except Exception as e:
                    print(f"Warning: Error searching for pattern {pattern} in {search_root}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
        
        # Sort by filename
        # If no files found in configured input path, try common fallback locations
        if not files and step_type == 'sentenceCorrect':
            # For sentenceCorrect, first check default recall texts directory
            recall_texts_dir = PROJECT_ROOT / 'data' / '5_recall_texts'
            if recall_texts_dir.exists() and recall_texts_dir.is_dir():
                fallback_file = recall_texts_dir / f"{item_id}.txt"
                if fallback_file.exists() and fallback_file.is_file():
                    found_files.add(fallback_file.name)
                    files.append({
                        'name': fallback_file.name,
                        'path': _path_for_client(fallback_file),
                        'size': fallback_file.stat().st_size
                    })
            # If still no files, check Excel files in data directory
            if not files:
                excel_files = sorted((PROJECT_ROOT / 'data').glob('summary_*.xlsx')) if (PROJECT_ROOT / 'data').exists() else []
            else:
                excel_files = []
            for excel_file in excel_files:
                if excel_file.exists():
                    # Check if this Excel file might contain data for this item
                    try:
                        import pandas as pd
                        df = pd.read_excel(excel_file, sheet_name=None)
                        for sheet_name, sheet_df in df.items():
                            if 'sub' in sheet_df.columns and 'ID' in sheet_df.columns:
                                # Check if item_id matches any row
                                for _, row in sheet_df.iterrows():
                                    sub = row.get('sub', '')
                                    sub_id = row.get('ID', '')
                                    if pd.notna(sub) and pd.notna(sub_id):
                                        sub_str = str(int(sub)) if isinstance(sub, (int, float)) else str(sub).strip()
                                        sub_id_str = str(int(sub_id)) if isinstance(sub_id, (int, float)) else str(sub_id).strip()
                                        if sub_str.startswith('sub'):
                                            filename_base = f"{sub_str}_{sub_id_str}"
                                        else:
                                            filename_base = f"sub{sub_str}_{sub_id_str}"
                                        if filename_base == item_id:
                                            if excel_file.name not in found_files:
                                                found_files.add(excel_file.name)
                                                files.append({
                                                    'name': excel_file.name,
                                                    'path': _path_for_client(excel_file),
                                                    'size': excel_file.stat().st_size
                                                })
                                            break
                    except Exception as e:
                        print(f"Warning: Could not read Excel file {excel_file}: {e}")
                        continue
        
        # For textMatching, also check for story events files (required input)
        # This should happen regardless of whether parsed files were found
        if step_type == 'textMatching':
            # Story events are required input (from eventSegment output)
            events_dir = get_output_dir_for_step_type('eventSegment') or STORY_EVENTS_DIR
            event_filename_base = f"{item_id}_events"
            
            if events_dir.exists() and events_dir.is_dir():
                story_file_found = False
                # Try exact match first (most common case)
                story_file = events_dir / f"{event_filename_base}.xlsx"
                if story_file.exists() and story_file.is_file():
                    story_file_found = True
                    if story_file.name not in found_files:
                        found_files.add(story_file.name)
                        try:
                            files.append({
                                'name': story_file.name,
                                'path': _path_for_client(story_file),
                                'size': story_file.stat().st_size
                            })
                            print(f"DEBUG: Added story events file: {story_file.name}")
                        except Exception as e:
                            files.append({
                                'name': story_file.name,
                                'path': _path_for_client(story_file),
                                'size': 0
                            })
                            print(f"DEBUG: Added story events file (no stats): {story_file.name}")
                
                # If exact match not found, try method-suffixed files
                if not story_file_found:
                    story_events_patterns = [
                        f"{event_filename_base}_*-edit.xlsx",
                        f"{event_filename_base}_rule-based.xlsx",
                        f"{event_filename_base}_api.xlsx",
                        f"{event_filename_base}-*.xlsx"
                    ]
                    for pattern in story_events_patterns:
                        try:
                            matching_files = list(events_dir.glob(pattern))
                            for file_path in matching_files:
                                if file_path.is_file() and file_path.name not in found_files:
                                    found_files.add(file_path.name)
                                    try:
                                        files.append({
                                            'name': file_path.name,
                                            'path': _path_for_client(file_path),
                                            'size': file_path.stat().st_size
                                        })
                                        print(f"DEBUG: Added story events file: {file_path.name}")
                                        story_file_found = True
                                        break  # Use the first matching file
                                    except Exception as e:
                                        files.append({
                                            'name': file_path.name,
                                            'path': _path_for_client(file_path),
                                            'size': 0
                                        })
                                        print(f"DEBUG: Added story events file (no stats): {file_path.name}")
                                        story_file_found = True
                                        break
                            if story_file_found:
                                break
                        except Exception as e:
                            print(f"Warning: Error searching for story events pattern {pattern}: {e}")
                            continue
                if not story_file_found:
                    print(f"DEBUG: Story events file not found for {item_id} in {events_dir}")
        
        files.sort(key=lambda x: x['name'])
        
        # If no files found with patterns, try a more lenient search
        lenient_roots = [input_dir]
        if step_type == 'textMatching':
            _prev_lenient = input_dir / '_prev'
            if _prev_lenient.is_dir():
                lenient_roots.append(_prev_lenient)

        if not files and any(d.exists() and d.is_dir() for d in lenient_roots):
            print(f"DEBUG: No files found with patterns, trying lenient search...")
            # Try to find any file that contains the item_id (case-insensitive)
            # Also try matching without underscores, and try extracting subject ID from filename
            item_id_variations = [
                item_id,
                item_id.lower(),
                item_id.replace('_', ''),
                item_id.replace('_', '').lower()
            ]
            # Extract just the numeric parts if item_id is like "subN_XXXX"
            import re
            match = re.search(r'sub(\d+)[_\s]*(\d+)', item_id, re.IGNORECASE)
            if match:
                sub_num = match.group(1)
                id_num = match.group(2)
                item_id_variations.extend([
                    f"sub{sub_num}_{id_num}",
                    f"sub{sub_num}{id_num}",
                    f"{sub_num}_{id_num}",
                    f"{sub_num}{id_num}"
                ])

            for lenient_dir in lenient_roots:
                if not lenient_dir.exists() or not lenient_dir.is_dir():
                    continue
                for file_path in lenient_dir.iterdir():
                    if file_path.is_file():
                        # Skip hidden files and common non-data files
                        if file_path.name.startswith('.') or file_path.suffix in ['.py', '.pyc', '.log', '.json']:
                            continue

                        file_name_lower = file_path.name.lower()
                        file_stem_lower = file_path.stem.lower()

                        matched = False
                        for variation in item_id_variations:
                            if variation.lower() in file_name_lower or variation.lower() in file_stem_lower:
                                matched = True
                                break

                        if matched:
                            # Skip human-edit files unless that step treats edits as valid primary input
                            if is_user_edit_file(file_path.name) and step_type not in ('textParsing', 'textMatching'):
                                continue
                            if step_type == 'textParsing':
                                file_lower = file_path.name.lower()
                                if any(exclude in file_lower for exclude in ['_recall_audio', '_recall_transcribed', '_audio_transcribed', '_transcribed_audio']):
                                    print(f"DEBUG: Excluding audio transcription file in lenient search: {file_path.name}")
                                    continue
                                if file_path.suffix.lower() != '.txt':
                                    continue
                            if step_type == 'textMatching':
                                if '_parsed' not in file_path.name.lower() or file_path.suffix.lower() != '.xlsx':
                                    continue

                            if file_path.name not in found_files:
                                found_files.add(file_path.name)
                                try:
                                    files.append({
                                        'name': file_path.name,
                                        'path': _path_for_client(file_path),
                                        'size': file_path.stat().st_size
                                    })
                                    print(f"DEBUG: Found file via lenient search: {file_path.name}")
                                except Exception:
                                    files.append({
                                        'name': file_path.name,
                                        'path': _path_for_client(file_path),
                                        'size': 0
                                    })
        
        files.sort(key=lambda x: x['name'])
        
        print(f"DEBUG: Found {len(files)} files for {item_id}, step {step_index}, type {step_type}")
        if files:
            print(f"DEBUG: Files found: {[f['name'] for f in files]}")
        else:
            print(f"DEBUG: No files found. Checked directory: {input_dir}")
            if input_dir.exists():
                all_files = [f.name for f in input_dir.iterdir() if f.is_file()]
                print(f"DEBUG: All files in directory: {all_files[:20]}")  # Show first 20
        
        return jsonify({
            'success': True,
            'files': files
        })
    except Exception as e:
        import traceback
        print(f"DEBUG: Exception in get_input_files: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/step/<int:step_index>/input-variants', methods=['GET'])
def get_step_input_variants(step_index):
    """Return available input-variant suffixes for ``step_index`` intersected across ``item_ids``.

    Query:
        ?item_ids=id1,id2,...   (comma-separated; required)

    Response:
        {
            "success": true,
            "step_index": N,
            "step_type": "...",
            "streams": [
                {
                    "key": "main" | "parsed_recall" | "story_events",
                    "label": "...",
                    "env_var": "BATCH_INPUT_VARIANT" | "BATCH_STORY_EVENTS_VARIANT",
                    "variants": [
                        {"suffix": "", "label": "canonical", "count": N},
                        {"suffix": "_spell-ollama_gemma4_e4b", "label": "spell-ollama gemma4 e4b", "count": N},
                        ...
                    ]
                },
                ...
            ]
        }
    """
    try:
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400

        if step_index < 0 or step_index >= len(pipeline_config['steps']):
            return jsonify({'success': False, 'error': 'Invalid step index'}), 400

        step = pipeline_config['steps'][step_index]
        step_type = step_runtime_key(step)
        input_path = step.get('inputPath', '')

        raw_ids = request.args.get('item_ids', '').strip()
        item_ids = [i for i in (s.strip() for s in raw_ids.split(',')) if i]
        if not item_ids:
            return jsonify({'success': False, 'error': 'item_ids query param required'}), 400

        data = enumerate_step_input_variants(step_type, item_ids, input_path, step=step)
        return jsonify({
            'success': True,
            'step_index': step_index,
            'step_type': step_type,
            'streams': data.get('streams', []),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/item/<item_id>/step/<int:step_index>/output-files', methods=['GET'])
def get_output_files(item_id, step_index):
    """Get output files for a specific item and step."""
    try:
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        
        if step_index < 0 or step_index >= len(pipeline_config['steps']):
            return jsonify({'success': False, 'error': 'Invalid step index'}), 400
        
        step = pipeline_config['steps'][step_index]
        step_type = step_runtime_key(step)
        output_path = step.get('outputPath', '')
        
        if not output_path:
            return jsonify({'success': True, 'files': []})
        
        # Normalize output path - remove trailing slashes
        output_path = output_path.rstrip('/')
        if output_path.startswith('data/') or output_path.startswith('output/'):
            output_dir = PROJECT_ROOT / output_path
        elif output_path.startswith('/'):
            output_dir = Path(output_path)
        else:
            output_dir = PROJECT_ROOT / output_path
        
        if not output_dir.exists() or not output_dir.is_dir():
            return jsonify({'success': True, 'files': []})
        
        # Get file patterns based on step type
        patterns = []
        if step_type == 'audioTranscribe:story':
            patterns = [f'*{item_id}*_story_audio*.txt', f'*{item_id}*_story_audio*.json']
        elif step_type == 'audioTranscribe:recall':
            patterns = [f'*{item_id}*_recall_audio*.txt', f'*{item_id}*_recall_audio*.json']
        elif step_type == 'eventSegment':
            patterns = [f'{item_id}_events*.xlsx', f'{item_id}_events-*.xlsx']
        elif step_type == 'sentenceCorrect':
            patterns = [f'*{item_id}*.txt']
        elif step_type == 'textParsing':
            patterns = [f'*{item_id}*_parsed.xlsx']
        elif step_type == 'textMatching':
            patterns = [f'*{item_id}*_rate-recall.xlsx', f'*{item_id}*_rate-recall-*.xlsx']
        elif step_type == 'causalRating':
            patterns = [f'{item_id}_causal-*.xlsx', f'{item_id}_causal.xlsx']
        else:
            # Default: look for any file containing item_id
            patterns = [f'*{item_id}*']
        
        files = []
        found_files = set()
        
        for pattern in patterns:
            matching_files = list(output_dir.glob(pattern))
            for file_path in matching_files:
                if file_path.is_file() and file_path.name not in found_files:
                    found_files.add(file_path.name)
                    try:
                        file_size = file_path.stat().st_size
                    except:
                        file_size = 0
                    
                    files.append({
                        'name': file_path.name,
                        'path': _path_for_client(file_path),
                        'size': file_size
                    })
        
        files.sort(key=lambda x: x['name'])
        
        return jsonify({
            'success': True,
            'files': files
        })
    except Exception as e:
        import traceback
        print(f"DEBUG: Exception in get_output_files: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/item/<item_id>/step/<int:step_index>/completed-methods', methods=['GET'])
def get_completed_methods(item_id, step_index):
    """Get which methods have been completed for a given item and step, including API trial counts."""
    try:
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        if step_index < 0 or step_index >= len(pipeline_config['steps']):
            return jsonify({'success': False, 'error': 'Invalid step index'}), 400

        step = pipeline_config['steps'][step_index]
        step_type = step_runtime_key(step)
        output_path = step.get('outputPath', '')
        if not output_path:
            return jsonify({'success': True, 'methods': {}})

        output_path = output_path.rstrip('/')
        if output_path.startswith('data/') or output_path.startswith('output/'):
            output_dir = PROJECT_ROOT / output_path
        elif output_path.startswith('/'):
            output_dir = Path(output_path)
        else:
            output_dir = PROJECT_ROOT / output_path

        if not output_dir.exists() or not output_dir.is_dir():
            return jsonify({'success': True, 'methods': {}})

        methods = {}

        if step_type == 'eventSegment':
            for f in output_dir.glob(f'{item_id}_events-*.xlsx'):
                if is_user_edit_file(f.name):
                    continue
                stem = f.stem
                suffix_part = stem.replace(f'{item_id}_events-', '')
                trial_match = re.match(r'^(.+?)_trial(\d+)$', suffix_part)
                if trial_match:
                    method_id = trial_match.group(1)
                    trial_num = int(trial_match.group(2))
                else:
                    method_id = suffix_part
                    trial_num = 1
                is_api = method_id.startswith('api_')
                display_method = 'api' if is_api else method_id
                if display_method not in methods:
                    methods[display_method] = {'completed': True, 'is_api': is_api, 'trials': 0, 'files': []}
                methods[display_method]['trials'] = max(methods[display_method]['trials'], trial_num)
                methods[display_method]['files'].append(f.name)
            # Legacy filename {id}_events.xlsx (no method suffix) may be manual, old default output,
            # or an edit — do not attribute it to fine-grained only (that incorrectly disabled/confused
            # the method modal vs coarse/clause, which always run from the transcript in-script).
            if has_any_edit_file(output_dir, item_id, '_events', '.xlsx'):
                methods.setdefault('manual', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
        elif step_type == 'causalRating':
            for f in output_dir.glob(f'{item_id}_causal-*.xlsx'):
                if is_user_edit_file(f.name):
                    continue
                stem = f.stem
                suffix_part = stem.replace(f'{item_id}_causal-', '')
                trial_match = re.match(r'^(.+?)_trial(\d+)$', suffix_part)
                if trial_match:
                    method_id = trial_match.group(1)
                    trial_num = int(trial_match.group(2))
                else:
                    method_id = suffix_part
                    trial_num = 1
                is_api = method_id.startswith('api_')
                display_method = 'api' if is_api else method_id
                if display_method not in methods:
                    methods[display_method] = {'completed': True, 'is_api': is_api, 'trials': 0, 'files': []}
                methods[display_method]['trials'] = max(methods[display_method]['trials'], trial_num)
                methods[display_method]['files'].append(f.name)
        elif step_type == 'sentenceCorrect':
            for f in output_dir.glob(f'{item_id}*.txt'):
                if is_user_edit_file(f.name):
                    continue
                if f.name == f'{item_id}.txt':
                    methods.setdefault('auto', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
                    methods['auto']['files'].append(f.name)
                elif '_spell-gemma-hf_' in f.name:
                    methods.setdefault('gemma-hf', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
                    methods['gemma-hf']['files'].append(f.name)
                elif '_spell-ollama_' in f.name:
                    methods.setdefault('gemma-ollama', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
                    methods['gemma-ollama']['files'].append(f.name)
        elif step_type == 'textParsing':
            for f in output_dir.glob(f'{item_id}_parsed*.xlsx'):
                if is_user_edit_file(f.name):
                    continue
                if f.name == f'{item_id}_parsed.xlsx':
                    methods.setdefault('auto', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
                    methods['auto']['files'].append(f.name)
                elif '_parsed-ollama_' in f.name or '_parsed-e4b' in f.name:
                    methods.setdefault('gemma-ollama', {'completed': True, 'is_api': False, 'trials': 1, 'files': []})
                    methods['gemma-ollama']['files'].append(f.name)
        else:
            all_output_files = []
            patterns_for_type = {
                'audioTranscribe:story': [f'{item_id}*.txt'],
                'audioTranscribe:recall': [f'{item_id}*.txt'],
                'textMatching': [f'{item_id}_rate-recall*.xlsx'],
            }
            for pattern in patterns_for_type.get(step_type, [f'*{item_id}*']):
                all_output_files.extend(output_dir.glob(pattern))
            if all_output_files:
                methods['auto'] = {'completed': True, 'is_api': False, 'trials': 1, 'files': [f.name for f in all_output_files]}

        return jsonify({'success': True, 'methods': methods})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# Test route to verify server is working
@app.route('/api/test-execute', methods=['GET', 'POST'])
def test_execute():
    """Test route to verify execute endpoint is accessible."""
    return jsonify({'success': True, 'message': 'Execute route is accessible'})

# Test route to verify item/step routing works
@app.route('/api/item/<item_id>/step/<int:step_index>/test', methods=['GET', 'POST'])
def test_item_step_route(item_id, step_index):
    """Test route to verify item/step routing pattern works."""
    return jsonify({
        'success': True, 
        'message': 'Item/step route pattern works',
        'item_id': item_id,
        'step_index': step_index
    })

def _execute_manual_step(item_id, step_type, input_path, output_path):
    """Handle 'manual' method: copy input file content to output directory as a starter file for user editing."""
    try:
        input_dir = PROJECT_ROOT / input_path if input_path and not Path(input_path).is_absolute() else Path(input_path) if input_path else None
        output_dir = PROJECT_ROOT / output_path if output_path and not Path(output_path).is_absolute() else Path(output_path) if output_path else None
        
        if not output_dir:
            return jsonify({'success': False, 'error': 'No output path configured'}), 400
        output_dir.mkdir(parents=True, exist_ok=True)
        
        if step_type == 'sentenceCorrect':
            # Copy raw recall text as-is to the output directory
            if not input_dir or not input_dir.exists():
                return jsonify({'success': False, 'error': 'Input directory not found'}), 404
            src_file = None
            for pattern in [f"{item_id}.txt", f"{item_id}*.txt", f"*{item_id}*.txt"]:
                matches = list(input_dir.glob(pattern))
                matches = [f for f in matches if not is_user_edit_file(f.name)]
                if matches:
                    src_file = matches[0]
                    break
            if not src_file:
                return jsonify({'success': False, 'error': f'Input file not found for {item_id}'}), 404
            raw_text = src_file.read_text(encoding='utf-8').strip()
            out_file = output_dir / f"{item_id}.txt"
            with open(out_file, 'w', encoding='utf-8') as f:
                f.write(f"{item_id}.txt\n")
                f.write(raw_text)
            return jsonify({'success': True, 'message': f'Input text copied to output for manual editing ({len(raw_text)} chars)'})
        
        elif step_type == 'textParsing':
            # Read recall text and create single-segment parsed file for hand-editing.
            corrected_text = _resolve_recall_text_for_manual(item_id, input_dir)
            if not corrected_text:
                return jsonify({
                    'success': False,
                    'error': (
                        f'No recall text found for {item_id}. '
                        f'Add {item_id}.txt under the sentenceCorrect input folder '
                        f'(e.g. data/5_recall_texts/) or run sentenceCorrect first.'
                    ),
                }), 404
            df = pd.DataFrame({
                'recalled_events': [''] * 1,
                'recall_in_temporal_order': [corrected_text]
            })
            out_file = output_dir / f"{item_id}_parsed.xlsx"
            df.to_excel(out_file, index=False, engine='openpyxl', na_rep='')
            return jsonify({'success': True, 'message': f'Recall text placed as single segment for manual parsing'})
        
        elif step_type == 'textMatching':
            # Read parsed segments and create rated file with empty event matches
            parsed_texts = get_parsed_texts(item_id)
            if not parsed_texts:
                fallback_text = _resolve_recall_text_for_manual(item_id, input_dir)
                if fallback_text:
                    parsed_texts = [{'text': fallback_text}]
            if not parsed_texts:
                return jsonify({'success': False, 'error': f'No recall text found for {item_id}. Run textParsing first or add recall text under data/5_recall_texts/.'}), 404
            segments = [seg.get('text', '') for seg in parsed_texts]
            df = pd.DataFrame({
                'recalled_events': [''] * len(segments),
                'recall_in_temporal_order': segments
            })
            out_file = output_dir / f"{item_id}_rate-recall.xlsx"
            df.to_excel(out_file, index=False, engine='openpyxl', na_rep='')
            return jsonify({'success': True, 'message': f'Parsed segments copied with empty event matches for manual rating ({len(segments)} segments)'})
        
        elif step_type == 'eventSegment':
            # Read story transcript and create single-event events file
            transcript = get_story_transcript(item_id, is_story=True)
            if not transcript:
                transcript = get_story_transcript(item_id, is_story=False)
            if not transcript:
                if not input_dir or not input_dir.exists():
                    return jsonify({'success': False, 'error': f'Story transcript not found for {item_id}'}), 404
                for pattern in [f"{item_id}.txt", f"{item_id}*.txt", f"{item_id}*.csv", f"{item_id}*.tsv", f"{item_id}*.xlsx", f"*{item_id}*.txt", f"*{item_id}*.csv", f"*{item_id}*.xlsx"]:
                    matches = list(input_dir.glob(pattern))
                    matches = [f for f in matches if not is_user_edit_file(f.name)]
                    if matches:
                        from helpers.flexible_io import read_document_text
                        transcript = read_document_text(matches[0])
                        break
            if not transcript:
                return jsonify({'success': False, 'error': f'Story transcript not found for {item_id}'}), 404
            df = pd.DataFrame({
                'event': [1],
                'story_texts': [transcript]
            })
            # Write a method-suffixed file (like clause/fine/coarse/api) instead of the
            # canonical {item_id}_events.xlsx. Using the bare name would overwrite an
            # existing segmented events file with this single-event starter — violating
            # the "never overwrite without a suffix" versioning rule — and would not be
            # recognised as a completed 'manual' run by the dashboard's method scan,
            # which keys off the {item_id}_events-*.xlsx pattern.
            out_file = output_dir / f"{item_id}_events-manual.xlsx"
            df.to_excel(out_file, index=False, engine='openpyxl')
            return jsonify({'success': True, 'message': f'Story transcript placed as single event for manual segmentation'})
        
        elif step_type == 'causalRating':
            # Create an empty causal rating scaffold file
            df = pd.DataFrame(columns=['event_A_number', 'event_B_number', 'rating', 'reasoning'])
            out_file = output_dir / f"{item_id}_causal-manual.xlsx"
            df.to_excel(out_file, index=False, engine='openpyxl', na_rep='')
            return jsonify({'success': True, 'message': f'Empty causal rating file created for manual entry'})
        
        elif step_type in ('audioTranscribe:story', 'audioTranscribe:recall'):
            # Create an empty transcript file for the user to fill in manually
            out_file = output_dir / f"{item_id}.txt"
            with open(out_file, 'w', encoding='utf-8') as f:
                f.write('')
            return jsonify({'success': True, 'message': f'Empty transcript file created for manual transcription'})
        
        else:
            return jsonify({'success': False, 'error': f'Manual method not supported for step type: {step_type}'}), 400
    
    except Exception as e:
        print(f"Error in manual step execution: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# Execute step route - MUST be defined before error handler
@app.route('/api/item/<item_id>/step/<int:step_index>/execute', methods=['POST'])
def execute_step(item_id, step_index):
    """Execute a single step for a specific item - works exactly like batch_process but for one item."""
    print(f"=== EXECUTE STEP ROUTE HIT ===")
    print(f"DEBUG: execute_step called with item_id='{item_id}', step_index={step_index}")
    try:
        import subprocess
        import sys
        import os
        from flask import request
        
        # Get request data (method, API key, and step options)
        request_data = request.get_json() or {}
        method = request_data.get('method')
        api_key = request_data.get('api_key')
        step_options = request_data.get('options', {})
        
        print(f"DEBUG: Method={method}, API key provided={bool(api_key)}, options={step_options}")
        
        # Get pipeline configuration - same as batch_process
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        
        if step_index < 0 or step_index >= len(pipeline_config['steps']):
            return jsonify({'success': False, 'error': 'Invalid step index'}), 400
        
        # Get step configuration - same as batch_process
        step = pipeline_config['steps'][step_index]
        step_type = step_runtime_key(step)
        input_path = step.get('inputPath', '')
        output_path = step.get('outputPath', '')

        if method and step_type == 'sentenceCorrect' and method in ('gemma-hf', 'gemma'):
            method = 'gemma-ollama'
        
        print(f"Starting single-item processing for {step_type} (item: {item_id})")
        print(f"  Input: {input_path}")
        for field, env_var, resolved in _resolve_step_extra_input_dirs(step):
            print(f"  {field}: {step.get(field, '')} -> {env_var}={resolved}")
        print(f"  Output: {output_path}")
        if method:
            print(f"  Method: {method}")
        
        # Handle manual method: copy input content to output for user editing
        if method == 'manual':
            return _execute_manual_step(item_id, step_type, input_path, output_path)
        
        # Treat 'auto' as no specific method (run default script behavior)
        if method == 'auto':
            method = None
        
        # Map step types to scripts - same as batch_process
        script_map = {
            'audioTranscribe:story': '1_audio-transcribe.py',
            'audioTranscribe:recall': '1_audio-transcribe.py',
            'eventSegment': '2_story-event-segment.py',
            'sentenceCorrect': '3_spell-grammar-correct.py',
            'textParsing': '4_parse-texts.py',
            'textMatching': '5_recall-rater.py',
            'causalRating': '6_causal-rater.py'
        }
        
        script_name = script_map.get(step_type)
        if not script_name:
            return jsonify({'success': False, 'error': f'Processing not available for step type: {step_type}'}), 400
        
        script_path = SCRIPTS_DIR / script_name
        if not script_path.exists():
            return jsonify({'success': False, 'error': f'Script not found: {script_name}'}), 404
        
        # Build command - same as batch_process
        cmd = [sys.executable, str(script_path)]
        
        # Set environment variables - same as batch_process, but ADD BATCH_ITEM_ID for single item
        env = os.environ.copy()
        
        # Add method argument if provided (for scripts that support it)
        if method and step_type == 'eventSegment':
            cmd.extend(['--method', method])
            # For API method, pass model and prompt version via env vars
            if method == 'api':
                api_model = request_data.get('model')
                prompt_version = request_data.get('prompt_version')
                if api_model:
                    cmd.extend(['--model', api_model])
                    print(f"  Model: {api_model}")
                if prompt_version:
                    cmd.extend(['--prompt-version', prompt_version])
                    print(f"  Prompt version: {prompt_version}")
                # Determine which API key env var to set based on model provider
                if api_key:
                    provider = 'anthropic'
                    if api_model:
                        if api_model in _EVENT_SEGMENT_OLLAMA_MODEL_KEYS:
                            provider = 'ollama'
                        else:
                            provider = provider_for_model(api_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    elif provider == 'ollama':
                        pass
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
                    print(f"DEBUG: API key set for {provider}")
        elif method and step_type == 'textMatching':
            if method == 'test-mode':
                env['TEST_MODE'] = '1'
            elif method == 'api':
                env.pop('TEST_MODE', None)
                recall_model = request_data.get('model')
                recall_prompt = request_data.get('prompt_version')
                if recall_model:
                    env['RECALL_RATING_MODEL'] = recall_model
                    print(f"  RECALL_RATING_MODEL={recall_model}")
                if recall_prompt:
                    # script 5 loads scripts/prompt/<name>.txt via RECALL_RATING_PROMPT (no extension)
                    env['RECALL_RATING_PROMPT'] = (
                        recall_prompt[:-4] if recall_prompt.endswith('.txt') else recall_prompt
                    )
                    print(f"  RECALL_RATING_PROMPT={env['RECALL_RATING_PROMPT']}")
                # Route the supplied key to the provider matching the selected model.
                if api_key:
                    provider = provider_for_model(recall_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
                    print(f"DEBUG: API key set for {provider}")
            elif method == 'gemma-ollama':
                env.pop('TEST_MODE', None)
                env['RECALL_RATING_BACKEND'] = 'ollama'
            elif method == 'rmatch':
                env.pop('TEST_MODE', None)
                env['RECALL_RATING_BACKEND'] = 'rmatch'
                opts = step_options or {}
                rmatch_model = opts.get('rmatch_model') or opts.get('model')
                if rmatch_model:
                    env['RMATCH_MODEL_NAME'] = str(rmatch_model)
                    print(f"  RMATCH_MODEL_NAME={rmatch_model}")
                for src_key, env_key in (
                    ('rmatch_quantization', 'RMATCH_QUANTIZATION'),
                    ('rmatch_window_size', 'RMATCH_WINDOW_SIZE'),
                    ('rmatch_prompt', 'RMATCH_PROMPT'),
                    ('rmatch_batch_size', 'RMATCH_BATCH_SIZE'),
                    ('rmatch_force_cpu', 'RMATCH_FORCE_CPU'),
                ):
                    v = opts.get(src_key)
                    if v not in (None, ''):
                        env[env_key] = str(v)
        elif method and step_type == 'sentenceCorrect':
            env['SPELL_GRAM_METHOD'] = method
        elif method and step_type == 'textParsing':
            if method == 'gemma-ollama':
                env['RECALL_PARSE_METHOD'] = 'ollama'
        elif step_type == 'causalRating':
            if method:
                cmd.extend(['--method', method])
            if method == 'api':
                causal_model = request_data.get('model')
                causal_prompt = request_data.get('prompt_version')
                if causal_model:
                    env['CAUSAL_RATING_MODEL'] = causal_model
                if causal_prompt:
                    env['CAUSAL_RATING_PROMPT'] = causal_prompt
                if api_key:
                    provider = 'anthropic'
                    if causal_model:
                        provider = provider_for_model(causal_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
        
        # Set API key if provided (generic fallback for non-eventSegment steps)
        if api_key and step_type != 'eventSegment':
            if step_type == 'textMatching' and method in ('gemma-ollama', 'rmatch'):
                pass
            else:
                env['ANTHROPIC_API_KEY'] = api_key
                print("DEBUG: API key set from request")
        
        _apply_step_path_env(env, step)
        env['BATCH_STEP_TYPE'] = step_type
        env['BATCH_ITEM_ID'] = item_id  # This is the only difference from batch_process

        # Forward input-variant selections from the launcher dropdown(s).
        # ``input_variant`` is the empty-or-suffix string for the primary input stream.
        # ``story_events_variant`` is the (optional) suffix for the story-events stream
        # consumed by textMatching and causalRating.
        input_variant = request_data.get('input_variant')
        if input_variant is not None:
            env['BATCH_INPUT_VARIANT'] = str(input_variant)
            print(f"  BATCH_INPUT_VARIANT={env['BATCH_INPUT_VARIANT']!r}")
        story_events_variant = request_data.get('story_events_variant')
        if story_events_variant is not None:
            env['BATCH_STORY_EVENTS_VARIANT'] = str(story_events_variant)
            print(f"  BATCH_STORY_EVENTS_VARIANT={env['BATCH_STORY_EVENTS_VARIANT']!r}")

        # Forward step options as environment variables
        if step_options:
            for opt_key, opt_val in step_options.items():
                if opt_val == '' or opt_val is None:
                    continue
                env_key = opt_key.upper()
                if isinstance(opt_val, bool):
                    env[env_key] = '1' if opt_val else '0'
                else:
                    env[env_key] = str(opt_val)
                print(f"  Option: {env_key}={env[env_key]}")

        from narraters.runtime_install import prepare_web_step

        prep_err = prepare_web_step(
            step_type=step_type,
            method=method,
            request_data=request_data,
            step_options=step_options,
        )
        if prep_err:
            return jsonify({"success": False, "error": prep_err}), 400

        gemma_report = _gemma4_preflight_report(step_type, method, step_options)
        if gemma_report is not None:
            if not gemma_report.get("ok"):
                return jsonify({
                    "success": False,
                    "error": "Gemma-4 environment check failed: "
                    + "; ".join(gemma_report.get("errors") or []),
                    "gemma_check": gemma_report,
                }), 400
            print(f"DEBUG: Gemma-4 preflight OK: {gemma_report.get('details')}")
            for w in gemma_report.get("warnings") or []:
                print(f"DEBUG: Gemma-4 preflight warning: {w}")

        print(f"Running: {' '.join(cmd)}")
        print(f"  BATCH_ITEM_ID={item_id}")
        print(f"  BATCH_INPUT_DIR={env.get('BATCH_INPUT_DIR')}")
        print(f"  BATCH_OUTPUT_DIR={env.get('BATCH_OUTPUT_DIR')}")
        
        # Run the script - same as batch_process
        result = subprocess.run(
            cmd,
            cwd=str(WORKSPACE_ROOT),
            env=env,
            capture_output=True,
            text=True,
            timeout=3600  # 1 hour timeout
        )
        
        print(f"Script completed with returncode={result.returncode}")
        if result.stdout:
            print(f"Output: {result.stdout[:500]}")
        if result.stderr:
            print(f"Error: {result.stderr[:500]}")
        
        # Return result - same as batch_process
        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': f'Step executed successfully for {item_id}',
                'output': result.stdout
            })
        else:
            display_error = _format_subprocess_step_error(result)
            
            # Log full error for debugging
            print(f"DEBUG: Full stderr:\n{result.stderr}")
            print(f"DEBUG: Full stdout:\n{result.stdout}")
            
            return jsonify({
                'success': False,
                'error': display_error,
                'output': result.stdout if result.stdout else 'No output',
                'full_stderr': result.stderr,
                'full_stdout': result.stdout,
                'returncode': result.returncode
            }), 500
            
    except subprocess.TimeoutExpired:
        print(f"DEBUG: Subprocess timed out for item_id='{item_id}', step_index={step_index}")
        return jsonify({'success': False, 'error': 'Processing timed out'}), 500
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        error_type = type(e).__name__
        error_message = str(e)
        print(f"DEBUG: Exception in execute_step: {error_type}: {error_message}")
        print(f"DEBUG: Full traceback:\n{error_traceback}")
        return jsonify({
            'success': False, 
            'error': f'{error_type}: {error_message}',
            'traceback': error_traceback
        }), 500


@app.route('/api/method-resource-check', methods=['POST'])
def api_method_resource_check():
    """Preflight: is the chosen method too heavy for this device?

    The UI calls this *before* launching a step. Returns a structured
    assessment ({heavy, severity, title, message, suggestion, ...}) so the
    frontend can pop a warning and offer a lighter method. This never runs or
    downloads a model; on any internal failure it fails open (severity 'ok')
    so it can't block legitimate runs.
    """
    try:
        data = request.get_json(silent=True) or {}
        step_type = str(data.get('step_type', '')).strip()
        method = data.get('method')
        options = data.get('options') if isinstance(data.get('options'), dict) else {}
        from helpers.resource_preflight import assess_method

        report = assess_method(step_type, method, options)
        return jsonify({'success': True, 'assessment': report})
    except Exception as e:
        print(f"method-resource-check error (failing open): {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': True,
            'assessment': {
                'heavy': False, 'severity': 'ok', 'title': '', 'message': '',
                'reasons': [], 'suggestion': {'method': None, 'label': ''},
                'details': {'preflight_error': str(e)},
            },
        })


# Debug route to catch 404s and see what URL was requested
# This must be defined AFTER all routes
@app.errorhandler(404)
def handle_404(e):
    """Handle 404 errors and log the requested URL."""
    try:
        print(f"\n{'='*70}")
        print(f"=== 404 ERROR ===")
        if hasattr(request, 'url'):
            print(f"Requested URL: {request.url}")
            print(f"Request path: {request.path}")
            print(f"Request method: {request.method}")
            print(f"Request args: {dict(request.args)}")
        print(f"\nAvailable routes containing 'execute' or 'item':")
        for rule in app.url_map.iter_rules():
            if 'execute' in rule.rule.lower() or 'item' in rule.rule.lower():
                print(f"  {rule.rule} -> {rule.endpoint} {list(rule.methods)}")
        print(f"{'='*70}\n")
        
        # Return JSON for API routes, HTML for others
        if hasattr(request, 'path') and request.path.startswith('/api/'):
            available_routes = [str(rule.rule) for rule in app.url_map.iter_rules() if 'execute' in rule.rule.lower() or 'item' in rule.rule.lower()]
            return jsonify({
                'success': False, 
                'error': f'Route not found: {request.path}', 
                'requested_url': request.url if hasattr(request, 'url') else 'unknown',
                'available_routes': available_routes,
                'message': 'The requested API endpoint was not found. Make sure the server has been restarted with the latest code.'
            }), 404
    except Exception as handler_error:
        # If error handler fails, try to return JSON anyway for API routes
        print(f"Error in 404 handler: {handler_error}")
        try:
            if hasattr(request, 'path') and request.path.startswith('/api/'):
                return jsonify({
                    'success': False,
                    'error': f'Route not found: {request.path if hasattr(request, "path") else "unknown"}',
                    'handler_error': str(handler_error)
                }), 404
        except:
            pass
    
    # Return default Flask 404 for non-API routes
    return e if e else ('Not Found', 404)


@app.errorhandler(500)
def handle_500(e):
    """Handle 500 errors and return JSON for API routes."""
    import traceback
    try:
        error_traceback = traceback.format_exc()
        if e:
            error_type = type(e).__name__ if hasattr(e, '__class__') else 'UnknownError'
            error_message = str(e) if hasattr(e, '__str__') else 'Internal server error'
        else:
            error_type = 'InternalServerError'
            error_message = 'An internal server error occurred'
        
        print(f"\n{'='*70}")
        print(f"=== 500 ERROR ===")
        if hasattr(request, 'url'):
            print(f"Requested URL: {request.url}")
            print(f"Request path: {request.path}")
            print(f"Request method: {request.method}")
        print(f"Error: {error_type}: {error_message}")
        print(f"Traceback:\n{error_traceback}")
        print(f"{'='*70}\n")
        
        # Return JSON for API routes, HTML for others
        if hasattr(request, 'path') and request.path.startswith('/api/'):
            return jsonify({
                'success': False,
                'error': f'{error_type}: {error_message}',
                'traceback': error_traceback,
                'path': request.path if hasattr(request, 'path') else 'unknown'
            }), 500
    except Exception as handler_error:
        # If error handler itself fails, try to return basic JSON
        print(f"Error in 500 handler: {handler_error}")
        try:
            if hasattr(request, 'path') and request.path.startswith('/api/'):
                return jsonify({
                    'success': False,
                    'error': 'Internal server error (error handler failed)',
                    'path': request.path if hasattr(request, 'path') else 'unknown'
                }), 500
        except:
            pass
    
    # Return default Flask 500 for non-API routes
    return e if e else ('Internal Server Error', 500)


@app.route('/api/batch-process/<step_type>', methods=['POST'])
def batch_process(step_type):
    """Batch process all files for a given step type."""
    try:
        import subprocess
        import sys
        
        # Get request data (method, API key, and step options)
        request_data = request.get_json() or {}
        method = request_data.get('method')
        api_key = request_data.get('api_key')
        step_options = request_data.get('options', {})
        
        # Get pipeline configuration to find input/output paths
        pipeline_config = get_pipeline_config()
        if not pipeline_config or not pipeline_config.get('steps'):
            return jsonify({'success': False, 'error': 'Pipeline configuration not found'}), 400
        
        # Find the step configuration
        step_config = None
        for step in pipeline_config['steps']:
            if step.get('type') == step_type:
                step_config = step
                break
        
        if not step_config:
            return jsonify({'success': False, 'error': f'Step type {step_type} not found in pipeline configuration'}), 400
        
        input_path = step_config.get('inputPath', '')
        output_path = step_config.get('outputPath', '')

        if method and step_type == 'sentenceCorrect' and method in ('gemma-hf', 'gemma'):
            method = 'gemma-ollama'
        
        # Map step types to scripts
        script_map = {
            'audioTranscribe:story': '1_audio-transcribe.py',
            'audioTranscribe:recall': '1_audio-transcribe.py',
            'eventSegment': '2_story-event-segment.py',
            'sentenceCorrect': '3_spell-grammar-correct.py',
            'textParsing': '4_parse-texts.py',
            'textMatching': '5_recall-rater.py',
            'causalRating': '6_causal-rater.py'
        }
        
        script_name = script_map.get(step_type)
        if not script_name:
            return jsonify({'success': False, 'error': f'Batch processing not available for step type: {step_type}'}), 400
        
        script_path = SCRIPTS_DIR / script_name
        if not script_path.exists():
            return jsonify({'success': False, 'error': f'Script not found: {script_name}'}), 404
        
        print(f"Starting batch processing for {step_type} using {script_name}")
        print(f"  Input: {input_path}")
        for field, env_var, resolved in _resolve_step_extra_input_dirs(step_config):
            print(f"  {field}: {step_config.get(field, '')} -> {env_var}={resolved}")
        print(f"  Output: {output_path}")
        if method:
            print(f"  Method: {method}")
        
        # Build command
        cmd = [sys.executable, str(script_path)]
        
        # Set environment variables for scripts to use
        import os
        env = os.environ.copy()
        _apply_step_path_env(env, step_config)
        env['BATCH_STEP_TYPE'] = step_type
        
        # Add method argument if provided
        if method and step_type == 'eventSegment':
            cmd.extend(['--method', method])
            if method == 'api':
                api_model = request_data.get('model')
                prompt_version = request_data.get('prompt_version')
                if api_model:
                    cmd.extend(['--model', api_model])
                if prompt_version:
                    cmd.extend(['--prompt-version', prompt_version])
                if api_key:
                    provider = 'anthropic'
                    if api_model:
                        if api_model in _EVENT_SEGMENT_OLLAMA_MODEL_KEYS:
                            provider = 'ollama'
                        else:
                            provider = provider_for_model(api_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    elif provider == 'ollama':
                        pass
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
        elif method and step_type == 'textMatching':
            if method == 'test-mode':
                env['TEST_MODE'] = '1'
            elif method == 'api':
                env.pop('TEST_MODE', None)
                recall_model = request_data.get('model')
                recall_prompt = request_data.get('prompt_version')
                if recall_model:
                    env['RECALL_RATING_MODEL'] = recall_model
                    print(f"  RECALL_RATING_MODEL={recall_model}")
                if recall_prompt:
                    # script 5 loads scripts/prompt/<name>.txt via RECALL_RATING_PROMPT (no extension)
                    env['RECALL_RATING_PROMPT'] = (
                        recall_prompt[:-4] if recall_prompt.endswith('.txt') else recall_prompt
                    )
                    print(f"  RECALL_RATING_PROMPT={env['RECALL_RATING_PROMPT']}")
                # Route the supplied key to the provider matching the selected model.
                if api_key:
                    provider = provider_for_model(recall_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
                    print(f"DEBUG: API key set for {provider}")
            elif method == 'gemma-ollama':
                env.pop('TEST_MODE', None)
                env['RECALL_RATING_BACKEND'] = 'ollama'
            elif method == 'rmatch':
                env.pop('TEST_MODE', None)
                env['RECALL_RATING_BACKEND'] = 'rmatch'
                opts = step_options or {}
                rmatch_model = opts.get('rmatch_model') or opts.get('model')
                if rmatch_model:
                    env['RMATCH_MODEL_NAME'] = str(rmatch_model)
                    print(f"  RMATCH_MODEL_NAME={rmatch_model}")
                for src_key, env_key in (
                    ('rmatch_quantization', 'RMATCH_QUANTIZATION'),
                    ('rmatch_window_size', 'RMATCH_WINDOW_SIZE'),
                    ('rmatch_prompt', 'RMATCH_PROMPT'),
                    ('rmatch_batch_size', 'RMATCH_BATCH_SIZE'),
                    ('rmatch_force_cpu', 'RMATCH_FORCE_CPU'),
                ):
                    v = opts.get(src_key)
                    if v not in (None, ''):
                        env[env_key] = str(v)
        elif method and step_type == 'sentenceCorrect':
            env['SPELL_GRAM_METHOD'] = method
        elif method and step_type == 'textParsing':
            if method == 'gemma-ollama':
                env['RECALL_PARSE_METHOD'] = 'ollama'
        elif step_type == 'causalRating':
            if method:
                cmd.extend(['--method', method])
            if method == 'api':
                causal_model = request_data.get('model')
                causal_prompt = request_data.get('prompt_version')
                if causal_model:
                    env['CAUSAL_RATING_MODEL'] = causal_model
                if causal_prompt:
                    env['CAUSAL_RATING_PROMPT'] = causal_prompt
                if api_key:
                    provider = 'anthropic'
                    if causal_model:
                        provider = provider_for_model(causal_model)
                    if provider == 'openai':
                        env['OPENAI_API_KEY'] = api_key
                    else:
                        env['ANTHROPIC_API_KEY'] = api_key
        
        if api_key and step_type not in ('eventSegment',):
            if step_type == 'textMatching' and method == 'gemma-ollama':
                pass
            else:
                env['ANTHROPIC_API_KEY'] = api_key

        # Forward input-variant selections from the launcher dropdown(s).
        input_variant = request_data.get('input_variant')
        if input_variant is not None:
            env['BATCH_INPUT_VARIANT'] = str(input_variant)
            print(f"  BATCH_INPUT_VARIANT={env['BATCH_INPUT_VARIANT']!r}")
        story_events_variant = request_data.get('story_events_variant')
        if story_events_variant is not None:
            env['BATCH_STORY_EVENTS_VARIANT'] = str(story_events_variant)
            print(f"  BATCH_STORY_EVENTS_VARIANT={env['BATCH_STORY_EVENTS_VARIANT']!r}")

        # Forward step options as environment variables
        if step_options:
            for opt_key, opt_val in step_options.items():
                if opt_val == '' or opt_val is None:
                    continue
                env_key = opt_key.upper()
                if isinstance(opt_val, bool):
                    env[env_key] = '1' if opt_val else '0'
                else:
                    env[env_key] = str(opt_val)
                print(f"  Option: {env_key}={env[env_key]}")

        from narraters.runtime_install import prepare_web_step

        prep_err = prepare_web_step(
            step_type=step_type,
            method=method,
            request_data=request_data,
            step_options=step_options,
        )
        if prep_err:
            return jsonify({"success": False, "error": prep_err}), 400

        gemma_report = _gemma4_preflight_report(step_type, method, step_options)
        if gemma_report is not None:
            if not gemma_report.get("ok"):
                return jsonify({
                    "success": False,
                    "error": "Gemma-4 environment check failed: "
                    + "; ".join(gemma_report.get("errors") or []),
                    "gemma_check": gemma_report,
                }), 400
            print(f"DEBUG: Gemma-4 preflight OK: {gemma_report.get('details')}")
            for w in gemma_report.get("warnings") or []:
                print(f"DEBUG: Gemma-4 preflight warning: {w}")

        # Run the script
        result = subprocess.run(
            cmd,
            cwd=str(WORKSPACE_ROOT),
            env=env,
            capture_output=True,
            text=True,
            timeout=3600  # 1 hour timeout
        )
        
        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': f'Batch processing completed for {step_type}',
                'output': result.stdout
            })
        else:
            display_error = _format_subprocess_step_error(result, prefix='Batch processing failed')
            return jsonify({
                'success': False,
                'error': display_error,
                'output': result.stdout,
                'full_stderr': result.stderr,
                'full_stdout': result.stdout,
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'Batch processing timed out'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pipeline/check-gemma-environment', methods=['GET', 'POST'])
def check_gemma_environment_api():
    """Return local Ollama Gemma-4 E4B readiness (default tag gemma4:e4b)."""
    try:
        from helpers.gemma_environment import check_ollama_gemma_e4b_environment

        if request.method == 'POST':
            body = request.get_json(silent=True) or {}
            tag = body.get('model_tag') or body.get('ollama_model') or body.get('model')
        else:
            tag = request.args.get('model_tag') or request.args.get('ollama_model') or request.args.get('model')
        report = check_ollama_gemma_e4b_environment(model_tag=tag)
        return jsonify({'success': True, **report})
    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def get_pipeline_config():
    """Get pipeline configuration from file."""
    pipeline_file = _pipeline_config_path()
    if not pipeline_file.exists():
        # Backwards compatibility: config may still live beside the package.
        pipeline_file = _LEGACY_PIPELINE_FILE
    if not pipeline_file.exists():
        return None
    
    try:
        with open(pipeline_file, 'r', encoding='utf-8') as f:
            return normalize_pipeline_config(json.load(f))
    except Exception as e:
        print(f"Error loading pipeline config: {e}")
        return None


USER_DATA_DIR = WORKSPACE_ROOT / 'user_data'


def _userlog_fmt_ms(ms):
    try:
        ms = int(ms)
    except (TypeError, ValueError):
        return ''
    s, msr = divmod(ms, 1000)
    m, s = divmod(s, 60)
    return f"{m:02d}:{s:02d}.{msr:03d}"


def _human_ms(ms):
    try:
        ms = int(ms)
    except (TypeError, ValueError):
        return None
    s, msr = divmod(ms, 1000)
    m, s = divmod(s, 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}.{msr // 100}s"


def _userlog_describe_target(t):
    if not t:
        return ''
    bits = []
    if t.get('text'):
        bits.append(f"“{t['text']}”")
    if t.get('data-rating'):
        bits.append(f"rating={t['data-rating']}")
    sv = t.get('data-segment-index')
    if sv not in (None, ''):
        bits.append(f"seg#{sv}")
    if t.get('data-story-event-num'):
        bits.append(f"event#{t['data-story-event-num']}")
    if t.get('placeholder'):
        bits.append(f"[{t['placeholder']}]")
    if t.get('id'):
        bits.append(f"#{t['id']}")
    elif t.get('cls'):
        bits.append("." + str(t['cls']).split()[0])
    panel = f" @{t['panel']}" if t.get('panel') else ''
    return f"<{t.get('tag','?')}{panel}> " + " ".join(bits)


def _render_userlog_text(data, events):
    """Render the operation log as a plain, human-readable text report."""
    L = []
    L.append("narRaters — user operation log")
    L.append("(read-only record of one rater's operations for one exported step)")
    L.append("=" * 78)
    L.append(f" subject  : {data.get('subject_id')}")
    L.append(f" step     : {data.get('step')}")
    L.append(f" rater    : {session.get('username', 'human')}")
    L.append(f" export   : {data.get('export_path')}")
    L.append(f" started  : {data.get('started_at')}")
    L.append(f" ended    : {data.get('ended_at')}")
    L.append(f" DURATION : {_human_ms(data.get('duration_ms'))}  ({data.get('duration_ms')} ms)")
    L.append(f" ops      : {len(events)}")
    L.append("=" * 78)
    L.append(f"{'time':>9}  {'op':<7} detail")
    L.append("-" * 78)
    for ev in events:
        t = _userlog_fmt_ms(ev.get('t', 0))
        op = ev.get('type', '?')
        detail = _userlog_describe_target(ev.get('target', {}))
        if 'value' in ev:
            v = ev['value']
            detail += "  -> " + ("checked" if v is True else "unchecked" if v is False else json.dumps(v, ensure_ascii=False))
        if ev.get('key'):
            detail += f"  [key={ev['key']}]"
        L.append(f"{t:>9}  {op:<7} {detail}")
    L.append("-" * 78)
    return "\n".join(L) + "\n"


@app.route('/api/userlog/save', methods=['POST'])
def save_user_log():
    """Persist a per-step user-operation log as a plain, READ-ONLY text file under
    ``narRaters/user_data/`` (a folder separate from output/). The file records the
    rater's operations (clicks / typed text / numbers / checkbox toggles) with the
    time from the start of editing, the interface (tab) and panel, and the total
    duration — directly openable, no decryption needed.
    """
    try:
        data = request.get_json(silent=True) or {}
        events = data.get('events') or []
        if not events:
            return jsonify({'success': True, 'skipped': 'no operations recorded'})

        # Name the log after the exported file, but store it in user_data/ (not output/).
        export_path = str(data.get('export_path') or '').strip()
        if export_path:
            stem = Path(export_path).stem or 'export'
        else:
            stem = f"{data.get('subject_id', 'item')}_{data.get('step', 'step')}"
        try:
            USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        out_path = USER_DATA_DIR / f"{stem}_userlog.txt"

        text = _render_userlog_text(data, events)
        # Write then make the file read-only so it can be opened but not edited.
        try:
            os.chmod(out_path, 0o644)  # ensure writable if it already exists read-only
        except OSError:
            pass
        out_path.write_text(text, encoding='utf-8')
        try:
            os.chmod(out_path, 0o444)  # read-only
        except OSError:
            pass
        print(f"Wrote user log: {out_path} ({len(events)} ops, {_human_ms(data.get('duration_ms'))})")
        return jsonify({'success': True, 'path': str(out_path), 'n_events': len(events)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        # Telemetry must never block the user's export — report but don't 500 hard.
        return jsonify({'success': False, 'error': str(e)}), 200


if __name__ == '__main__':
    import os as _os
    _debug = _os.environ.get('FLASK_DEBUG', '0') == '1'
    if _debug:
        print("narRater - Web Viewer")
        print(f"Workspace root: {WORKSPACE_ROOT}")
        print(f"Package root: {PACKAGE_ROOT}")
    # Default to loopback. The web UI has open file-write and subprocess-spawn
    # endpoints; binding to 0.0.0.0 would expose them to the local network.
    # Override with NARRATERS_HOST=0.0.0.0 if LAN access is intentionally needed.
    _host = _os.environ.get('NARRATERS_HOST', '127.0.0.1')
    _port = int(_os.environ.get('NARRATERS_PORT', '5000'))
    _browse = '127.0.0.1' if _host in ('127.0.0.1', 'localhost', '::1') else _host
    print(f"narRaters ready — open http://{_browse}:{_port}/pipeline-config")
    app.run(debug=_debug, host=_host, port=_port, use_reloader=False)

